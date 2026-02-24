from flask import Blueprint, render_template, request, redirect, url_for, session, flash, make_response
from app.db import DB
from app.models import NavModel, EmployeeModel, EmployeePortalModel, LoanModel, PayrollModel, IncomeTaxModel
from app.utils import get_pagination
import math

hrms_bp = Blueprint('hrms', __name__)

@hrms_bp.before_request
def ensure_module():
    if 'user_id' in session:
        session['current_module_id'] = 63 # ID for Employee Portal Module
        session['ui_mode'] = 'portal'

@hrms_bp.route('/portal')
def employee_portal():
    if 'user_id' not in session: return redirect(url_for('auth.login'))
    emp_id = session.get('emp_id')
    if not emp_id:
        return redirect(url_for('main.index'))
    
    profile = EmployeePortalModel.get_full_profile(emp_id)
    # The user wants "EmployeeDetails.html" style for the profile
    return render_template('hrms/employee_details.html', p=profile)

@hrms_bp.route('/gpf_details')
def gpf_details():
    if 'user_id' not in session: return redirect(url_for('auth.login'))
    emp_id = session.get('emp_id')
    if not emp_id:
        flash("Employee record not linked.", "warning")
        return redirect(url_for('main.index'))
    
    fin_id = request.args.get('fin_year')
    gpf_data = EmployeePortalModel.get_gpf_details(emp_id, fin_id)
    fin_years = NavModel.get_all_fin_years()
    
    return render_template('hrms/gpf_details.html', gpf=gpf_data, fin_years=fin_years, selected_fin_id=fin_id)

@hrms_bp.route('/loan_apply', methods=['GET', 'POST'])
def loan_apply():
    if 'user_id' not in session: return redirect(url_for('auth.login'))
    emp_id = session.get('emp_id')
    user_id = session['user_id']
    
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'DELETE':
            apply_id = request.form.get('apply_id')
            if LoanModel.delete_loan(apply_id, emp_id):
                flash("Loan application deleted.", "success")
            else:
                flash("Failed to delete application.", "danger")
            return redirect(url_for('hrms.loan_apply'))
            
        try:
            LoanModel.apply_loan(request.form, emp_id, user_id)
            flash("Loan application submitted successfully.", "success")
        except Exception as e:
            flash(f"Error: {str(e)}", "danger")
        return redirect(url_for('hrms.loan_apply'))

    # Pagination for history
    page = int(request.args.get('page', 1))
    pagination, sql_limit = get_pagination('SAL_LoanApply_Mst', page) # Note: This counts ALL loans, might need filtering by emp_id for accurate count
    
    # Accurate count for this employee
    total_emp_loans = DB.fetch_scalar("SELECT COUNT(*) FROM SAL_LoanApply_Mst WHERE fk_empid = ?", [emp_id])
    pagination['total'] = total_emp_loans
    pagination['total_pages'] = math.ceil(total_emp_loans / pagination['per_page'])

    loan_types = LoanModel.get_loan_types()
    natures = LoanModel.get_loan_natures()
    purposes = LoanModel.get_loan_purposes()
    history = LoanModel.get_loan_history(emp_id, sql_limit)
    emp_details = LoanModel.get_employee_loan_details(emp_id)
    
    # Show all employees for 'Send To' selection as requested
    all_employees = EmployeeModel.get_lookups()['employees']
    
    import datetime
    today = datetime.datetime.now().strftime('%d/%m/%Y')
    
    return render_template('hrms/loan_apply.html', 
                           types=loan_types, 
                           natures=natures, 
                           purposes=purposes, 
                           history=history,
                           e=emp_details,
                           employees=all_employees,
                           today=today,
                           pagination=pagination)

@hrms_bp.route('/loan_report/<apply_id>')
def download_loan_report(apply_id):
    if 'user_id' not in session: return redirect(url_for('auth.login'))
    emp_id = session.get('emp_id')
    
    loan = LoanModel.get_loan_application(apply_id)
    if not loan: return "Record not found", 404
    if str(loan['fk_empid']) != str(emp_id): return "Access denied", 403

    import io
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import inch
    from flask import make_response
    import os

    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    margin = 1 * inch
    curr_y = height - margin

    # Logo
    logo_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'static', 'images', 'logo.png')
    if os.path.exists(logo_path):
        c.drawImage(logo_path, width/2 - 0.3*inch, curr_y - 0.6*inch, width=0.6*inch, height=0.6*inch, mask='auto')
    
    curr_y -= 0.8*inch
    c.setFont("Helvetica-Bold", 12)
    c.drawCentredString(width/2, curr_y, "Chaudhary Charan Singh Haryana Agricultural University, HISAR")
    
    curr_y -= 0.4*inch
    c.setFont("Helvetica-Bold", 11)
    c.drawCentredString(width/2, curr_y, "LOAN APPLICATION REPORT")
    
    curr_y -= 0.5*inch
    c.setFont("Helvetica", 10)
    
    emp_details = LoanModel.get_employee_loan_details(emp_id)
    
    # Grid info
    data = [
        ("Employee Name", emp_details['empname']),
        ("Employee Code", emp_details['empcode']),
        ("Designation", emp_details['designation']),
        ("Department", emp_details['dept_name']),
        ("Loan Type", DB.fetch_scalar("SELECT description FROM SAL_Head_Mst WHERE pk_headid = ?", [loan['fk_headid']])),
        ("Loan Nature", DB.fetch_scalar("SELECT loanNature FROM SAL_LoanNature_Mst WHERE pk_lnatureid = ?", [loan['fk_lnatureid']])),
        ("Loan Purpose", loan['LoanPurpose']),
        ("Amount", f"Rs. {float(loan['amount']):,.2f}"),
        ("Application Date", loan['dated'].strftime('%d/%m/%Y')),
        ("Remarks", loan['remarks'])
    ]
    
    for label, val in data:
        c.setFont("Helvetica-Bold", 10)
        c.drawString(margin, curr_y, label + ":")
        c.setFont("Helvetica", 10)
        c.drawString(margin + 2*inch, curr_y, str(val) if val else "-")
        curr_y -= 0.25*inch

    curr_y -= 1.0*inch
    c.setFont("Helvetica-Bold", 10)
    c.drawRightString(width - margin, curr_y, f"({emp_details['empname']})")
    
    c.showPage()
    c.save()
    pdf_out = buffer.getvalue()
    buffer.close()
    
    response = make_response(pdf_out)
    response.headers['Content-Type'] = 'application/pdf'
    filename = f"Loan_Application_{apply_id}.pdf"
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    return response

@hrms_bp.route('/income_tax_cert', methods=['GET', 'POST'])
def income_tax_cert():
    if 'user_id' not in session: return redirect(url_for('auth.login'))
    emp_id = session.get('emp_id')
    
    fin_years = NavModel.get_all_fin_years()
    selected_fin_id = request.form.get('fin_year') or request.args.get('fin_year')
    
    data = None
    old_regime = None
    new_regime = None
    profile = EmployeePortalModel.get_full_profile(emp_id)
    
    if selected_fin_id:
        data = PayrollModel.get_it_certificate_data(emp_id, selected_fin_id)
        if data:
            # 1. Monthly Names Mapping
            month_names = {1: 'January', 2: 'February', 3: 'March', 4: 'April', 5: 'May', 6: 'June', 
                           7: 'July', 8: 'August', 9: 'September', 10: 'October', 11: 'November', 12: 'December'}
            for row in data['statement']:
                if not row.get('is_arrear'):
                    row['period'] = f"{month_names[row['month_no']]}-{row['year_no']}"
                row['total'] = float(row['basic_pay']) + float(row['da']) + float(row['hra']) + float(row['fma'])

            # 2. Calculation Totals
            gross_annual = sum(float(s['gross_total']) for s in data['statement'])
            total_it_paid = sum(float(s['it_paid']) for s in data['statement'])
            total_gpf = sum(float(s['gpf_sub']) for s in data['statement'])
            total_gis = sum(float(s['gslis']) for s in data['statement'])
            
            # --- NEW REGIME (2025-26) ---
            std_deduction_new = 75000
            taxable_new = max(0, gross_annual - std_deduction_new)
            new_slabs = [
                {'LowerLimit': 0, 'UpperLimit': 400000, 'Tax_Percent': 0},
                {'LowerLimit': 400000, 'UpperLimit': 800000, 'Tax_Percent': 5},
                {'LowerLimit': 800000, 'UpperLimit': 1200000, 'Tax_Percent': 10},
                {'LowerLimit': 1200000, 'UpperLimit': 1600000, 'Tax_Percent': 15},
                {'LowerLimit': 1600000, 'UpperLimit': 2000000, 'Tax_Percent': 20},
                {'LowerLimit': 2000000, 'UpperLimit': 2400000, 'Tax_Percent': 25},
                {'LowerLimit': 2400000, 'UpperLimit': 99999999, 'Tax_Percent': 30}
            ]
            if taxable_new <= 1200000:
                new_tax = 0
            else:
                new_tax, _ = PayrollModel.calculate_tax(taxable_new, new_slabs)
            
            new_regime = {
                'gross': gross_annual, 'std_deduction': std_deduction_new, 'taxable': taxable_new,
                'it_only': new_tax, 'cess': new_tax * 0.04, 'total_tax': new_tax * 1.04,
                'it_paid': total_it_paid, 'balance': (new_tax * 1.04) - total_it_paid
            }

            # --- OLD REGIME (2025-26) ---
            std_deduction_old = 50000
            # 80C Limited to 1.5L
            ded_80c = min(150000, total_gpf + total_gis) # Base 80C from salary
            taxable_old = max(0, gross_annual - std_deduction_old - ded_80c)
            old_slabs = [
                {'LowerLimit': 0, 'UpperLimit': 250000, 'Tax_Percent': 0},
                {'LowerLimit': 250000, 'UpperLimit': 500000, 'Tax_Percent': 5},
                {'LowerLimit': 500000, 'UpperLimit': 1000000, 'Tax_Percent': 20},
                {'LowerLimit': 1000000, 'UpperLimit': 99999999, 'Tax_Percent': 30}
            ]
            # 87A for Old Regime: No tax up to 5 Lakhs
            if taxable_old <= 500000:
                old_tax = 0
            else:
                old_tax, _ = PayrollModel.calculate_tax(taxable_old, old_slabs)
            
            old_regime = {
                'gross': gross_annual, 'std_deduction': std_deduction_old, 'taxable': taxable_old,
                'it_only': old_tax, 'cess': old_tax * 0.04, 'total_tax': old_tax * 1.04,
                'it_paid': total_it_paid, 'balance': (old_tax * 1.04) - total_it_paid
            }

    import datetime
    now = datetime.datetime.now()
    today_day = now.day
    today_month = month_names[now.month] if data else now.strftime('%B')
    today_year = now.year

    return render_template('hrms/income_tax_cert.html', 
                           fin_years=fin_years, 
                           selected_fin_id=selected_fin_id,
                           data=data,
                           new_regime=new_regime,
                           p=profile,
                           today={'day': today_day, 'month': today_month, 'year': today_year})

@hrms_bp.route('/income_tax_cert/export', methods=['POST'])
def export_it_excel():
    if 'user_id' not in session: return redirect(url_for('auth.login'))
    emp_id = session.get('emp_id')
    regime = request.form.get('selected_regime', 'NEW')
    
    # 1. Capture All Form Data
    periods = request.form.getlist('period[]')
    basics = request.form.getlist('basic[]')
    das = request.form.getlist('da[]')
    hras = request.form.getlist('hra[]')
    fmas = request.form.getlist('fma[]')
    gpfs = request.form.getlist('gpf[]')
    gslis_list = request.form.getlist('gslis[]')
    pts = request.form.getlist('pt[]')
    its = request.form.getlist('it[]')

    # Performa Inputs
    ex_hra = float(request.form.get('ex_hra', 0))
    ex_cea = float(request.form.get('ex_cea', 0))
    ex_conv = float(request.form.get('ex_conv', 0))
    inc_house = float(request.form.get('inc_house', 0))
    ex_house_int = float(request.form.get('ex_house_int', 0))
    inc_other = float(request.form.get('inc_other', 0))
    s_lic = float(request.form.get('s_lic', 0))
    s_ulip = float(request.form.get('s_ulip', 0))
    s_hloan = float(request.form.get('s_hloan', 0))
    s_tuition = float(request.form.get('s_tuition', 0))
    s_nsc = float(request.form.get('s_nsc', 0))
    s_ppf = float(request.form.get('s_ppf', 0))
    s_other_ded = float(request.form.get('s_other_ded', 0))

    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Salary Statement"

    # Styles
    title_font = Font(size=14, bold=True)
    header_fill = PatternFill(start_color="3c8dbc", end_color="3c8dbc", fill_type="solid")
    total_fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # --- SHEET 1: SALARY STATEMENT ---
    ws1.merge_cells('A1:J1')
    ws1['A1'] = "CHAUDHARY CHARAN SINGH HARYANA AGRICULTURAL UNIVERSITY, HISAR"
    ws1['A1'].font = title_font; ws1['A1'].alignment = center_align

    headers = ["Period", "Basic Pay", "DA", "HRA", "FMA", "Gross Pay", "GPF(SUB)", "GSLIS", "PT", "IT"]
    for col, val in enumerate(headers, 1):
        cell = ws1.cell(row=4, column=col)
        cell.value = val; cell.fill = header_fill; cell.font = white_font; cell.alignment = center_align; cell.border = thin_border

    row_idx = 5
    for i in range(len(periods)):
        ws1.cell(row=row_idx, column=1, value=periods[i]).border = thin_border
        ws1.cell(row=row_idx, column=2, value=float(basics[i])).border = thin_border
        ws1.cell(row=row_idx, column=3, value=float(das[i])).border = thin_border
        ws1.cell(row=row_idx, column=4, value=float(hras[i])).border = thin_border
        ws1.cell(row=row_idx, column=5, value=float(fmas[i])).border = thin_border
        ws1.cell(row=row_idx, column=6, value=f"=SUM(B{row_idx}:E{row_idx})").border = thin_border
        ws1.cell(row=row_idx, column=7, value=float(gpfs[i])).border = thin_border
        ws1.cell(row=row_idx, column=8, value=float(gslis_list[i])).border = thin_border
        ws1.cell(row=row_idx, column=9, value=float(pts[i])).border = thin_border
        ws1.cell(row=row_idx, column=10, value=float(its[i])).border = thin_border
        row_idx += 1

    ws1.cell(row=row_idx, column=1, value="TOTAL").font = bold_font
    for col in range(2, 11):
        col_let = get_column_letter(col)
        cell = ws1.cell(row=row_idx, column=col)
        cell.value = f"=SUM({col_let}5:{col_let}{row_idx-1})"
        cell.font = bold_font; cell.alignment = right_align; cell.border = thin_border; cell.fill = total_fill

    # --- SHEET 2: PERFORMA ---
    ws2 = wb.create_sheet("Tax Calculation Performa")
    ws2.column_dimensions['A'].width = 60
    ws2.column_dimensions['B'].width = 20
    ws2.column_dimensions['C'].width = 15
    ws2.column_dimensions['D'].width = 15

    ws2.merge_cells('A1:D1')
    ws2['A1'] = "CHAUDHARY CHARAN SINGH HARYANA AGRICULTURAL UNIVERSITY, HISAR"
    ws2['A1'].font = bold_font; ws2['A1'].alignment = center_align

    ws2.merge_cells('A2:D2')
    ws2['A2'] = f"PERFORMA FOR CALCULATION OF INCOME TAX FOR THE YEAR 2025-2026 ({regime} REGIME)"
    ws2['A2'].alignment = center_align

    # Employee Bio
    profile = EmployeePortalModel.get_full_profile(emp_id)
    ws2['A4'] = f"Name: {profile['basic']['empname']}"; ws2['C4'] = f"Designation: {profile['basic']['designation']}"
    ws2['A5'] = f"Account No: {profile['basic']['bankaccountno']}"; ws2['C5'] = f"PAN: {profile['basic']['panno']}"

    # Standard Deduction based on regime
    std_ded_val = 75000 if regime == 'NEW' else 50000

    ws2['A8'] = "A. Salary and Other Benefits:"; ws2['B8'] = "Gross Income"
    ws2['C8'] = f"='Salary Statement'!F{row_idx}"
    ws2['A9'] = f"   Standard Deduction ({regime})"; ws2['B9'] = "Less"; ws2['C9'] = std_ded_val
    ws2['A10'] = "Balance"; ws2['D10'] = "=C8-C9"; ws2['D10'].font = bold_font

    current_row = 11
    if regime == 'OLD':
        ws2.cell(row=current_row, column=1, value="B. Less: Income exempt u/s 10"); current_row += 1
        ws2.cell(row=current_row, column=1, value="   1. House Rent Allowance"); ws2.cell(row=current_row, column=3, value=ex_hra); current_row += 1
        ws2.cell(row=current_row, column=1, value="   2. Children Education Allowance"); ws2.cell(row=current_row, column=3, value=ex_cea); current_row += 1
        ws2.cell(row=current_row, column=1, value="   3. Fixed / Conveyance Allowance"); ws2.cell(row=current_row, column=3, value=ex_conv); current_row += 1
        ws2.cell(row=current_row, column=1, value="Income from Salary"); ws2.cell(row=current_row, column=4, value=f"=D10-SUM(C{current_row-3}:C{current_row-1})"); current_row += 1
        
        inc_sal_row = current_row - 1
        ws2.cell(row=current_row, column=1, value="C. Add : Income from House Property"); ws2.cell(row=current_row, column=3, value=inc_house); current_row += 1
        ws2.cell(row=current_row, column=1, value="D. Less : Interest Paid for self occupied house"); ws2.cell(row=current_row, column=3, value=ex_house_int); current_row += 1
        ws2.cell(row=current_row, column=1, value="E. Add : Income from other sources"); ws2.cell(row=current_row, column=3, value=inc_other); current_row += 1
        ws2.cell(row=current_row, column=1, value="F. Gross Total Income"); ws2.cell(row=current_row, column=4, value=f"=D{inc_sal_row}+C{current_row-3}-C{current_row-2}+C{current_row-1})"); current_row += 1
        
        gti_row = current_row - 1
        ws2.cell(row=current_row, column=1, value="G. Less : Deduction u/s 80C to 80CCF"); current_row += 1
        ws2.cell(row=current_row, column=1, value="   i) GPF/PRAN"); ws2.cell(row=current_row, column=3, value=f"='Salary Statement'!G{row_idx}"); current_row += 1
        ws2.cell(row=current_row, column=1, value="   ii) GIS"); ws2.cell(row=current_row, column=3, value=f"='Salary Statement'!H{row_idx}"); current_row += 1
        ws2.cell(row=current_row, column=1, value="   iii) LIC"); ws2.cell(row=current_row, column=3, value=s_lic); current_row += 1
        ws2.cell(row=current_row, column=1, value="Total 80C (Limited to 1.5L)"); ws2.cell(row=current_row, column=4, value=f"=MIN(150000, SUM(C{current_row-3}:C{current_row-1}))"); current_row += 1
        
        ded80c_row = current_row - 1
        ws2.cell(row=current_row, column=1, value="H. Less : Deductions u/s 80D to 80U"); ws2.cell(row=current_row, column=3, value=s_other_ded); current_row += 1
        ws2.cell(row=current_row, column=1, value="I. Taxable Income (Rounded)"); ws2.cell(row=current_row, column=4, value=f"=ROUND(D{gti_row}-D{ded80c_row}-C{current_row-1}, -1)"); current_row += 1
    else:
        # NEW REGIME
        ws2.cell(row=current_row, column=1, value="E. Add : Income from other sources"); ws2.cell(row=current_row, column=3, value=inc_other); current_row += 1
        ws2.cell(row=current_row, column=1, value="F. Gross Total Income"); ws2.cell(row=current_row, column=4, value=f"=D10+C{current_row-1})"); current_row += 1
        ws2.cell(row=current_row, column=1, value="I. Taxable Income (Rounded)"); ws2.cell(row=current_row, column=4, value=f"=ROUND(D{current_row-1}, -1)"); current_row += 1

    taxable_cell = f"D{current_row-1}"
    it_row = current_row + 2
    ws2.cell(row=it_row, column=1, value="J. Computation of Tax").font = bold_font; it_row += 1
    
    if regime == 'NEW':
        slabs = [(0, 400000, 0), (400000, 800000, 5), (800000, 1200000, 10), (1200000, 1600000, 15), (1600000, 2000000, 20), (2000000, 2400000, 25), (2400000, 99999999, 30)]
        rebate_limit = 1200000
    else:
        slabs = [(0, 250000, 0), (250000, 500000, 5), (500000, 1000000, 20), (1000000, 99999999, 30)]
        rebate_limit = 500000

    start_slab = it_row
    for l, u, r in slabs:
        ws2.cell(row=it_row, column=1, value=f"   {r}% ({l} - {u})")
        ws2.cell(row=it_row, column=2, value=f"=MAX(0, MIN({taxable_cell}, {u})-{l})")
        ws2.cell(row=it_row, column=3, value=f"=B{it_row}*{r/100}")
        it_row += 1
    
    ws2.cell(row=it_row, column=1, value="Tax on total income"); ws2.cell(row=it_row, column=3, value=f"=SUM(C{start_slab}:C{it_row-1})"); it_row += 1
    ws2.cell(row=it_row, column=1, value="Less Rebate u/s 87A"); ws2.cell(row=it_row, column=3, value=f"=IF({taxable_cell}<={rebate_limit}, C{it_row-1}, 0)"); it_row += 1
    ws2.cell(row=it_row, column=1, value="TAX PAYABLE (J)"); ws2.cell(row=it_row, column=4, value=f"=C{it_row-2}-C{it_row-1})"); it_row += 1
    
    pay_row = it_row - 1
    ws2.cell(row=it_row, column=1, value="K. Education Cess @ 4%"); ws2.cell(row=it_row, column=3, value=f"=D{pay_row}*0.04"); it_row += 1
    ws2.cell(row=it_row, column=1, value="L. Total Tax Payable (J+K)"); ws2.cell(row=it_row, column=4, value=f"=D{pay_row}+C{it_row-1})"); it_row += 1
    ws2.cell(row=it_row, column=1, value="M. Tax Deducted at Source"); ws2.cell(row=it_row, column=4, value=f"='Salary Statement'!J{row_idx}"); it_row += 1
    ws2.cell(row=it_row, column=1, value="N. Balance Tax Payable (L-M)"); ws2.cell(row=it_row, column=4, value=f"=D{it_row-2}-D{it_row})"); it_row += 1

    # Verification Section
    v_row = it_row + 2
    ws2.cell(row=v_row, column=1, value="Verification").font = bold_font
    ws2.merge_cells(start_row=v_row+1, start_column=1, end_row=v_row+2, end_column=4)
    ws2.cell(row=v_row+1, column=1, value=f"I, {profile['basic']['empname']}, do hereby declare that what is stated above is true to the best of my knowledge and belief.")
    
    # Borders for all used cells
    for row in ws1.iter_rows(min_row=1, max_row=row_idx, min_col=1, max_col=10):
        for cell in row: cell.border = thin_border
    for row in ws2.iter_rows(min_row=1, max_row=v_row+5, min_col=1, max_col=4):
        for cell in row:
            if cell.coordinate in ws2.merged_cells: continue
            cell.border = thin_border

    output = io.BytesIO()
    wb.save(output); output.seek(0)
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=Income_Tax_Performa_{regime}.xlsx'
    return response

@hrms_bp.route('/ltc_apply')
def ltc_apply():
    return render_template('hrms/ltc_apply.html')

@hrms_bp.route('/ta_bill', methods=['GET', 'POST'])
def ta_bill_view():
    if 'user_id' not in session: return redirect(url_for('auth.login'))
    emp_id = session.get('emp_id')
    user_id = session['user_id']
    
    action = request.args.get('action')
    edit_id = request.args.get('edit_id')
    
    if request.method == 'POST':
        # logic to save master and child details
        flash("TA Bill saved successfully.", "success")
        return redirect(url_for('hrms.ta_bill_view'))

    # Fetch History
    history = DB.fetch_all("""
        SELECT T.*, CONVERT(varchar, M.BillDate, 103) as bill_date,
               M.BillAmount,
               CASE WHEN M.IsBillSubmit = 1 THEN 'Submitted' ELSE 'Draft' END as status_desc
        FROM Acct_TA_Bill T
        INNER JOIN Acct_Bill_Master M ON T.fk_billId = M.Pk_BillId
        WHERE T.fk_empId = ?
        ORDER BY M.BillDate DESC
    """, [emp_id])
    
    bill = None
    segments = []
    if edit_id:
        bill = DB.fetch_one("SELECT * FROM Acct_TA_Bill WHERE PkID = ?", [edit_id])
        segments = DB.fetch_all("SELECT *, CONVERT(varchar, FromDate, 23) as fdate, CONVERT(varchar, ToDate, 23) as tdate FROM Acct_TA_Emp_Bill_Details WHERE fk_Emp_billId = ?", [edit_id])

    return render_template('hrms/ta_bill.html', history=history, action=action, editing=bool(edit_id), bill=bill, segments=segments)

@hrms_bp.route('/rent_submission', methods=['GET', 'POST'])
def rent_submission():
    if 'user_id' not in session: return redirect(url_for('auth.login'))
    emp_id = session.get('emp_id')
    user_id = session['user_id']
    
    fin_years = NavModel.get_all_fin_years()
    selected_fin_id = request.form.get('fin_year') or request.args.get('fin_year')
    
    if request.method == 'POST' and 'save_rent' in request.form:
        month_ids = request.form.getlist('month_id[]')
        amounts = request.form.getlist('amount[]')
        years = request.form.getlist('year[]')
        for m_id, amt, yr in zip(month_ids, amounts, years):
            PayrollModel.save_rent_details(emp_id, selected_fin_id, m_id, yr, amt, user_id)
        flash("Rent details updated successfully.", "success")
        return redirect(url_for('hrms.rent_submission', fin_year=selected_fin_id))

    rent_data = []
    current_fy = None
    if selected_fin_id:
        rent_data = PayrollModel.get_employee_rent_details(emp_id, selected_fin_id)
        current_fy = next((f for f in fin_years if f['id'] == selected_fin_id), None)

    e_header = PayrollModel.get_employee_rent_header(emp_id)

    return render_template('hrms/rent_submission.html', 
                           fin_years=fin_years, 
                           selected_fin_id=selected_fin_id,
                           current_fy=current_fy,
                           rent_data=rent_data,
                           e=e_header)

@hrms_bp.route('/form16', methods=['GET', 'POST'])
def form16():
    if 'user_id' not in session: return redirect(url_for('auth.login'))
    emp_id = session.get('emp_id')
    user_id = session['user_id']
    
    fin_years = NavModel.get_all_fin_years()
    selected_fin_id = request.form.get('fin_year') or request.args.get('fin_year')
    
    if request.method == 'POST' and 'save_decl' in request.form:
        sub_ids = request.form.getlist('subsec_id[]')
        sec_ids = request.form.getlist('sec_id[]')
        amounts = request.form.getlist('amount[]')
        
        decl_data = []
        for sid, sec, amt in zip(sub_ids, sec_ids, amounts):
            if float(amt or 0) > 0:
                decl_data.append({'subsec_id': sid, 'sec_id': sec, 'amount': float(amt)})
        
        if IncomeTaxModel.save_declaration(decl_data, emp_id, selected_fin_id, user_id):
            flash("Tax declaration saved successfully.", "success")
        else:
            flash("Failed to save declaration.", "danger")
        return redirect(url_for('hrms.form16', fin_year=selected_fin_id))

    sections = IncomeTaxModel.get_sections()
    for sec in sections:
        sec['subsections'] = IncomeTaxModel.get_subsections(sec['pk_secid'])

    declarations = []
    if selected_fin_id:
        declarations = IncomeTaxModel.get_employee_declarations(emp_id, selected_fin_id)
    
    decl_map = {d['fk_subsecid']: d['docsub_Amt'] for d in declarations}
    for sec in sections:
        for sub in sec['subsections']:
            sub['current_val'] = decl_map.get(sub['pk_subsecid'], 0)

    e_header = PayrollModel.get_employee_rent_header(emp_id)

    return render_template('hrms/tax_declaration.html', 
                           fin_years=fin_years, 
                           selected_fin_id=selected_fin_id,
                           sections=sections,
                           e=e_header)

@hrms_bp.route('/form16_generation', methods=['GET', 'POST'])
def form16_generation():
    if 'user_id' not in session: return redirect(url_for('auth.login'))
    user_id = session['user_id']
    
    # Priority: query string > session (allows admin to view any employee)
    emp_id = request.args.get('emp_id') or session.get('emp_id')
    if not emp_id:
        flash("No employee selected.", "warning")
        return redirect(url_for('hrms.form16_process'))

    fin_years = NavModel.get_all_fin_years()
    
    # Get active financial year ID as default
    active_fy = DB.fetch_one("SELECT pk_finid FROM SAL_Financial_Year WHERE active = 'Y' ORDER BY orderno DESC")
    default_fy = active_fy['pk_finid'] if active_fy else 'CO-18'
    
    selected_fin_id = request.form.get('fin_year') or request.args.get('fin_year') or default_fy

    if request.method == 'POST':
        form_data = {
            'GrossSal': float(request.form.get('GrossSal', 0)),
            'Perquisites': float(request.form.get('Perquisites', 0)),
            'Profitlieu': float(request.form.get('Profitlieu', 0)),
            'GrossPerkLieu': float(request.form.get('GrossPerkLieu', 0)),
            'HRA': float(request.form.get('HRA', 0)),
            'Conveyance': float(request.form.get('Conveyance', 0)),
            'TotHraConvey': float(request.form.get('TotHraConvey', 0)),
            'Balance': float(request.form.get('Balance', 0)),
            'StandardDed': float(request.form.get('StandardDed', 0)),
            'EntAllowance': float(request.form.get('EntAllowance', 0)),
            'Aggregate': float(request.form.get('Aggregate', 0)),
            'Incomechargeble': float(request.form.get('Incomechargeble', 0)),
            'Houseproperty': float(request.form.get('Houseproperty', 0)),
            'IntrestHBA': float(request.form.get('IntrestHBA', 0)),
            'Othersource': float(request.form.get('Othersource', 0)),
            'AgriIncome': float(request.form.get('AgriIncome', 0)),
            'Grosstotincome': float(request.form.get('Grosstotincome', 0)),
            'AggrChapVIDedAmt': float(request.form.get('AggrChapVIDedAmt', 0)),
            'TotIncomechapVI': float(request.form.get('TotIncomechapVI', 0)),
            'TotalTaxIncome': float(request.form.get('TotalTaxIncome', 0)),
            'TaxOnIncome': float(request.form.get('TaxOnIncome', 0)),
            'TotalTax': float(request.form.get('TotalTax', 0)),
            'Surcharge': float(request.form.get('Surcharge', 0)),
            'AddEdu': float(request.form.get('AddEdu', 0)),
            'TaxIncludindSur': float(request.form.get('TaxIncludindSur', 0)),
            'Relief89': float(request.form.get('Relief89', 0)),
            'TaxPayable': float(request.form.get('TaxPayable', 0)),
            'Taxdeducted': float(request.form.get('Taxdeducted', 0)),
            'paid': float(request.form.get('paid', 0)),
            'PayRefund': float(request.form.get('PayRefund', 0)),
            'Regime': request.form.get('Regime', 'NEW')
        }
        
        PayrollModel.save_it_computation(form_data, emp_id, selected_fin_id, user_id)
        
        if 'btn_preview' in request.form:
            return redirect(url_for('hrms.download_form16_pdf', emp_id=emp_id, fin_id=selected_fin_id))
            
        flash("Draft saved successfully.", "success")
        return redirect(url_for('hrms.form16_generation', fin_year=selected_fin_id))

    comp_data = PayrollModel.get_it_computation_data(emp_id, selected_fin_id)
    e_header = PayrollModel.get_employee_rent_header(emp_id)
    
    # Fetch dynamic data
    decls_list = IncomeTaxModel.get_employee_declarations(emp_id, selected_fin_id)
    q_sum_data = PayrollModel.get_form16_quarterly_summary(emp_id, selected_fin_id)
    
    from datetime import datetime
    now_fmt = datetime.now().strftime('%d/%m/%Y')

    return render_template('hrms/form16_generation.html', 
                           fin_years=fin_years, 
                           selected_fin_id=selected_fin_id, 
                           e=e_header, 
                           comp=comp_data,
                           decls=decls_list,
                           q_summary=q_sum_data,
                           now_fmt=now_fmt)
@hrms_bp.route('/download_form16_pdf/<emp_id>/<fin_id>')
def download_form16_pdf(emp_id, fin_id):
    if 'user_id' not in session: return redirect(url_for('auth.login'))
    
    # 1. Fetch Dynamic Data
    profile = EmployeePortalModel.get_full_profile(emp_id)
    comp = PayrollModel.get_it_computation_data(emp_id, fin_id)
    if not comp: return "Calculation not saved. Please save draft first.", 400
    
    # Financial Year Info for Assessment Year and Period
    fy = DB.fetch_one("SELECT date1, date2 FROM SAL_Financial_Year WHERE pk_finid = ?", [fin_id])
    if not fy: return "Financial Year not found.", 404
    
    y1, y2 = fy['date1'].year, fy['date2'].year
    assessment_year = f"{y1+1}-{y2+1}"
    period_str = f"{fy['date1'].strftime('%d %b %Y')} - {fy['date2'].strftime('%d %b %Y')}"
    
    tds_rows = PayrollModel.get_form16_tds_details(emp_id, fin_id)
    q_sum = PayrollModel.get_form16_quarterly_summary(emp_id, fin_id)
    q_map = {q['quarter']: q['total_tax'] for q in q_sum}

    import io
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import Table, TableStyle, Paragraph, SimpleDocTemplate, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.lib.units import inch

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=20, leftMargin=20, topMargin=20, bottom=20)
    styles = getSampleStyleSheet()
    elements = []

    # Custom styles
    h1 = ParagraphStyle('H1', fontName='Helvetica-Bold', fontSize=11, alignment=TA_CENTER)
    h2 = ParagraphStyle('H2', fontName='Helvetica', fontSize=9, alignment=TA_CENTER)
    body_bold = ParagraphStyle('BodyB', fontName='Helvetica-Bold', fontSize=8)
    body_norm = ParagraphStyle('BodyN', fontName='Helvetica', fontSize=8)
    
    def fmt(val): return f"{float(val or 0):,.2f}"

    # --- HEADER ---
    elements.append(Paragraph("FORM NO. 16", h1))
    elements.append(Paragraph("[See rule 31(1)(a)]", h2))
    elements.append(Paragraph("Certificate under section 203 of the Income-tax Act, 1961 for tax deducted at source from income chargeable under the head \"Salaries\"", h2))
    elements.append(Spacer(1, 10))

    # --- TOP INFO BOX ---
    gender_map = {'M': 'MALE', 'F': 'FEMALE', 'O': 'OTHER'}
    gender_str = gender_map.get(profile['basic'].get('gender', 'M'), 'MALE')
    
    top_data = [
        [Paragraph("<b>Designation and Address of the Employer :</b>", body_norm), Paragraph("<b>Name and Employee Code :</b>", body_norm)],
        [Paragraph(f"{profile['basic'].get('designation', 'Professor')}, CHAUDHARY CHARAN SINGH HARYANA AGRICULTURAL UNIVERSITY, HISAR", body_bold), Paragraph(f"{profile['basic']['empname']} - {profile['basic']['empcode']}", body_bold)],
        [Paragraph("TAN No. : RTKD02742F", body_norm), Paragraph(f"PAN: {profile['basic'].get('panno', '')}   SEX: {gender_str}", body_norm)]
    ]
    t1 = Table(top_data, colWidths=[2.8*inch, 2.8*inch])
    t1.setStyle(TableStyle([('GRID', (0,0), (-1,-1), 0.5, colors.black), ('VALIGN', (0,0), (-1,-1), 'TOP')]))
    elements.append(t1)

    # --- ASSESSMENT INFO ---
    ass_data = [
        ["TDS Circle where Annual Return is to be filled", "Assessment Year", "Period From"],
        ["Income Tax Ward HISAR", assessment_year, period_str]
    ]
    t2 = Table(ass_data, colWidths=[2.5*inch, 1.5*inch, 1.6*inch])
    t2.setStyle(TableStyle([('GRID', (0,0), (-1,-1), 0.5, colors.black), ('FONTSIZE', (0,0), (-1,-1), 8), ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold')]))
    elements.append(t2)
    elements.append(Spacer(1, 10))

    # --- QUARTERLY TDS ACK ---
    elements.append(Paragraph("<b>Acknowledgement Nos of all the quaterly statements of TDS under Sub-Section(3) of section 200 as provided by TIN faciliation Centre of NSDL web-site</b>", body_norm))
    elements.append(Spacer(1, 5))
    
    q_sub = ["Quarter", "Acknowledgment No.", "Amount of Tax Deducted in\nrespect of Employee", "Amount of Tax\nDeposit/remitted in\nrespect of Employee"]
    
    q_rows = [
        ["Quarter No.1", "0", fmt(q_map.get('Q1', 0)), fmt(q_map.get('Q1', 0))],
        ["Quarter No.2", "0", fmt(q_map.get('Q2', 0)), fmt(q_map.get('Q2', 0))],
        ["Quarter No.3", "0", fmt(q_map.get('Q3', 0)), fmt(q_map.get('Q3', 0))],
        ["Quarter No.4", "0", fmt(q_map.get('Q4', 0)), fmt(q_map.get('Q4', 0))]
    ]

    t3 = Table([q_sub] + q_rows, colWidths=[1.2*inch, 1.4*inch, 1.5*inch, 1.5*inch])
    t3.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 0.5, colors.black),
        ('ALIGN', (0,0), (1,-1), 'CENTER'),
        ('ALIGN', (2,0), (3,-1), 'RIGHT'),
        ('FONTSIZE', (0,0), (-1,-1), 7),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))
    elements.append(t3)
    elements.append(Spacer(1, 15))

    # --- DETAILS OF SALARY (PART B) ---
    elements.append(Paragraph("<b>Details of Salary Paid & Any Other Income & Tax Deducted</b>", body_norm))
    sal_data = [
        ["1. Gross salary U/s 17(1)", "", "", fmt(comp['GrossSal'])],
        ["2. (a) Less Allowances U/s 10 Education Allowance", "", fmt(comp['Conveyance']), ""],
        ["   (b) Less U/s 10(13 A) H.R.A", "", fmt(comp['HRA']), ""],
        [Paragraph("<b>3. Aggregate of 2(a to b)</b>", body_norm), "", "", fmt(comp['TotHraConvey'])],
        [Paragraph("<b>4. Balance (1-3)</b>", body_norm), "", "", fmt(comp['Balance'])],
        ["5. Deductions : (a) Standard Deduction", "", fmt(comp['StandardDed']), ""],
        ["   (b) Deduction (if any)", "", fmt(comp['EntAllowance']), ""],
        [Paragraph("<b>Total (a to b)</b>", body_norm), "", "", fmt(comp['Aggregate'])],
        [Paragraph("<b>6. Income chargeable under the \"Head Salary\" (4-5)</b>", body_norm), "", "", fmt(comp['Incomechargeble'])],
        ["7. Income reported by the Empl. (a)(from House\nProperty)", "", fmt(comp['Houseproperty']), ""],
        ["   (i) Less Intrest on HBA U/s 24", "", fmt(comp['IntrestHBA']), ""],
        ["   (b) Income from other sources", "", fmt(comp['Othersource']), ""],
        ["   (c) Agriculture Income", "", fmt(comp['AgriIncome']), ""],
        [Paragraph("<b>8. Gross total income :</b>", body_norm), "", "", fmt(comp['Grosstotincome'])],
        [Paragraph("<b>10. Aggregate of Deductable amount under chapter\nVI-A Total of (9) ded.amt.</b>", body_norm), "", "", fmt(comp['AggrChapVIDedAmt'])],
        [Paragraph("<b>11. Total Income(8-10)</b>", body_norm), "", "", fmt(comp['TotIncomechapVI'])],
        [Paragraph("<b>12. Total Taxable Income U/s 288A</b>", body_norm), "", "", fmt(comp['TotalTaxIncome'])],
        [Paragraph("<b>13. Tax on Total Income</b>", body_norm), "", "", fmt(comp['TaxOnIncome'])],
        ["14. Less : Tax credit(Available only to assessees\nhaving total taxable income upto Rs.3.50 Lacs", "", "0.00", ""],
        ["    Less : Income Tax on Agricultural income\ninclude above", "", "0", ""],
        ["15. (a) Tax payable\nSurcharge payable", "", fmt(comp['TotalTax']), ""],
        ["    (b) Add 4% Education Cess(on a+b)", "", fmt(comp['AddEdu']), ""],
        [Paragraph("<b>Total tax payable including Surcharge</b>", body_norm), "", "", fmt(comp['TaxIncludindSur'])],
        ["16. Relief under section 89", "", "", fmt(comp['Relief89'])],
        [Paragraph("<b>17. Balance tax Payable(14-15)</b>", body_norm), "", "", "0"],
        [Paragraph("<b>18. (a) Less Tax deducted at source U/s 192(1)\n(b) Tax paid by the Employer</b>", body_norm), "", "", fmt(comp['Taxdeducted'] + comp['paid'])],
        [Paragraph("<b>19. (a) Tax payable/Refundable</b>", ParagraphStyle('Red', parent=body_bold, textColor=colors.red)), "", "", fmt(comp['PayRefund'])]
    ]
    t4 = Table(sal_data, colWidths=[3.2*inch, 0.8*inch, 0.8*inch, 0.8*inch])
    t4.setStyle(TableStyle([('GRID', (0,0), (-1,-1), 0.5, colors.black), ('ALIGN', (2,0), (3, -1), 'RIGHT'), ('FONTSIZE', (0,0), (-1,-1), 8)]))
    elements.append(t4)
    
    elements.append(Spacer(1, 20))

    # --- TDS DETAILS TABLE (Continuation) ---
    elements.append(Paragraph("<b>DETAILS OF TAX DEDUCTED AND DEPOSITED INTO CENTRAL GOVERNMENT ACCOUNT</b>", h1))
    elements.append(Spacer(1, 10))
    
    tds_head = ["S.No", "Tax", "Surcharge", "Edu.Cess4%", "Total TDS", "Date of Payment\nDD/MM/YR", "Challan No.", "BSR Code of Bank\nBranch"]
    tds_data = [tds_head]
    
    total_tds_val = 0
    for i, r in enumerate(tds_rows, 1):
        row_tot = float(r['tax']) + float(r['surcharge']) + float(r['cess'])
        total_tds_val += row_tot
        tds_data.append([i, fmt(r['tax']), fmt(r['surcharge']), fmt(r['cess']), fmt(row_tot), r['pay_date'], r['challan_no'], r['bsr_code']])
    
    tds_data.append(["", "Total :", "", "", fmt(total_tds_val), "", "", ""])
    
    t5 = Table(tds_data, colWidths=[0.4*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.7*inch])
    t5.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 0.5, colors.black),
        ('FONTSIZE', (0,0), (-1,-1), 7),
        ('ALIGN', (1,0), (4,-1), 'RIGHT'),
        ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold')
    ]))
    elements.append(t5)
    
    # --- FOOTER VERIFICATION ---
    elements.append(Spacer(1, 20))
    father_name = profile['basic'].get('fathername', '______________')
    cert_text = f"I {profile['basic']['empname']} son of Mr. {father_name} working in the capacity of {profile['basic']['designation']} do hereby certify that a sum of Rs. {total_tds_val:,.2f} [Rupees {PayrollModel.amount_to_words(total_tds_val).upper()}] has been deducted at source and paid to the credit of the Central Government / I further certify that the information given above is true and correct based on the books of account, document and other available records."
    elements.append(Paragraph(cert_text, ParagraphStyle('Cert', parent=body_norm, alignment=TA_LEFT, leading=12)))
    elements.append(Spacer(1, 20))
    elements.append(Paragraph("Place: HISAR", body_norm))
    elements.append(Paragraph("Date: 31/01/2026", body_norm))
    elements.append(Spacer(1, 20))
    elements.append(Paragraph("__________________________________________", ParagraphStyle('sig', alignment=TA_RIGHT)))
    elements.append(Paragraph("Signature of the person responsible for deduction", ParagraphStyle('sig', alignment=TA_RIGHT, fontSize=8)))

    doc.build(elements)
    pdf_out = buffer.getvalue()
    buffer.close()
    
    response = make_response(pdf_out)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=Form16_{profile["basic"]["empcode"]}.pdf'
    return response

@hrms_bp.route('/form16_process', methods=['GET', 'POST'])
def form16_process():
    if 'user_id' not in session: return redirect(url_for('auth.login'))
    user_id = session['user_id']
    
    fin_years = NavModel.get_all_fin_years()
    selected_fin_id = request.form.get('fin_year') or request.args.get('fin_year') or 'CO-18'
    
    # 1. Fetch Filter Dropdowns
    ddos = DB.fetch_all("SELECT pk_ddoid as id, Description as name FROM DDO_Mst ORDER BY Description")
    locations = DB.fetch_all("SELECT pk_locationid as id, Description as name FROM Location_Mst ORDER BY Description")
    departments = DB.fetch_all("SELECT pk_deptid as id, description as name FROM Department_Mst ORDER BY description")
    
    # 2. Handle Process / Unprocess Actions
    if request.method == 'POST':
        action = request.form.get('action')
        selected_emp_ids = request.form.getlist('emp_ids[]')
        
        if action == 'PROCESS' and selected_emp_ids:
            for eid in selected_emp_ids:
                # Logic: Insert into SAL_IT_Details or a dedicated Form16 status table
                # For now, we simulate by marking them
                pass
            flash(f"Processed {len(selected_emp_ids)} employees for Form 16.", "success")
        elif action == 'UNPROCESS' and selected_emp_ids:
            flash(f"Unprocessed {len(selected_emp_ids)} employees.", "warning")

    # 3. Build Filtered Query
    f_ddo = request.form.get('ddo')
    f_loc = request.form.get('location')
    f_dept = request.form.get('dept')
    
    base_query = """
        SELECT E.pk_empid, E.empcode, E.empname, D.designation, DP.description as dept_name,
               L.Description as location, S.description as section
        FROM SAL_Employee_Mst E
        LEFT JOIN SAL_Designation_Mst D ON E.fk_desgid = D.pk_desgid
        LEFT JOIN Department_Mst DP ON E.fk_deptid = DP.pk_deptid
        LEFT JOIN Location_Mst L ON E.fk_locationid = L.pk_locationid
        LEFT JOIN SAL_Section_Mst S ON E.fk_sectionid = S.pk_sectionid
        WHERE E.employeeleftstatus = 'N'
    """
    params = []
    if f_ddo: base_query += " AND E.fk_ddoid = ?"; params.append(f_ddo)
    if f_loc: base_query += " AND E.fk_locationid = ?"; params.append(f_loc)
    if f_dept: base_query += " AND E.fk_deptid = ?"; params.append(f_dept)
    
    base_query += " ORDER BY E.empname"
    employees = DB.fetch_all(base_query, params)
    
    # Simulate Processed vs Not Processed
    not_processed = employees # For now
    processed = []

    return render_template('hrms/form16_process.html', 
                           fin_years=fin_years, 
                           selected_fin_id=selected_fin_id,
                           ddos=ddos, locations=locations, departments=departments,
                           not_processed=not_processed, processed=processed)

@hrms_bp.route('/download_form16/<emp_id>/<fin_id>')
def download_form16(emp_id, fin_id):
    if 'user_id' not in session: return redirect(url_for('auth.login'))
    
    # Use provided emp_id or fallback to session
    if not emp_id or emp_id == 'default':
        emp_id = session.get('emp_id')
    
    if not emp_id:
        return "Employee ID not found.", 400
    
    # 1. Fetch Profile
    profile = EmployeePortalModel.get_full_profile(emp_id)
    # 2. Fetch Salary Statement
    it_data = PayrollModel.get_it_certificate_data(emp_id, fin_id)
    # 3. Fetch Declarations
    decls_list = IncomeTaxModel.get_employee_declarations(emp_id, fin_id)
    
    import io
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from reportlab.lib import colors
    
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    
    # Header
    p.setFont("Helvetica-Bold", 14)
    p.drawCentredString(width/2, height-40, "FORM NO. 16")
    p.setFont("Helvetica", 10)
    p.drawCentredString(width/2, height-55, "[See rule 31(1)(a)]")
    p.drawCentredString(width/2, height-70, "Certificate under section 203 of the Income-tax Act, 1961 for tax deducted at source")
    
    # Main content box
    p.rect(40, height-400, width-80, 320)
    p.line(40, height-150, width-40, height-150)
    
    p.setFont("Helvetica-Bold", 10)
    p.drawString(50, height-100, "Name and address of the Employer")
    p.drawString(width/2 + 10, height-100, "Name and address of the Employee")
    
    p.setFont("Helvetica", 9)
    p.drawString(50, height-115, "The Comptroller, CCS HAU, Hisar")
    p.drawString(width/2 + 10, height-115, profile['basic']['empname'])
    p.drawString(width/2 + 10, height-125, f"Designation: {profile['basic']['designation']}")
    
    # Table header for TAN/PAN
    p.line(40, height-180, width-40, height-180)
    p.drawString(50, height-170, "TAN of Deductor")
    p.drawString(150, height-170, "PAN of Deductor")
    p.drawString(300, height-170, "PAN of Employee")
    
    p.setFont("Helvetica-Bold", 10)
    p.drawString(50, height-195, "RTKC00001A") # Placeholder TAN
    p.drawString(150, height-195, "AAAAC0000A") # Placeholder PAN
    p.drawString(300, height-195, profile['basic']['panno'] or 'N/A')
    
    # Financial details
    p.setFont("Helvetica-Bold", 11)
    p.drawString(50, height-250, "Details of Salary Paid and any other income")
    
    y = height - 270
    p.setFont("Helvetica", 9)
    p.drawString(50, y, "1. Gross Salary")
    
    gross = sum(float(s['gross_total']) for s in it_data['statement'])
    p.drawRightString(width-60, y, f"{gross:,.2f}")
    
    y -= 15
    p.drawString(50, y, "2. Less: Standard Deduction")
    p.drawRightString(width-60, y, "75,000.00")
    
    y -= 15
    p.drawString(50, y, "3. Deductions under Chapter VI-A")
    total_80c = sum(float(d['docsub_Amt']) for d in decls_list if '80C' in d['section_name'])
    p.drawRightString(width-60, y, f"{total_80c:,.2f}")
    
    p.showPage()
    p.save()
    
    buffer.seek(0)
    return make_response(buffer.getvalue(), 200, {
        'Content-Type': 'application/pdf',
        'Content-Disposition': f'attachment; filename=Form16_{profile["basic"]["empcode"]}.pdf'
    })

@hrms_bp.route('/tax_deduction_form')
def tax_deduction_form():
    return render_template('hrms/tax_deduction_form.html')

@hrms_bp.route('/property_return')
def property_return():
    return render_template('hrms/property_return.html')

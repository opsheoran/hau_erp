import os
from werkzeug.utils import secure_filename
from flask import Blueprint, render_template, request, jsonify, session, redirect, url_for, flash, make_response, send_file, current_app
import math
import io
from datetime import datetime
import pandas as pd
from app.db import DB
from app.models import (
    AcademicsModel, NavModel, InfrastructureModel, ClassificationModel, 
    CourseModel, ActivityCertificateModel, StudentConfigModel, 
    CourseActivityModel, PackageMasterModel, BoardMasterModel, 
    CertificateMasterModel, PgsCourseLimitModel, SeatDetailModel, 
    AdmissionModel, AdvisoryModel, StudentModel, CourseAllocationModel, ResearchModel, ThesisModel,
    AdvisoryStatusModel, IGradeModel, BatchModel, EmployeeModel, AdvisorApprovalModel, TeacherApprovalModel, DswApprovalModel, LibraryApprovalModel, FeeApprovalModel, DeanApprovalModel, SyllabusModel, AdvisorAllocationModel
)
from functools import wraps
from app.utils import get_page_url, get_pagination, get_pagination_range, clean_json_data, english_to_hindi

academics_bp = Blueprint('academics', __name__)

@academics_bp.before_request
def ensure_module():
    if 'user_id' not in session:
        return redirect(url_for('auth.login'))
    session['current_module_id'] = 55

# --- THE DEFINITIVE HIERARCHY AS PER USER INSTRUCTION ---
ACADEMIC_MENU_CONFIG = {
    'Master and Config': {
        'Common Master': {
            'Common Masters preparation': [
                'Academic Session Master', 'Faculty Master', 'Degree Year Master', 
                'Semester Master [Add Semester]', 'Specialization Master [PG/PHD]', 
                'College Type Master', 'Degree Type Master', 'Qualifying/Institution Quota', 
                'Employee- Degree Map'
            ]
        },
        'Administrative Configuration': {
            'Academic Configurations Management': [
                'College Master [Create College]', 'Degree Master [Create Degree]', 
                'Degree Cycle Master', 'Col-Deg-Specialization Master', 'Degree wise Credit Hours', 'Class - Batch Master', 
                'Moderation Marks Detail', 'Degree Wise Credit Hours(Course Plan)'
            ]
        }
    },
    'Course Details': {
        'Course Details': {
            'Course Management': [
                'Course Type Master', 'Paper/Course Title Master', 'Course/Subject Master', 
                'Activity Course Master', 'Activity Master', 'Package Master', 
                'Course Detail Report', 'Syllabus Creation [Master]', 'PGS Course Limit [Master]'
            ]
        }
    },
    'Admission': {
        'Admission Master': {
            'Admission Master Management': [
                'Academic Session Master [A]', 'Faculty Master [A]', 'Degree Year Master [A]', 
                'Semester Master [Add Semester] [A]', 'Specialization Master [PG/PHD] [A]', 
                'College Type Master [A]', 'Degree Type Master [A]', 'Qualifying/Institution Quota [A]', 
                'Employee- Degree Map [A]',
                'Nationality Master', 'Entitlement Master', 'Category Master', 'Board Master', 
                'Certificates Master', 'College Degree Wise Seat Detail', 'Student Registration', 
                'Student Password', 'View Student Password', 'Back Course Entry', 'College Degree Seat Report'
            ]
        },
        'Post Admission Activities': {
            'Admission No. Generation and Management': [
                'Admission No Configuration', 'Admission No Generation', 'Batch Assignment', 'Degree Complete Detail'
            ],
            'Student Bio Data Management': [
                'Student BioData', 'Student Biodata Updation', 'Achievement/ Disciplinary ', 
                'Achievement / Disciplinary Approval', 'Student Activity Management', 'Student Personal Detail Report'
            ],
            'Advisory Committee Management for PG/PHD Programme': [
                'Limit Assignment', 'Major Advisor', 'Specialization Assignment', 'Member of Minor and supporting', 
                'Dean PGS approval (advisory committee)', 'HOD Approval', 'College Dean Approval', 'Prepare Course Plan', 
                'Dean PGS approval (Course plan)', 'Course Allocation(For PG)', 'PG Mandates Submission by HOD',
                'Programme of work(PG)', 'Addition/withdrawal Approval by Teacher', 'Addition/withdrawal Approval by Major Advisor',
                'Student Thesis Detail', 'Addition/Withdrawal Approval Status', 'Advisory Creation And Approval Status',
                'I-Grade Approval by Teacher', 'I-Grade Approval By Dean Pgs', 'I Grade Approval Status'
            ]
        }
    },
    'Academics': {
        'Academics Masters and Cofiguration': {
            'Calender Event/Notification Management': ['Event Master', 'Event Assignment'],
            'Academic Miscellaneous Activities': [
                'Semester Registration', 'Academic Counselling Meeting', 'Previous Question Paper Uploads', 
                'SMS And Mail', 'Student Extension', 'Extension Management', 'Student Transfer Details', 
                'Registration Cancel', 'Student Semester Change', 'Course Approval Status', 
                'rechecking Approval By Hod(PG)', 'Rechecking Approval By Advisor[UG]', 
                'Rechecking Approval By Dean[UG]', 'Revised Result'
            ]
        },
        'Student Course Allocation': {
            'Student Subject/Course Allocation and Management': [
                'Course Offer (By HOD)', 'Course Allocation For UG[Reguler]', 'Advisor Allocation(For UG)', 
                'Teacher Course Assignment', 'Course Allocation Approval For UG/PG[By Advisor]', 
                'Course Allocation Approval For UG/PG[By Teacher]', 'Course Allocation Approval For UG/PG[By DSW]', 
                'Course Allocation Approval For UG/PG[By Library]', 'Course Allocation Fee Approval For UG/PG', 
                'Course Allocation Approval For UG/PG [By Dean]', 'Course Allocation Approval For PG [By DeanPGS]', 
                'Course Allocation Approval For UG/PG [By Registrar]', 'Course Allocation(HOD on Request)', 
                'Compartment & Year Back', 'Course Allocation Report-Regular', 'Activity Allocation', 
                'Course Allocation Report-Regular(Consolidate)', 'Course Allocation Re-Appear', 
                'Course Allocation Report-Compart', 'Course Offer And Teacher Course Assignment Report(For HOD)', 
                'Course Allocation Report-Regular(Consolidate-PG)', 'NSS/NCC Course updation'
            ]
        }
    },
    'Report': {
        'Admission Reports': { 'Admission miscellaneous Reports': [] },
        'Academics Reports': { 'Academic miscellaneous Reports': [] }
    }
}

def permission_required(page_caption):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'user_id' not in session:
                return redirect(url_for('auth.login'))
            user_id = session.get('user_id')
            loc_id = session.get('selected_loc')
            perm = NavModel.check_permission(user_id, loc_id, page_caption)
            if not perm or not perm.get('AllowView'):
                return redirect(url_for('main.index'))
            # Write protection (only when action indicates a write).
            if request.method == 'POST':
                action = (request.form.get('action') or '').strip().upper()
                write_actions = {
                    'SAVE', 'SUBMIT', 'ADD', 'CREATE', 'INSERT', 'UPDATE', 'EDIT',
                    'DELETE', 'REMOVE', 'DEALLOCATE', 'ALLOCATE',
                    'APPROVE', 'HOLD', 'REJECT', 'CANCEL', 'PROCESS', 'UNPROCESS'
                }
                if action in write_actions:
                    if action in {'DELETE', 'REMOVE'} and not perm.get('AllowDelete'):
                        flash('You do not have Delete permission for this page.', 'danger')
                        return redirect(url_for('main.index'))
                    if action in {'ADD', 'CREATE', 'INSERT'} and not perm.get('AllowAdd'):
                        flash('You do not have Add permission for this page.', 'danger')
                        return redirect(url_for('main.index'))
                    if action in {'SAVE', 'SUBMIT', 'UPDATE', 'EDIT', 'APPROVE', 'HOLD', 'REJECT', 'DEALLOCATE', 'ALLOCATE', 'PROCESS', 'UNPROCESS', 'CANCEL'} and not (perm.get('AllowAdd') or perm.get('AllowUpdate')):
                        flash('You do not have permission to perform this action.', 'danger')
                        return redirect(url_for('main.index'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator

# --- CLEAN ROUTES ---

@academics_bp.route('/session_master', methods=['GET', 'POST'])
@permission_required('Academic Session Master')
def session_master():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'DELETE':
            if InfrastructureModel.delete_session(request.form.get('id')):
                flash('Session deleted successfully!', 'success')
            else:
                flash('Error deleting session.', 'danger')
        else:
            if InfrastructureModel.save_session(request.form):
                flash('Session saved successfully!', 'success')
            else:
                flash('Error saving session.', 'danger')
        return redirect(url_for('academics.session_master'))
    
    page = request.args.get('page', 1, type=int)
    per_page = 10
    sessions, total = InfrastructureModel.get_sessions_paginated(page=page, per_page=per_page)
    
    pagination = {
        'page': page,
        'per_page': per_page,
        'total': total,
        'total_pages': math.ceil(total / per_page) if total else 1,
        'has_prev': page > 1,
        'has_next': page < (math.ceil(total / per_page) if total else 1)
    }

    page_range = get_pagination_range(page, pagination['total_pages'])

    # clean_json_data converts datetimes to 'DD/MM/YYYY' strings that the frontend can safely handle
    sessions = clean_json_data(sessions)

    return render_template('academics/session_master.html',
                           sessions=sessions,
                           sessions_json=sessions,
                           pagination=pagination,
                           page_range=page_range)

@academics_bp.route('/faculty_master', methods=['GET', 'POST'])
@permission_required('Faculty Master')
def faculty_master():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'DELETE':
            if AcademicsModel.delete_faculty(request.form.get('id')):
                flash('Faculty deleted successfully!', 'success')
            else:
                flash('Error deleting faculty.', 'danger')
        else:
            if AcademicsModel.save_faculty(request.form):
                flash('Faculty saved successfully!', 'success')
            else:
                flash('Error saving faculty.', 'danger')
        return redirect(url_for('academics.faculty_master'))
    faculties = AcademicsModel.get_faculties()
    return render_template('academics/faculty_master.html', faculties=faculties)

@academics_bp.route('/degree_year_master', methods=['GET', 'POST'])
@permission_required('Degree Year Master')
def degree_year_master():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'DELETE':
            if AcademicsModel.delete_degree_year(request.form.get('id')):
                flash('Degree Year deleted successfully!', 'success')
            else:
                flash('Error deleting degree year.', 'danger')
        else:
            if AcademicsModel.save_degree_year(request.form):
                flash('Degree Year saved successfully!', 'success')
            else:
                flash('Error saving degree year.', 'danger')
        return redirect(url_for('academics.degree_year_master'))
    
    page = request.args.get('page', 1, type=int)
    per_page = 10
    years, total = AcademicsModel.get_degree_years_paginated(page=page, per_page=per_page)
    
    pagination = {
        'page': page,
        'per_page': per_page,
        'total': total,
        'total_pages': math.ceil(total / per_page) if total else 1,
        'has_prev': page > 1,
        'has_next': page < (math.ceil(total / per_page) if total else 1)
    }
    
    page_range = get_pagination_range(page, pagination['total_pages'])

    return render_template('academics/degree_year_master.html', 
                           years=years, 
                           pagination=pagination,
                           page_range=page_range)

@academics_bp.route('/semester_master', methods=['GET', 'POST'])
@permission_required('Semester Master [Add Semester]')
def semester_master():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'DELETE':
            if InfrastructureModel.delete_semester(request.form.get('id')):
                flash('Semester deleted successfully!', 'success')
            else:
                flash('Error deleting semester.', 'danger')
        else:
            if InfrastructureModel.save_semester(request.form):
                flash('Semester saved successfully!', 'success')
            else:
                flash('Error saving semester.', 'danger')
        return redirect(url_for('academics.semester_master'))
    
    page = request.args.get('page', 1, type=int)
    per_page = 10
    semesters, total = InfrastructureModel.get_semesters_paginated(page=page, per_page=per_page)
    years = AcademicsModel.get_degree_years()
    
    pagination = {
        'page': page,
        'per_page': per_page,
        'total': total,
        'total_pages': math.ceil(total / per_page) if total else 1,
        'has_prev': page > 1,
        'has_next': page < (math.ceil(total / per_page) if total else 1)
    }
    
    page_range = get_pagination_range(page, pagination['total_pages'])
    
    return render_template('academics/semester_master.html', 
                           semesters=semesters, 
                           years=years,
                           pagination=pagination,
                           page_range=page_range)

@academics_bp.route('/branch_master', methods=['GET', 'POST'])
@permission_required('Specialization Master [PG/PHD]')
def branch_master():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'DELETE':
            if AcademicsModel.delete_branch(request.form.get('id')):
                flash('Specialization deleted successfully!', 'success')
            else:
                flash('Error deleting specialization.', 'danger')
        else:
            if AcademicsModel.save_branch(request.form):
                flash('Specialization saved successfully!', 'success')
            else:
                flash('Error saving specialization.', 'danger')
        return redirect(url_for('academics.branch_master'))
    
    page = request.args.get('page', 1, type=int)
    per_page = 10
    branches, total = AcademicsModel.get_branches_paginated(page=page, per_page=per_page)
    faculties = AcademicsModel.get_faculties()
    
    pagination = {
        'page': page,
        'per_page': per_page,
        'total': total,
        'total_pages': math.ceil(total / per_page) if total else 1,
        'has_prev': page > 1,
        'has_next': page < (math.ceil(total / per_page) if total else 1)
    }
    
    page_range = get_pagination_range(page, pagination['total_pages'])
    
    return render_template('academics/branch_master.html', 
                           branches=branches, 
                           faculties=faculties,
                           pagination=pagination,
                           page_range=page_range)

@academics_bp.route('/college_type_master', methods=['GET', 'POST'])
@permission_required('College Type Master')
def college_type_master():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'DELETE':
            if ClassificationModel.delete_college_type(request.form.get('id')):
                flash('College Type deleted successfully!', 'success')
            else:
                flash('Error deleting college type.', 'danger')
        else:
            if ClassificationModel.save_college_type(request.form):
                flash('College Type saved successfully!', 'success')
            else:
                flash('Error saving college type.', 'danger')
        return redirect(url_for('academics.college_type_master'))
    
    page = request.args.get('page', 1, type=int)
    per_page = 10
    types, total = ClassificationModel.get_college_types_paginated(page=page, per_page=per_page)
    
    pagination = {
        'page': page,
        'per_page': per_page,
        'total': total,
        'total_pages': math.ceil(total / per_page) if total else 1,
        'has_prev': page > 1,
        'has_next': page < (math.ceil(total / per_page) if total else 1)
    }
    
    page_range = get_pagination_range(page, pagination['total_pages'])
    
    return render_template('academics/college_type_master.html', types=types, pagination=pagination, page_range=page_range)
@academics_bp.route('/degree_type_master', methods=['GET', 'POST'])
@permission_required('Degree Type Master')
def degree_type_master():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'DELETE':
            if ClassificationModel.delete_degree_type(request.form.get('id')):
                flash('Degree Type deleted successfully!', 'success')
            else:
                flash('Error deleting degree type.', 'danger')
        else:
            if ClassificationModel.save_degree_type(request.form):
                flash('Degree Type saved successfully!', 'success')
            else:
                flash('Error saving degree type.', 'danger')
        return redirect(url_for('academics.degree_type_master'))
    
    page = request.args.get('page', 1, type=int)
    per_page = 10
    types, total = ClassificationModel.get_degree_types_paginated(page=page, per_page=per_page)
    
    pagination = {
        'page': page,
        'per_page': per_page,
        'total': total,
        'total_pages': math.ceil(total / per_page) if total else 1,
        'has_prev': page > 1,
        'has_next': page < (math.ceil(total / per_page) if total else 1)
    }
    
    page_range = get_pagination_range(page, pagination['total_pages'])
    
    return render_template('academics/degree_type_master.html', 
                           types=types, 
                           pagination=pagination,
                           page_range=page_range)

@academics_bp.route('/rank_master', methods=['GET', 'POST'])
@permission_required('Qualifying/Institution Quota')
def rank_master():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'DELETE':
            if StudentConfigModel.delete_rank(request.form.get('id')):
                flash('Quota deleted successfully!', 'success')
            else:
                flash('Error deleting quota.', 'danger')
        else:
            if StudentConfigModel.save_rank(request.form):
                flash('Quota saved successfully!', 'success')
            else:
                flash('Error saving quota.', 'danger')
        return redirect(url_for('academics.rank_master'))
    
    page = request.args.get('page', 1, type=int)
    per_page = 10
    items, total = StudentConfigModel.get_ranks(page=page, per_page=per_page)
    
    pagination = {
        'page': page,
        'per_page': per_page,
        'total': total,
        'total_pages': math.ceil(total / per_page) if total else 1,
        'has_prev': page > 1,
        'has_next': page < (math.ceil(total / per_page) if total else 1)
    }
    
    page_range = get_pagination_range(page, pagination['total_pages'])
    
    return render_template('academics/rank_master.html', items=items, pagination=pagination, page_range=page_range)

@academics_bp.route('/api/transliterate_hindi', methods=['POST'])
def api_transliterate_hindi():
    """Convert English text to Hindi (Devanagari) transliteration.

    POST JSON: {"text": "Ramesh Kumar"}
    Response:  {"hindi": "रमेश कुमार"}
    """
    data = request.get_json(silent=True) or {}
    text = data.get('text', '')
    return jsonify({'hindi': english_to_hindi(text)})


@academics_bp.route('/api/college/<college_id>/degrees')
def get_college_degrees_api(college_id):
    from app.utils import clean_json_data
    from app.db import DB
    
    is_pg_phd = request.args.get('is_pg_phd') == '1'
    if is_pg_phd:
        degrees = DB.fetch_all("""
            SELECT DISTINCT D.pk_degreeid as id, D.degreename as name
            FROM SMS_CollegeDegreeBranchMap_Mst M
            INNER JOIN SMS_Degree_Mst D ON M.fk_degreeid = D.pk_degreeid
            WHERE M.fk_collegeid = ? AND D.degreename NOT LIKE '%---%' AND D.fk_degreetypeid IN (1, 2, 4)
            ORDER BY D.degreename
        """, [college_id])
    else:
        degrees = AcademicsModel.get_college_degrees(college_id)
    return jsonify(clean_json_data(degrees))

@academics_bp.route('/api/search_teachers')
def search_teachers_api():
    term = request.args.get('q', '').strip()
    include_all = request.args.get('all') == '1'
    emp_id = session.get('emp_id')
    
    # Use EmployeeModel.search_employees with only_teachers=True
    results = EmployeeModel.search_employees(term, only_teachers=True)
    
    # Filter by department if not include_all (for departmental dropdown population)
    if not include_all and emp_id:
        ctx = AcademicsModel.get_hod_department_context(emp_id)
        hod_dept_ids = {str(d['id']) for d in ctx.get('hr_departments', [])}
        filtered = []
        for r in results:
            if str(r.get('fk_deptid')) in hod_dept_ids:
                filtered.append(r)
        results = filtered

    # Ensure format matches what ddlEmployee expect: {id, name}
    formatted = []
    for r in results:
        formatted.append({
            'id': r.get('id') or r.get('pk_empid'),
            'name': r.get('name') or (f"{r.get('empname')} | {r.get('empcode')}")
        })

    return jsonify(clean_json_data(formatted))

@academics_bp.route('/employee_degree_map', methods=['GET', 'POST'])
@permission_required('Employee- Degree Map')
def employee_degree_map():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'DELETE':
            if AcademicsModel.delete_employee_degree_mapping(request.form.get('id')):
                flash('Mapping deleted successfully!', 'success')
            else:
                flash('Error deleting mapping.', 'danger')
        else:
            if AcademicsModel.save_employee_degree_mapping(request.form):
                flash('Employee-Degree mapping saved successfully!', 'success')
            else:
                flash('Error saving mapping.', 'danger')
        return redirect(url_for('academics.employee_degree_map'))
    
    page = request.args.get('page', 1, type=int)
    per_page = 10
    items, total = AcademicsModel.get_employee_degree_mappings_paginated(page=page, per_page=per_page)
    
    pagination = {
        'page': page,
        'per_page': per_page,
        'total': total,
        'total_pages': math.ceil(total / per_page) if total else 1,
        'has_prev': page > 1,
        'has_next': page < (math.ceil(total / per_page) if total else 1)
    }
    
    page_range = get_pagination_range(page, pagination['total_pages'])

    lookups = {
        'colleges': AcademicsModel.get_colleges_simple(),
        'degrees': AcademicsModel.get_all_degrees(),
        'employees': DB.fetch_all("""
            SELECT U.pk_userId as id, E.empname + ' [ User Name -{ ' + U.loginname + ' } ]' as name
            FROM UM_Users_Mst U
            INNER JOIN SAL_Employee_Mst E ON U.fk_empId = E.pk_empid
            WHERE U.active = 1
            ORDER BY E.empname
        """)
    }
    
    return render_template('academics/employee_degree_map.html', items=items, lookups=lookups, pagination=pagination, page_range=page_range)

@academics_bp.route('/college_master', methods=['GET', 'POST'])
@permission_required('College Master [Create College]')
def college_master():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'DELETE':
            if AcademicsModel.delete_college(request.form.get('id')):
                flash('College deleted successfully!', 'success')
            else:
                flash('Error deleting college.', 'danger')
        else:
            if AcademicsModel.save_college(request.form):
                flash('College saved successfully!', 'success')
            else:
                flash('Error saving college.', 'danger')
        return redirect(url_for('academics.college_master'))
    
    page = request.args.get('page', 1, type=int)
    per_page = 10
    items, total = AcademicsModel.get_colleges(page=page, per_page=per_page)
    
    pagination = {
        'page': page,
        'per_page': per_page,
        'total': total,
        'total_pages': math.ceil(total / per_page) if total else 1,
        'has_prev': page > 1,
        'has_next': page < (math.ceil(total / per_page) if total else 1)
    }
    
    page_range = get_pagination_range(page, pagination['total_pages'])

    lookups = {
        'types': ClassificationModel.get_college_types(),
        'cities': AcademicsModel.get_cities(),
        'locations': DB.fetch_all("SELECT pk_locid as id, locname as name, * FROM Location_Mst ORDER BY locname")
    }
    
    return render_template('academics/college_master.html', items=items, lookups=lookups, pagination=pagination, page_range=page_range)

@academics_bp.route('/api/get_college_details/<int:college_id>')
def get_college_details_data_api(college_id):
    data = AcademicsModel.get_college_full_details(college_id)
    return jsonify(data)

@academics_bp.route('/api/get_all_employees')
def get_all_employees_api():
    sql = "SELECT pk_empid as id, empname + ' || ' + ISNULL(empcode, '') as name FROM SAL_Employee_Mst WHERE employeeleftstatus='N' AND stopsalary=0 ORDER BY empname"
    emps = DB.fetch_all(sql)
    return jsonify(emps)

@academics_bp.route('/degree_master', methods=['GET', 'POST'])
@permission_required('Degree Master [Create Degree]')
def degree_master():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'DELETE':
            if AcademicsModel.delete_degree(request.form.get('id')):
                flash('Degree deleted successfully!', 'success')
            else:
                flash('Error deleting degree.', 'danger')
        else:
            if AcademicsModel.save_degree(request.form):
                flash('Degree saved successfully!', 'success')
            else:
                flash('Error saving degree.', 'danger')
        return redirect(url_for('academics.degree_master'))
    
    page = request.args.get('page', 1, type=int)
    per_page = 10
    items, total = AcademicsModel.get_degrees_paginated(page=page, per_page=per_page)
    
    pagination = {
        'page': page,
        'per_page': per_page,
        'total': total,
        'total_pages': math.ceil(total / per_page) if total else 1,
        'has_prev': page > 1,
        'has_next': page < (math.ceil(total / per_page) if total else 1)
    }
    
    page_range = get_pagination_range(page, pagination['total_pages'])

    lookups = {
        'types': ClassificationModel.get_degree_types()
    }
    
    return render_template('academics/degree_master.html', items=items, lookups=lookups, pagination=pagination, page_range=page_range)

@academics_bp.route('/degree_cycle_master', methods=['GET', 'POST'])
@permission_required('Degree Cycle Master')
def degree_cycle_master():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'DELETE':
            if AcademicsModel.delete_degree_cycle(request.form.get('id')):
                flash('Degree cycle deleted successfully!', 'success')
            else:
                flash('Error deleting cycle.', 'danger')
        else:
            if AcademicsModel.save_degree_cycle(request.form):
                flash('Degree cycle saved successfully!', 'success')
            else:
                flash('Error saving cycle.', 'danger')
        return redirect(url_for('academics.degree_cycle_master'))
    
    page = request.args.get('page', 1, type=int)
    per_page = 10
    items, total = AcademicsModel.get_degree_cycles_paginated(page=page, per_page=per_page)
    
    pagination = {
        'page': page,
        'per_page': per_page,
        'total': total,
        'total_pages': math.ceil(total / per_page) if total else 1,
        'has_prev': page > 1,
        'has_next': page < (math.ceil(total / per_page) if total else 1)
    }
    
    page_range = get_pagination_range(page, pagination['total_pages'])

    lookups = {
        'degrees': AcademicsModel.get_all_degrees(),
        'branches': AcademicsModel.get_branches(),
        'years': AcademicsModel.get_degree_years(),
        'semesters': InfrastructureModel.get_all_semesters()
    }
    
    return render_template('academics/degree_cycle_master.html', items=items, lookups=lookups, pagination=pagination, page_range=page_range)

@academics_bp.route('/course_type_master', methods=['GET', 'POST'])
@permission_required('Course Type Master')
def course_type_master():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'DELETE':
            if ClassificationModel.delete_course_type(request.form.get('id')):
                flash('Course Type deleted successfully!', 'success')
            else:
                flash('Error deleting course type.', 'danger')
        else:
            if ClassificationModel.save_course_type(request.form):
                flash('Course Type saved successfully!', 'success')
            else:
                flash('Error saving course type.', 'danger')
        return redirect(url_for('academics.course_type_master'))
    
    page = request.args.get('page', 1, type=int)
    per_page = 10
    types, total = ClassificationModel.get_course_types_paginated(page=page, per_page=per_page)
    
    pagination = {
        'page': page,
        'per_page': per_page,
        'total': total,
        'total_pages': math.ceil(total / per_page) if total else 1,
        'has_prev': page > 1,
        'has_next': page < (math.ceil(total / per_page) if total else 1)
    }
    
    page_range = get_pagination_range(page, pagination['total_pages'])
    
    return render_template('academics/course_type_master.html', types=types, pagination=pagination, page_range=page_range)

@academics_bp.route('/paper_title_master', methods=['GET', 'POST'])
@permission_required('Paper/Course Title Master')
def paper_title_master():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'DELETE':
            if CourseModel.delete_paper_title(request.form.get('id')):
                flash('Paper Title deleted successfully!', 'success')
            else:
                flash('Error deleting paper title.', 'danger')
        else:
            if CourseModel.save_paper_title(request.form):
                flash('Paper Title saved successfully!', 'success')
            else:
                flash('Error saving paper title.', 'danger')
        return redirect(url_for('academics.paper_title_master'))
    
    page = request.args.get('page', 1, type=int)
    per_page = 10
    titles, total = CourseModel.get_paper_titles(page=page, per_page=per_page)
    
    pagination = {
        'page': page,
        'per_page': per_page,
        'total': total,
        'total_pages': math.ceil(total / per_page) if total else 1,
        'has_prev': page > 1,
        'has_next': page < (math.ceil(total / per_page) if total else 1)
    }
    
    page_range = get_pagination_range(page, pagination['total_pages'])
    
    return render_template('academics/paper_title_master.html', titles=titles, pagination=pagination, page_range=page_range)

@academics_bp.route('/course_master', methods=['GET', 'POST'])
@permission_required('Course/Subject Master')
def course_master():
    user_id = session['user_id']
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'DELETE':
            if CourseModel.delete_course(request.form.get('id')):
                flash('Course deleted successfully!', 'success')
            else:
                flash('Error deleting course.', 'danger')
        else:
            if CourseModel.save_course(request.form, user_id):
                flash('Course saved successfully!', 'success')
            else:
                flash('Error saving course.', 'danger')
        return redirect(url_for('academics.course_master'))

    page = request.args.get('page', 1, type=int)
    per_page = 10
    filters = {'dept_id': request.args.get('dept_id'), 'term': request.args.get('term')}
    courses, total = CourseModel.get_courses(filters, page=page, per_page=per_page)
    
    pagination = {
        'page': page,
        'per_page': per_page,
        'total': total,
        'total_pages': math.ceil(total / per_page) if total else 1,
        'has_prev': page > 1,
        'has_next': page < (math.ceil(total / per_page) if total else 1)
    }
    
    page_range = get_pagination_range(page, pagination['total_pages'])

    lookups = {
        'depts': AcademicsModel.get_departments(),
        'titles': CourseModel.get_all_paper_titles(),
        'types': ClassificationModel.get_course_types(),
        'sessions': InfrastructureModel.get_sessions(),
        'degrees': AcademicsModel.get_all_degrees(),
        'semesters': InfrastructureModel.get_all_semesters(),
        'branches': AcademicsModel.get_branches()
    }
    
    return render_template('academics/course_master.html', 
                           courses=courses, 
                           lookups=lookups, 
                           filters=filters,
                           pagination=pagination,
                           page_range=page_range)

@academics_bp.route('/api/get_course_details/<int:course_id>')
def get_course_details_api(course_id):
    details = CourseModel.get_course_details(course_id)
    return jsonify(clean_json_data(details))

@academics_bp.route('/entitlement_master', methods=['GET', 'POST'])
@permission_required('Entitlement Master')
def entitlement_master():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'DELETE':
            if StudentConfigModel.delete_entitlement(request.form.get('id')):
                flash('Entitlement deleted successfully!', 'success')
            else:
                flash('Error deleting entitlement.', 'danger')
        else:
            if StudentConfigModel.save_entitlement(request.form):
                flash('Entitlement saved successfully!', 'success')
            else:
                flash('Error saving entitlement.', 'danger')
        return redirect(url_for('academics.entitlement_master'))

    page = request.args.get('page', 1, type=int)
    per_page = 10
    items, total = StudentConfigModel.get_entitlements(page=page, per_page=per_page)
    
    pagination = {
        'page': page,
        'per_page': per_page,
        'total': total,
        'total_pages': math.ceil(total / per_page) if total else 1,
        'has_prev': page > 1,
        'has_next': page < (math.ceil(total / per_page) if total else 1)
    }
    
    page_range = get_pagination_range(page, pagination['total_pages'])
    
    return render_template('academics/entitlement_master.html', items=items, pagination=pagination, page_range=page_range)

@academics_bp.route('/category_master', methods=['GET', 'POST'])
@permission_required('Category Master')
def category_master():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'DELETE':
            if ClassificationModel.delete_category(request.form.get('id')):
                flash('Category deleted successfully!', 'success')
            else:
                flash('Error deleting category.', 'danger')
        else:
            if ClassificationModel.save_category(request.form):
                flash('Category saved successfully!', 'success')
            else:
                flash('Error saving category.', 'danger')
        return redirect(url_for('academics.category_master'))

    page = request.args.get('page', 1, type=int)
    per_page = 10
    items, total = ClassificationModel.get_categories(page=page, per_page=per_page)
    
    pagination = {
        'page': page,
        'per_page': per_page,
        'total': total,
        'total_pages': math.ceil(total / per_page) if total else 1,
        'has_prev': page > 1,
        'has_next': page < (math.ceil(total / per_page) if total else 1)
    }
    
    page_range = get_pagination_range(page, pagination['total_pages'])
    
    return render_template('academics/category_master.html', items=items, pagination=pagination, page_range=page_range)

@academics_bp.route('/nationality_master', methods=['GET', 'POST'])
@permission_required('Nationality Master')
def nationality_master():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'DELETE':
            if ClassificationModel.delete_nationality(request.form.get('id')):
                flash('Nationality deleted successfully!', 'success')
            else:
                flash('Error deleting nationality.', 'danger')
        else:
            if ClassificationModel.save_nationality(request.form):
                flash('Nationality saved successfully!', 'success')
            else:
                flash('Error saving nationality.', 'danger')
        return redirect(url_for('academics.nationality_master'))
    nationalities = ClassificationModel.get_nationalities()
    return render_template('academics/nationality_master.html', nationalities=nationalities)

@academics_bp.route('/certificate_master', methods=['GET', 'POST'])
@permission_required('Certificates Master')
def certificate_master():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'DELETE':
            DB.execute("DELETE FROM SMS_Certificate_Mst WHERE pk_certificateid = ?", [request.form.get('id')])
            flash('Certificate deleted successfully!', 'success')
        else:
            if AdmissionModel.save_certificate(request.form):
                flash('Certificate saved successfully!', 'success')
            else:
                flash('Error saving certificate.', 'danger')
        return redirect(url_for('academics.certificate_master'))

    page = request.args.get('page', 1, type=int)
    per_page = 10
    items, total = AdmissionModel.get_certificates(page=page, per_page=per_page)
    
    pagination = {
        'page': page, 'per_page': per_page, 'total': total,
        'total_pages': math.ceil(total / per_page) if total else 1,
        'has_prev': page > 1,
        'has_next': page < (math.ceil(total / per_page) if total else 1)
    }
    page_range = get_pagination_range(page, pagination['total_pages'])
    return render_template('academics/certificate_master.html', items=items, pagination=pagination, page_range=page_range)

@academics_bp.route('/api/student/profile_basic/<int:sid>')
def api_get_student_profile_basic(sid):
    data = StudentModel.get_student_profile_basic(sid)
    if data:
        return jsonify(clean_json_data(data))
    return jsonify({'error': 'Student not found'}), 404

@academics_bp.route('/student_biodata', methods=['GET', 'POST'])
@academics_bp.route('/student_biodata/<int:sid>', methods=['GET', 'POST'])
@permission_required('Student BioData')
def student_biodata(sid=None):
    if request.method == 'POST':
        if StudentModel.save_student_biodata(request.form, session['user_id']):
            flash('Student BioData saved successfully!', 'success')
        else:
            flash('Error saving student BioData.', 'danger')
        return redirect(url_for('academics.student_biodata', sid=request.form.get('pk_sid')))
    
    if not sid:
        sid = request.args.get('sid')
        
    student_data = clean_json_data(StudentModel.get_student_all_details(sid)) if sid else None
    lookups = StudentModel.get_student_lookups()
    
    return render_template('academics/student_biodata.html', student_data=student_data, lookups=lookups)

@academics_bp.route('/api/student/get_full_details')
def api_get_student_details():
    sid = request.args.get('sid')
    enrollment_no = request.args.get('enrollment_no')
    
    if enrollment_no:
        basic = StudentModel.get_student_by_enrollment(enrollment_no)
        if basic:
            sid = basic['pk_sid']
        else:
            return jsonify({'error': 'Student not found'}), 404
            
    if not sid:
        return jsonify({'error': 'No student identifier provided'}), 400
        
    data = StudentModel.get_student_all_details(sid)
    if data:
        # Format dates for JSON (clean_json_data also handles this, but keeping for explicitness)
        if data['basic'].get('dob'): data['basic']['dob'] = data['basic']['dob'].isoformat()
        if data['basic'].get('FatherDOB'): data['basic']['FatherDOB'] = data['basic']['FatherDOB'].isoformat()
        if data['basic'].get('MotherDOB'): data['basic']['MotherDOB'] = data['basic']['MotherDOB'].isoformat()
        if data['basic'].get('PerentMarriedDate'): data['basic']['PerentMarriedDate'] = data['basic']['PerentMarriedDate'].isoformat()
        
        return jsonify(clean_json_data(data))
    return jsonify({'error': 'Student details not found'}), 404

@academics_bp.route('/upload_certificate', methods=['POST'])
@permission_required('Student Registration')
def upload_certificate():
    if 'cert_file' not in request.files:
        return jsonify({'success': False, 'message': 'No file part'})
    
    file = request.files['cert_file']
    cert_id = request.form.get('upload_cert_id')
    sid = request.form.get('sid') # Optional, for new reg we might use session or temp ID
    
    if file.filename == '':
        return jsonify({'success': False, 'message': 'No selected file'})
    
    if file:
        filename = secure_filename(file.filename)
        # Store in temp directory or final path
        upload_path = os.path.join('app/static/uploads/certificates', filename)
        os.makedirs(os.path.dirname(upload_path), exist_ok=True)
        file.save(upload_path)
        
        # Save to DB if sid is present
        if sid:
            sql = """INSERT INTO Sms_StudentCertificateUpload_Dtl (fk_sid, UploadCertificateFileName, OriginalFileName, FilePath, fk_certificateId)
                     VALUES (?, ?, ?, ?, ?)"""
            DB.execute(sql, [sid, filename, file.filename, upload_path, cert_id])
            
        return jsonify({'success': True, 'filename': filename})

@academics_bp.route('/api/search_students')
@permission_required('Student Registration')
def api_search_students():
    filters = {
        'name': request.args.get('name', '').strip(),
        'gender': request.args.get('gender', '').strip(),
        'college_id': request.args.get('college_id', '').strip(),
        'session_id': request.args.get('session_id', '').strip(),
        'degree_id': request.args.get('degree_id', '').strip(),
        'semester_id': request.args.get('semester_id', '').strip(),
        'branch_id': request.args.get('branch_id', '').strip(),
        'seat_type_id': request.args.get('seat_type_id', '').strip(),
        'cat_id': request.args.get('cat_id', '').strip(),
        'admission_no': request.args.get('admission_no', '').strip(),
        'reg_status': request.args.get('reg_status', '').strip(),
    }
    # Remove empty values
    filters = {k: v for k, v in filters.items() if v}
    if not filters:
        return jsonify([])
    rows = StudentModel.get_students_by_filter(filters)
    result = []
    for r in rows:
        result.append({
            'pk_sid': r['pk_sid'],
            'fullname': r.get('fullname') or '',
            'AdmissionNo': r.get('AdmissionNo') or '',
            'enrollmentno': r.get('enrollmentno') or '',
            'collegename': r.get('collegename') or '',
            'degreename': r.get('degreename') or '',
            'semester_roman': (r.get('semester_roman') or '').strip(),
            'branchname': r.get('branchname') or '',
            'seatype': r.get('seatype') or '',
            'category': r.get('category') or '',
            'reg_status': r.get('reg_status') or '',
            'adm_session_name': r.get('adm_session_name') or '',
        })
    return jsonify(result)

@academics_bp.route('/student_registration', methods=['GET', 'POST'])
@permission_required('Student Registration')
def student_registration():
    import decimal as _decimal
    if request.method == 'POST':
        form_data = request.form.to_dict()
        pk_sid = form_data.get('pk_sid', '').strip()
        action = form_data.get('action', 'SAVE_BASIC').upper()

        if action == 'SAVE_QUALIFICATION':
            if not pk_sid:
                flash('Please save basic details first before saving qualifications.', 'warning')
            else:
                # Build list of qualification rows from form arrays
                exam_ids   = request.form.getlist('exam_id[]')
                boards     = request.form.getlist('board[]')
                years      = request.form.getlist('year[]')
                roll_nos   = request.form.getlist('roll_no[]')
                max_marks  = request.form.getlist('max_marks[]')
                marks_obt  = request.form.getlist('marks_obt[]')
                pers       = request.form.getlist('per[]')
                subjects   = request.form.getlist('subjects[]')
                qual_rows = []
                for i in range(len(exam_ids)):
                    qual_rows.append({
                        'exam_id':   exam_ids[i] if i < len(exam_ids) else '',
                        'board':     boards[i]   if i < len(boards)   else '',
                        'year':      years[i]    if i < len(years)    else '',
                        'roll_no':   roll_nos[i] if i < len(roll_nos) else '',
                        'max_marks': max_marks[i]if i < len(max_marks)else '',
                        'per':       pers[i]     if i < len(pers)     else '',
                        'subjects':  subjects[i] if i < len(subjects) else '',
                    })
                if StudentModel.save_student_qualifications(pk_sid, qual_rows):
                    flash('Qualifications saved successfully!', 'success')
                else:
                    flash('Error saving qualifications.', 'danger')
            return redirect(url_for('academics.student_registration') + f'?sid={pk_sid}&tab=qualification')

        # Default: save basic details
        if pk_sid:
            form_data['pk_sid'] = pk_sid
        result = StudentModel.save_student(form_data)
        if result is True or result == 1:
            flash('Student saved successfully!', 'success')
        else:
            flash(f'Error saving student: {result}', 'danger')
        sid = pk_sid or ''
        return redirect(url_for('academics.student_registration') + (f'?sid={sid}' if sid else ''))

    sd = None
    qualifications = []
    certificates = []
    sid = request.args.get('sid', '').strip()
    active_tab = request.args.get('tab', 'basic')
    if not sid:
        enroll = request.args.get('enroll', '').strip()
        if enroll:
            row = StudentModel.get_student_by_enrollment(enroll)
            if row:
                sid = str(row['pk_sid'])
    if sid:
        details = StudentModel.get_student_all_details(sid)
        if details and details.get('basic'):
            sd = dict(details['basic'])
            for k, v in list(sd.items()):
                if hasattr(v, 'strftime'):
                    sd[k] = v.strftime('%Y-%m-%d')
                elif isinstance(v, _decimal.Decimal):
                    sd[k] = int(v) if v == int(v) else float(v)
                elif isinstance(v, (bytes, bytearray, memoryview)):
                    sd[k] = None
        qualifications = StudentModel.get_student_qualifications(sid)
        certificates   = StudentModel.get_student_certificates(sid)

    lookups = StudentModel.get_student_lookups()
    return render_template('academics/student_registration.html',
                           lookups=lookups, sd=sd,
                           qualifications=qualifications,
                           certificates=certificates,
                           active_tab=active_tab)

@academics_bp.route('/student_password', methods=['GET', 'POST'])
@permission_required('Student Password')
def student_password():
    if request.method == 'POST':
        # Logic for resetting password
        flash('Password updated successfully!', 'success')
        return redirect(url_for('academics.student_password'))
    
    colleges = AcademicsModel.get_colleges_simple()
    degrees = AcademicsModel.get_all_degrees()
    sessions = InfrastructureModel.get_sessions()
    return render_template('academics/student_password.html', colleges=colleges, degrees=degrees, sessions=sessions)

@academics_bp.route('/view_student_password', methods=['GET', 'POST'])
@permission_required('View Student Password')
def view_student_password():
    student_password = None
    enrollment_no = None
    
    action = None
    if request.method == 'POST':
        enrollment_no = (request.form.get('enrollment_no') or '').strip()
        action = request.form.get('action')
        
        if not enrollment_no:
            flash('Please enter Enrollment No.', 'warning')
        elif action == 'Get Password':
            student_password = StudentModel.get_student_password(enrollment_no)
            if not student_password:
                flash('Student not found or password not available.', 'warning')
        elif action == 'Unlock Biodata':
            if StudentModel.unlock_student_biodata(enrollment_no):
                flash('Biodata unlocked successfully for ' + enrollment_no, 'success')
            else:
                flash('Student not found or could not unlock biodata.', 'danger')
        elif action == 'Card Entry Submit':
            StudentModel.update_card_entry_status(enrollment_no)
            flash('Card entry status updated for ' + enrollment_no, 'success')
            
    return render_template('academics/view_student_password.html', 
                           student_password=student_password,
                           enrollment_no=enrollment_no,
                           action=action)

@academics_bp.route('/validate_student/<enrollment_no>')
@permission_required('Back Course Entry')
def validate_student(enrollment_no):
    student = StudentModel.get_student_by_enrollment(enrollment_no)
    if student:
        return jsonify({'success': True, 'name': student['fullname'], 'sid': student['pk_sid']})
    return jsonify({'success': False, 'message': 'Student not found.'})

@academics_bp.route('/search_courses')
@permission_required('Back Course Entry')
def search_courses():
    code = request.args.get('code', '')
    query = "SELECT pk_courseid as id, coursecode + ' - ' + coursename as name FROM SMS_Course_Mst WHERE coursecode LIKE ? ORDER BY coursecode"
    courses = DB.fetch_all(query, [f'%{code}%'])
    return jsonify(courses)

@academics_bp.route('/back_course_entry', methods=['GET', 'POST'])
@permission_required('Back Course Entry')
def back_course_entry():
    if request.method == 'POST':
        flash('Back course entry saved.', 'success')
        return redirect(url_for('academics.back_course_entry'))
    
    lookups = {
        'colleges': AcademicsModel.get_colleges_simple(),
        'degrees': AcademicsModel.get_all_degrees(),
        'sessions': InfrastructureModel.get_sessions()
    }
    return render_template('academics/back_course_entry.html', lookups=lookups)

@academics_bp.route('/college_degree_seat_report', methods=['GET', 'POST'])
@permission_required('College Degree Seat Report')
def college_degree_seat_report():
    filters = {
        'session_id': request.args.get('session_id'),
        'college_id': request.args.get('college_id')
    }

    data, seat_types = [], []
    if all(filters.values()):
        data, seat_types = SeatDetailModel.get_seat_report(filters)

    fmt = request.args.get('fmt')
    if fmt and data:
        if fmt == 'excel':
            import xlsxwriter
            output = io.BytesIO()
            wb = xlsxwriter.Workbook(output, {'in_memory': True})
            ws = wb.add_worksheet('Seat Report')

            # ── Formats ──────────────────────────────────────────────────
            border = {'border': 1, 'border_color': '#AAAAAA'}

            fmt_title = wb.add_format({
                'bold': True, 'font_size': 14, 'align': 'center', 'valign': 'vcenter',
                'font_color': '#1F3864', 'font_name': 'Calibri'
            })
            fmt_sub = wb.add_format({
                'font_size': 10, 'align': 'center', 'valign': 'vcenter',
                'font_color': '#444444', 'font_name': 'Calibri'
            })
            fmt_hdr = wb.add_format({
                'bold': True, 'font_size': 10, 'align': 'center', 'valign': 'vcenter',
                'bg_color': '#2E75B6', 'font_color': '#FFFFFF',
                'font_name': 'Calibri', 'text_wrap': True, **border
            })
            fmt_cell = wb.add_format({
                'font_size': 9, 'align': 'left', 'valign': 'vcenter',
                'font_name': 'Calibri', 'text_wrap': True, **border
            })
            fmt_cell_c = wb.add_format({
                'font_size': 9, 'align': 'center', 'valign': 'vcenter',
                'font_name': 'Calibri', **border
            })
            fmt_cell_alt = wb.add_format({
                'font_size': 9, 'align': 'left', 'valign': 'vcenter',
                'bg_color': '#DEEAF1', 'font_name': 'Calibri', 'text_wrap': True, **border
            })
            fmt_cell_alt_c = wb.add_format({
                'font_size': 9, 'align': 'center', 'valign': 'vcenter',
                'bg_color': '#DEEAF1', 'font_name': 'Calibri', **border
            })
            fmt_total_lbl = wb.add_format({
                'bold': True, 'font_size': 10, 'align': 'right', 'valign': 'vcenter',
                'bg_color': '#1F3864', 'font_color': '#FFFFFF',
                'font_name': 'Calibri', **border
            })
            fmt_total_val = wb.add_format({
                'bold': True, 'font_size': 10, 'align': 'center', 'valign': 'vcenter',
                'bg_color': '#1F3864', 'font_color': '#FFFFFF',
                'font_name': 'Calibri', **border
            })

            session_name = data[0]['sessionname'] if data else ''
            college_name = data[0]['collegename'] if data else ''
            total_cols = 4 + len(seat_types)  # S.No, Degree, Specialization, Total, *seat_types

            # ── Title rows ───────────────────────────────────────────────
            ws.merge_range(0, 0, 0, total_cols - 1,
                           'HAU — College Degree Seat Report', fmt_title)
            ws.merge_range(1, 0, 1, total_cols - 1,
                           f'Session: {session_name}     College: {college_name}', fmt_sub)
            ws.merge_range(2, 0, 2, total_cols - 1,
                           f'Generated: {datetime.now().strftime("%d/%m/%Y %I:%M %p")}', fmt_sub)
            ws.set_row(0, 22)
            ws.set_row(1, 16)
            ws.set_row(2, 14)

            # ── Header row ───────────────────────────────────────────────
            HDR_ROW = 4
            headers = ['S.No.', 'Degree', 'Specialization', 'Total Seats'] + seat_types
            for col, h in enumerate(headers):
                ws.write(HDR_ROW, col, h, fmt_hdr)
            ws.set_row(HDR_ROW, 28)

            # ── Column widths ────────────────────────────────────────────
            ws.set_column(0, 0, 6)    # S.No.
            ws.set_column(1, 1, 30)   # Degree
            ws.set_column(2, 2, 28)   # Specialization
            ws.set_column(3, 3, 12)   # Total Seats
            for c in range(4, 4 + len(seat_types)):
                ws.set_column(c, c, 12)

            # ── Data rows ────────────────────────────────────────────────
            totals = {st: 0 for st in seat_types}
            grand_total = 0
            for i, row in enumerate(data):
                r = HDR_ROW + 1 + i
                alt = (i % 2 == 1)
                fl, fc = (fmt_cell_alt, fmt_cell_alt_c) if alt else (fmt_cell, fmt_cell_c)
                ws.write(r, 0, i + 1,                        fc)
                ws.write(r, 1, row['degreename'] or '',       fl)
                ws.write(r, 2, row['Branchname'] or 'All',    fl)
                ws.write(r, 3, row['totseat'] or 0,           fc)
                grand_total += (row['totseat'] or 0)
                for ci, st in enumerate(seat_types):
                    val = row['seat_types'].get(st, 0)
                    ws.write(r, 4 + ci, val if val else '-',  fc)
                    totals[st] += val if val else 0
                ws.set_row(r, 18)

            # ── Totals row ───────────────────────────────────────────────
            total_row = HDR_ROW + 1 + len(data)
            ws.merge_range(total_row, 0, total_row, 2, 'TOTAL', fmt_total_lbl)
            ws.write(total_row, 3, grand_total, fmt_total_val)
            for ci, st in enumerate(seat_types):
                ws.write(total_row, 4 + ci, totals[st] or '-', fmt_total_val)
            ws.set_row(total_row, 20)

            # ── Freeze panes below header ─────────────────────────────────
            ws.freeze_panes(HDR_ROW + 1, 0)

            wb.close()
            output.seek(0)
            return send_file(output, download_name='College_Degree_Seat_Report.xlsx',
                             as_attachment=True,
                             mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        elif fmt == 'pdf':
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib import colors
            from reportlab.lib.units import cm
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.enums import TA_CENTER, TA_LEFT

            buf = io.BytesIO()
            doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                                    leftMargin=1.5*cm, rightMargin=1.5*cm,
                                    topMargin=1.5*cm, bottomMargin=1.5*cm)
            styles = getSampleStyleSheet()

            # Reusable paragraph styles for table cells (enable word-wrap)
            hdr_style  = ParagraphStyle('hdr',  fontSize=8, leading=10,
                                        textColor=colors.white, fontName='Helvetica-Bold',
                                        alignment=TA_CENTER, wordWrap='CJK')
            cell_c     = ParagraphStyle('cellc', fontSize=8, leading=10,
                                        fontName='Helvetica', alignment=TA_CENTER, wordWrap='CJK')
            cell_l     = ParagraphStyle('celll', fontSize=8, leading=10,
                                        fontName='Helvetica', alignment=TA_LEFT,  wordWrap='CJK')

            def P(text, style): return Paragraph(str(text) if text is not None else '', style)

            story = []

            # Header: logo + title
            logo_path = os.path.join(current_app.root_path, 'static', 'images', 'logo.png')
            if os.path.exists(logo_path):
                logo = Image(logo_path, width=2.5*cm, height=2.5*cm)
                title_para = Paragraph(
                    '<b>HAU - College Degree Seat Report</b><br/>'
                    f'<font size=10>Session: {data[0]["sessionname"] if data else ""} &nbsp;&nbsp; '
                    f'College: {data[0]["collegename"] if data else ""}</font><br/>'
                    f'<font size=8>Generated: {datetime.now().strftime("%d/%m/%Y %I:%M %p")}</font>',
                    ParagraphStyle('titl', fontSize=14, leading=20, alignment=TA_CENTER)
                )
                header_tbl = Table([[logo, title_para]], colWidths=[3*cm, None])
                header_tbl.setStyle(TableStyle([
                    ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                    ('ALIGN',  (1,0), (1,0),   'CENTER'),
                ]))
                story.append(header_tbl)
            else:
                story.append(Paragraph('<b>HAU - College Degree Seat Report</b>', styles['Title']))
            story.append(Spacer(1, 0.4*cm))

            # Column widths
            fixed = [1.2*cm, 5*cm, 5*cm, 2*cm]
            remaining = 28*cm - sum(fixed)
            st_width = max(1.8*cm, (remaining / len(seat_types))) if seat_types else 2*cm
            col_widths = fixed + [st_width] * len(seat_types)

            # Header row — all cells are Paragraphs so long headers wrap too
            header_row = [
                P('S.No.',         hdr_style),
                P('Degree',        hdr_style),
                P('Specialization',hdr_style),
                P('Total Seats',   hdr_style),
            ] + [P(st, hdr_style) for st in seat_types]

            table_data = [header_row]
            for i, row in enumerate(data, 1):
                r = [
                    P(i,                          cell_c),
                    P(row['degreename'],           cell_l),
                    P(row['Branchname'] or 'All',  cell_l),
                    P(row['totseat'],              cell_c),
                ]
                for st in seat_types:
                    r.append(P(row['seat_types'].get(st, '-'), cell_c))
                table_data.append(r)

            tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
            tbl.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0),   colors.HexColor('#3c8dbc')),
                ('ROWBACKGROUNDS', (0,1), (-1,-1),
                                            [colors.white, colors.HexColor('#f0f7ff')]),
                ('GRID',       (0,0), (-1,-1),  0.4, colors.grey),
                ('VALIGN',     (0,0), (-1,-1),  'MIDDLE'),
                ('TOPPADDING', (0,0), (-1,-1),  4),
                ('BOTTOMPADDING', (0,0), (-1,-1), 4),
                ('LEFTPADDING',   (0,0), (-1,-1), 4),
                ('RIGHTPADDING',  (0,0), (-1,-1), 4),
            ]))
            story.append(tbl)

            doc.build(story)
            buf.seek(0)
            response = make_response(buf.read())
            response.headers['Content-Type'] = 'application/pdf'
            response.headers['Content-Disposition'] = 'inline; filename=College_Degree_Seat_Report.pdf'
            return response

    lookups = {
        'sessions': InfrastructureModel.get_sessions(),
        'colleges': AcademicsModel.get_colleges_simple()
    }
    return render_template('academics/college_degree_seat_report.html',
                           lookups=lookups, filters=filters,
                           data=data, seat_types=seat_types)

@academics_bp.route('/student_biodata_updation', methods=['GET', 'POST'])
@permission_required('Student Biodata Updation')
def student_biodata_updation():
    filters = {
        'college_id': request.args.get('college_id'),
        'session_id': request.args.get('session_id'),
        'degree_id': request.args.get('degree_id'),
        'admission_no': request.args.get('admission_no')
    }
    
    students = []
    if any(filters.values()):
        students = StudentModel.get_students_by_filter(filters)
        
    lookups = {
        'colleges': AcademicsModel.get_colleges_simple(),
        'sessions': InfrastructureModel.get_sessions(),
        'degrees': []
    }
    if filters['college_id']:
        lookups['degrees'] = AcademicsModel.get_college_pg_degrees(filters['college_id'])
    else:
        lookups['degrees'] = AcademicsModel.get_all_degrees()

    return render_template('academics/student_biodata_updation.html', 
                           lookups=lookups, filters=filters, students=students)

@academics_bp.route('/achievement_disciplinary', methods=['GET', 'POST'])
@permission_required('Achievement/ Disciplinary ')
def achievement_disciplinary():
    # This page usually searches by enrollment/admission no
    enrollment_no = request.args.get('enrollment_no')
    student = None
    if enrollment_no:
        student = StudentModel.get_student_by_enrollment(enrollment_no)
        
    return render_template('academics/achievement_disciplinary.html', student=student)

@academics_bp.route('/achievement_disciplinary_approval', methods=['GET', 'POST'])
@permission_required('Achievement / Disciplinary Approval')
def achievement_disciplinary_approval():
    # List based approval
    filters = {
        'college_id': request.args.get('college_id'),
        'session_id': request.args.get('session_id'),
        'degree_id': request.args.get('degree_id')
    }
    lookups = {
        'colleges': AcademicsModel.get_colleges_simple(),
        'sessions': InfrastructureModel.get_sessions(),
        'degrees': AcademicsModel.get_all_degrees()
    }
    return render_template('academics/achievement_disciplinary_approval.html', 
                           lookups=lookups, filters=filters)

@academics_bp.route('/student_activity_management', methods=['GET', 'POST'])
@permission_required('Student Activity Management')
def student_activity_management():
    # Placeholder for student activity management
    return render_template('academics/student_activity_management.html')

@academics_bp.route('/student_personal_detail_report', methods=['GET', 'POST'])
@permission_required('Student Personal Detail Report')
def student_personal_detail_report():
    lookups = {
        'colleges': AcademicsModel.get_colleges_simple(),
        'sessions': InfrastructureModel.get_sessions(),
        'degrees': AcademicsModel.get_all_degrees(),
        'semesters': InfrastructureModel.get_all_semesters()
    }

    if request.method == 'POST':
        filters = {
            'college_id': request.form.get('college_id', '0'),
            'session_id': request.form.get('session_id', '0'),
            'degree_id': request.form.get('degree_id', '0'),
            'semester_id': request.form.get('semester_id', '0'),
            'branch_id': request.form.get('branch_id', '0'),
            'rpt_type': request.form.get('rpt_type', '1'),
            'rpt_format': request.form.get('rpt_format', 'PDF'),
        }

        def fmt_date(val):
            if not val:
                return ''
            try:
                if hasattr(val, 'strftime'):
                    return val.strftime('%d/%m/%Y')
                return str(val)[:10]
            except Exception:
                return str(val) if val else ''

        try:
            data = StudentModel.get_student_personal_detail_report(filters)
        except Exception as e:
            import traceback; traceback.print_exc()
            flash(f'Error fetching data: {e}', 'danger')
            return render_template('academics/student_personal_detail_report.html', lookups=lookups)

        if not data:
            flash('No records found for the selected filters.', 'warning')
            return render_template('academics/student_personal_detail_report.html', lookups=lookups)

        # ========== EXCEL ==========
        if filters['rpt_format'] == 'Excel':
            try:
                from openpyxl import Workbook
                from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
                from openpyxl.utils import get_column_letter

                wb = Workbook()
                ws = wb.active
                ws.title = 'Student Personal Detail'

                navy   = PatternFill('solid', fgColor='1F3864')
                blue   = PatternFill('solid', fgColor='2E75B6')
                pale   = PatternFill('solid', fgColor='D9E1F2')
                white  = PatternFill('solid', fgColor='FFFFFF')
                alt    = PatternFill('solid', fgColor='EBF3FF')
                bold_white    = Font(name='Calibri', bold=True, color='FFFFFF', size=13)
                bold_white_sm = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
                bold_dark     = Font(name='Calibri', bold=True, color='1F3864', size=10)
                normal        = Font(name='Calibri', size=10)
                center        = Alignment(horizontal='center', vertical='center', wrap_text=True)
                left          = Alignment(horizontal='left',   vertical='center', wrap_text=True)
                thin          = Side(style='thin',   color='8EA9C1')
                thick         = Side(style='medium', color='1F3864')
                thin_border   = Border(left=thin, right=thin, top=thin, bottom=thin)

                num_cols = 22
                last_col = get_column_letter(num_cols)

                ws.append(['CHAUDHARY CHARAN SINGH HARYANA AGRICULTURAL UNIVERSITY, HISAR'])
                ws.merge_cells(f'A1:{last_col}1')
                ws['A1'].fill = navy
                ws['A1'].font = bold_white
                ws['A1'].alignment = center
                ws.row_dimensions[1].height = 28

                ws.append(['Student Personal Detail Report'])
                ws.merge_cells(f'A2:{last_col}2')
                ws['A2'].fill = blue
                ws['A2'].font = bold_white_sm
                ws['A2'].alignment = center
                ws.row_dimensions[2].height = 20

                ws.append([f'Generated on: {datetime.now().strftime("%d/%m/%Y %H:%M")}'])
                ws.merge_cells(f'A3:{last_col}3')
                ws['A3'].fill = pale
                ws['A3'].font = Font(name='Calibri', italic=True, size=9, color='1F3864')
                ws['A3'].alignment = center
                ws.row_dimensions[3].height = 16

                headers = [
                    'S.No.', 'Admission No', 'Enrollment No', 'Full Name', "Father's Name",
                    "Mother's Name", 'Date of Birth', 'Date of Admission', 'Gender',
                    'Qualifying Exam', 'Board / University', 'Postal Address',
                    'College', 'Degree', 'Specialization', 'Class', 'Session', 'Seat Type',
                    'Category', 'Phone No', 'Is Resident', 'Reg. Status'
                ]
                col_widths = [6, 14, 14, 25, 22, 22, 14, 16, 9, 25, 28, 35,
                              30, 25, 25, 8, 18, 14, 14, 14, 11, 12]
                ws.append(headers)
                for ci, (h, w) in enumerate(zip(headers, col_widths), start=1):
                    cell = ws.cell(row=4, column=ci)
                    cell.fill = pale
                    cell.font = bold_dark
                    cell.alignment = center
                    cell.border = Border(left=thick, right=thick, top=thick, bottom=thick)
                    ws.column_dimensions[get_column_letter(ci)].width = w
                ws.row_dimensions[4].height = 28
                ws.auto_filter.ref = f'A4:{last_col}4'
                ws.freeze_panes = 'A5'

                center_cols = {1, 7, 8, 9, 16, 20, 21, 22}
                for idx, row in enumerate(data, start=1):
                    fill = white if idx % 2 == 0 else alt
                    values = [
                        idx,
                        row.get('AdmissionNo', ''),
                        row.get('enrollmentno', ''),
                        row.get('fullname', ''),
                        row.get('fname', ''),
                        row.get('mname', ''),
                        fmt_date(row.get('dob')),
                        fmt_date(row.get('DateOfAdmission')),
                        row.get('gender', ''),
                        row.get('qual_exam', ''),
                        row.get('board_university', ''),
                        row.get('address', ''),
                        row.get('collegename', ''),
                        row.get('degreename', ''),
                        row.get('specialization', ''),
                        row.get('semester_roman', ''),
                        row.get('sessionname', ''),
                        row.get('seatype', ''),
                        row.get('category', ''),
                        row.get('phoneno', ''),
                        'YES' if row.get('is_resident') else 'NO',
                        row.get('reg_status', ''),
                    ]
                    ws.append(values)
                    r = 4 + idx
                    for ci in range(1, num_cols + 1):
                        cell = ws.cell(row=r, column=ci)
                        cell.fill = fill
                        cell.font = normal
                        cell.alignment = center if ci in center_cols else left
                        cell.border = thin_border

                footer_row = 5 + len(data)
                ws.append([f'Total Records: {len(data)}  |  HAU ERP System'])
                ws.merge_cells(f'A{footer_row}:{last_col}{footer_row}')
                ws[f'A{footer_row}'].fill = pale
                ws[f'A{footer_row}'].font = Font(name='Calibri', bold=True, size=9, color='1F3864')
                ws[f'A{footer_row}'].alignment = center
                ws.sheet_properties.pageSetUpPr.fitToPage = True
                ws.page_setup.orientation = 'landscape'
                ws.page_setup.fitToWidth = 1

                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                return send_file(output,
                                 download_name='Student_Personal_Detail_Report.xlsx',
                                 as_attachment=True,
                                 mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            except Exception as e:
                import traceback; traceback.print_exc()
                flash(f'Excel error: {e}', 'danger')
                return render_template('academics/student_personal_detail_report.html', lookups=lookups)

        # ========== PDF ==========
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch, mm

            navy_color = colors.HexColor('#1F3864')
            blue_color = colors.HexColor('#2E75B6')
            pale_color = colors.HexColor('#D9E1F2')
            alt_color  = colors.HexColor('#EBF3FF')

            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4,
                                    rightMargin=18*mm, leftMargin=18*mm,
                                    topMargin=20*mm, bottomMargin=25*mm)
            elements = []
            styles = getSampleStyleSheet()

            title_style   = ParagraphStyle('T', parent=styles['Normal'],
                                           fontName='Helvetica-Bold', fontSize=12, alignment=1,
                                           textColor=colors.white, spaceAfter=2)
            label_style   = ParagraphStyle('L', parent=styles['Normal'],
                                           fontName='Helvetica-Bold', fontSize=9,
                                           textColor=navy_color)
            value_style   = ParagraphStyle('V', parent=styles['Normal'],
                                           fontName='Helvetica', fontSize=9,
                                           textColor=colors.black)
            heading_style = ParagraphStyle('H', parent=styles['Normal'],
                                           fontName='Helvetica-Bold', fontSize=10,
                                           textColor=navy_color, alignment=1, spaceAfter=4)

            logo_path = os.path.join(current_app.root_path, 'static', 'images', 'logo.png')

            def page_footer(canvas, doc):
                canvas.saveState()
                canvas.setFont('Helvetica', 8)
                canvas.setFillColor(colors.grey)
                canvas.drawCentredString(A4[0] / 2.0, 12*mm,
                                         f'Page {canvas.getPageNumber()}  |  HAU ERP System  |  Confidential')
                canvas.restoreState()

            if os.path.exists(logo_path):
                img = Image(logo_path, width=0.75*inch, height=0.75*inch)
                hdr_data = [[img, Paragraph('CHAUDHARY CHARAN SINGH<br/>HARYANA AGRICULTURAL UNIVERSITY, HISAR', title_style)]]
            else:
                hdr_data = [['', Paragraph('CHAUDHARY CHARAN SINGH<br/>HARYANA AGRICULTURAL UNIVERSITY, HISAR', title_style)]]

            hdr_table = Table(hdr_data, colWidths=[0.9*inch, 4.8*inch])
            hdr_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), navy_color),
                ('VALIGN',     (0, 0), (-1, -1), 'MIDDLE'),
                ('ALIGN',      (0, 0), (0,  0),  'CENTER'),
                ('TOPPADDING',    (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ]))
            elements.append(hdr_table)
            elements.append(Spacer(1, 6))
            elements.append(Paragraph('STUDENT PERSONAL DETAIL REPORT', heading_style))
            elements.append(Spacer(1, 4))

            pw = doc.width

            def lv(label, value):
                return [Paragraph(label, label_style),
                        Paragraph(str(value) if value else '---', value_style)]

            for i, row in enumerate(data):
                dob_str  = fmt_date(row.get('dob'))
                doa_str  = fmt_date(row.get('DateOfAdmission'))
                resident = 'YES' if row.get('is_resident') else 'NO'
                adm_no   = row.get('AdmissionNo') or row.get('enrollmentno') or ''

                card_hdr = Table(
                    [[Paragraph(f"Admission No: {adm_no}", label_style),
                      Paragraph(f"Status: {row.get('reg_status', '')}", label_style)]],
                    colWidths=[pw * 0.5, pw * 0.5]
                )
                card_hdr.setStyle(TableStyle([
                    ('BACKGROUND',    (0, 0), (-1, -1), pale_color),
                    ('TOPPADDING',    (0, 0), (-1, -1), 4),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                    ('LEFTPADDING',   (0, 0), (-1, -1), 6),
                    ('RIGHTPADDING',  (0, 0), (-1, -1), 6),
                    ('LINEBELOW',     (0, 0), (-1, -1), 0.8, blue_color),
                ]))
                elements.append(card_hdr)

                left_col  = pw * 0.22
                right_col = pw * 0.28

                detail_data = [
                    lv('Full Name (Block Letters):', row.get('fullname', '')) +
                    lv('College:', row.get('collegename', '')),
                    lv("Father's Name:", row.get('fname', '')) +
                    lv('Degree:', row.get('degreename', '')),
                    lv("Mother's Name:", row.get('mname', '')) +
                    lv('Specialization:', row.get('specialization', '')),
                    lv('Date of Birth:', dob_str) +
                    lv('Class (Semester):', row.get('semester_roman', '')),
                    lv('Date of Admission:', doa_str) +
                    lv('Session:', row.get('sessionname', '')),
                    lv('Gender:', row.get('gender', '')) +
                    lv('Seat Type:', row.get('seatype', '')),
                    lv('Qualifying Examination:', row.get('qual_exam', '')) +
                    lv('Category:', row.get('category', '')),
                    lv('Board / University:', row.get('board_university', '')) +
                    lv('Phone No:', row.get('phoneno', '')),
                    lv('Is Resident:', resident) +
                    lv('Postal Address:', row.get('address', '')),
                ]

                card_tbl = Table(detail_data, colWidths=[left_col, right_col, left_col, right_col])
                card_tbl.setStyle(TableStyle([
                    ('FONTNAME',       (0, 0), (-1, -1), 'Helvetica'),
                    ('FONTSIZE',       (0, 0), (-1, -1), 9),
                    ('VALIGN',         (0, 0), (-1, -1), 'TOP'),
                    ('TOPPADDING',     (0, 0), (-1, -1), 3),
                    ('BOTTOMPADDING',  (0, 0), (-1, -1), 3),
                    ('LEFTPADDING',    (0, 0), (-1, -1), 5),
                    ('RIGHTPADDING',   (0, 0), (-1, -1), 5),
                    ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, alt_color]),
                    ('GRID',           (0, 0), (-1, -1), 0.4, colors.HexColor('#B8CCE4')),
                    ('LINEBELOW',      (0, -1), (-1, -1), 1.2, navy_color),
                ]))
                elements.append(card_tbl)
                elements.append(Spacer(1, 8))

                if (i + 1) % 3 == 0 and i + 1 < len(data):
                    elements.append(PageBreak())

            doc.build(elements, onFirstPage=page_footer, onLaterPages=page_footer)
            buffer.seek(0)
            response = make_response(buffer.getvalue())
            response.headers['Content-Type'] = 'application/pdf'
            response.headers['Content-Disposition'] = 'inline; filename=Student_Personal_Detail_Report.pdf'
            return response
        except Exception as e:
            import traceback; traceback.print_exc()
            flash(f'PDF error: {e}', 'danger')
            return render_template('academics/student_personal_detail_report.html', lookups=lookups)

    return render_template('academics/student_personal_detail_report.html', lookups=lookups)

@academics_bp.route('/moderation_marks', methods=['GET', 'POST'])
@permission_required('Moderation Marks Detail')
def moderation_marks():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'DELETE':
            if AcademicsModel.delete_moderation_marks(request.form.get('id')):
                flash('Moderation marks deleted successfully!', 'success')
            else:
                flash('Error deleting record.', 'danger')
        else:
            if AcademicsModel.save_moderation_marks(request.form):
                flash('Moderation marks saved successfully!', 'success')
            else:
                flash('Error saving record. Make sure a Degree Cycle exists for the selected degree/semester.', 'danger')
        return redirect(url_for('academics.moderation_marks'))
    
    page = request.args.get('page', 1, type=int)
    per_page = 10
    items, total = AcademicsModel.get_moderation_marks(page=page, per_page=per_page)
    
    pagination = {
        'page': page,
        'per_page': per_page,
        'total': total,
        'total_pages': math.ceil(total / per_page) if total else 1,
        'has_prev': page > 1,
        'has_next': page < (math.ceil(total / per_page) if total else 1)
    }
    
    page_range = get_pagination_range(page, pagination['total_pages'])

    lookups = {
        'degrees': AcademicsModel.get_all_degrees(),
        'semesters': InfrastructureModel.get_all_semesters(),
        'titles': CourseModel.get_all_paper_titles()
    }
    
    return render_template('academics/moderation_marks.html', items=items, lookups=lookups, pagination=pagination, page_range=page_range)

@academics_bp.route('/limit_assignment', methods=['GET', 'POST'])
@permission_required('Limit Assignment')
def limit_assignment():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'DELETE':
            # Simplified delete
            DB.execute("DELETE FROM SMS_AdvisoryStudentLimitConfiq WHERE pk_limitid = ?", [request.form.get('id')])
            flash('Limit assignment deleted successfully!', 'success')
        else:
            if AdvisoryModel.save_student_limit(request.form):
                flash('Limit assignment saved successfully!', 'success')
            else:
                flash('Error saving limit assignment.', 'danger')
        return redirect(url_for('academics.limit_assignment'))

    page = request.args.get('page', 1, type=int)
    per_page = 10
    items, total = AdvisoryModel.get_student_limits(page=page, per_page=per_page)
    
    pagination = {
        'page': page, 'per_page': per_page, 'total': total,
        'total_pages': math.ceil(total / per_page) if total else 1,
        'has_prev': page > 1,
        'has_next': page < (math.ceil(total / per_page) if total else 1)
    }
    page_range = get_pagination_range(page, pagination['total_pages'])

    lookups = {
        'colleges': AcademicsModel.get_colleges_simple(),
        'sessions': InfrastructureModel.get_sessions(),
        'degrees': AcademicsModel.get_all_degrees(),
        'branches': AcademicsModel.get_branches()
    }
    filters = {
        'college_id': request.args.get('college_id', ''),
        'session_id': request.args.get('session_id', ''),
        'degree_id': request.args.get('degree_id', '')
    }
    return render_template(
        'academics/limit_assignment.html',
        items=items,
        lookups=lookups,
        pagination=pagination,
        page_range=page_range,
        filters=filters
    )

@academics_bp.route('/major_advisor', methods=['GET', 'POST'])
@permission_required('Major Advisor')
def major_advisor():
    user_id = session.get('user_id')
    
    if request.method == 'POST':
        action = request.form.get('action')
        sid = request.form.get('sid')
        
        if action == 'delete' and sid:
            success, msg = AdvisoryModel.delete_major_advisor(sid)
            if success:
                flash(msg, 'success')
            else:
                flash(msg, 'danger')
            return redirect(url_for('academics.major_advisor', **request.args))
            
        advisor_id = request.form.get('advisor_id')
        if sid and advisor_id:
            if AdvisoryModel.save_major_advisor(sid, advisor_id, user_id):
                flash('Major Advisor assigned successfully.', 'success')
            else:
                flash('Error assigning Major Advisor.', 'danger')
        return redirect(url_for('academics.major_advisor', **request.args))

    # Filters for Student Selection
    college_id = request.args.get('college_id')
    loc_id = session.get('selected_loc')
    colleges = DB.fetch_all("SELECT pk_collegeid as id, collegename as name FROM SMS_College_Mst WHERE fk_locid = ? ORDER BY collegename", [loc_id]) if loc_id else AcademicsModel.get_colleges_simple()
    
    if not college_id and colleges:
        college_id = str(colleges[0]['id'])
        
    session_id = request.args.get('session_id')
    if not session_id:
        curr_session = InfrastructureModel.get_current_session_id()
        session_id = str(curr_session) if curr_session else None
        
    degree_id = request.args.get('degree_id')
    branch_id = request.args.get('branch_id')
    
    # Filters for Bottom Grid
    page = request.args.get('page', 1, type=int)
    college_id_f = request.args.get('college_id_f', college_id)
    session_id_f = request.args.get('session_id_f', session_id)
    degree_id_f = request.args.get('degree_id_f')
    branch_id_f = request.args.get('branch_id_f')
    admission_no_f = request.args.get('admission_no_f')
    status_f = request.args.get('status_f', '-1')

    students = []
    if college_id and session_id and degree_id:
        filters = {
            'college_id': college_id,
            'session_id': session_id,
            'degree_id': degree_id,
            'branch_id': branch_id
        }
        students = DB.fetch_all(
            "SELECT pk_sid, fullname, AdmissionNo, enrollmentno FROM SMS_Student_Mst "
            "WHERE fk_collegeid=? AND fk_adm_session=? AND fk_degreeid=? "
            "AND (fk_branchid=? OR ?='') ORDER BY fullname",
            [college_id, session_id, degree_id, branch_id or '', branch_id or '']
        )

    # Bottom Grid Students
    filters_f = {
        'college_id': college_id_f,
        'session_id': session_id_f,
        'degree_id': degree_id_f,
        'branch_id': branch_id_f,
        'admission_no': admission_no_f,
        'status': status_f
    }
    per_page = 10
    grid_students, total_f = AdvisoryModel.get_students_for_advisory(filters_f, page=page, per_page=per_page)
    
    import math
    pagination = {
        'page': page,
        'per_page': per_page,
        'total': total_f,
        'total_pages': math.ceil(total_f / per_page) if total_f else 1,
        'has_prev': page > 1,
        'has_next': page < (math.ceil(total_f / per_page) if total_f else 1)
    }

    # For auto-filling the advisor dropdown if edited
    edit_sid = request.args.get('edit_sid')
    edit_advisor = None
    if edit_sid:
        conn = DB.get_connection()
        try:
            mst = conn.execute("SELECT pk_adcid FROM SMS_Advisory_Committee_Mst WHERE fk_stid=?", [edit_sid]).fetchone()
            if mst:
                dtl = conn.execute("SELECT D.fk_empid, E.empname, E.empcode FROM SMS_Advisory_Committee_Dtl D JOIN SAL_Employee_Mst E ON D.fk_empid = E.pk_empid WHERE D.fk_adcid=? AND D.fk_statusid=1", [mst[0]]).fetchone()
                if dtl:
                    edit_advisor = {'id': dtl[0], 'name': f"{dtl[1]} || {dtl[2]}"}
        finally:
            conn.close()

    employees = []
    if branch_id and branch_id != '0' and branch_id != 'None':
        employees = DB.fetch_all("""
            SELECT E.pk_empid as id, E.empname + ' || ' + ISNULL(E.empcode, '') + ' (' + ISNULL(D.description, 'No Dept') + ')' as name 
            FROM SAL_Employee_Mst E 
            LEFT JOIN Department_Mst D ON E.fk_deptid = D.pk_deptid 
            JOIN SMS_BranchMst B ON E.fk_deptid = B.fk_deptidDdo 
            WHERE E.employeeleftstatus = 'N' AND B.Pk_BranchId = ? 
            ORDER BY E.empname
        """, [branch_id])
    else:
        employees = DB.fetch_all("""
            SELECT E.pk_empid as id, E.empname + ' || ' + ISNULL(E.empcode, '') + ' (' + ISNULL(D.description, 'No Dept') + ')' as name 
            FROM SAL_Employee_Mst E 
            LEFT JOIN Department_Mst D ON E.fk_deptid = D.pk_deptid 
            WHERE E.employeeleftstatus = 'N' 
            ORDER BY E.empname
        """)

    lookups = {
        'colleges': colleges,
        'sessions': InfrastructureModel.get_sessions(),
        'degrees': AcademicsModel.get_college_pg_degrees(college_id) if college_id else [],
        'branches': AcademicsModel.get_college_degree_specializations(college_id, degree_id) if (college_id and degree_id) else [],
        'degrees_f': AcademicsModel.get_college_pg_degrees(college_id_f) if college_id_f else [],
        'branches_f': AcademicsModel.get_college_degree_specializations(college_id_f, degree_id_f) if (college_id_f and degree_id_f) else [],
        'employees': employees
    }
    
    active_filters = {
        'college_id': college_id,
        'session_id': session_id,
        'degree_id': degree_id,
        'branch_id': branch_id,
        'college_id_f': college_id_f,
        'session_id_f': session_id_f,
        'degree_id_f': degree_id_f,
        'branch_id_f': branch_id_f,
        'admission_no_f': admission_no_f,
        'status_f': status_f
    }
    
    return render_template('academics/major_advisor.html', 
                           lookups=lookups,
                           students=clean_json_data(students),
                           grid_students=clean_json_data(grid_students),
                           pagination=pagination,
                           edit_advisor=edit_advisor,
                           edit_sid=edit_sid,
                           filters=active_filters)

@academics_bp.route('/dean_pgs_approval', methods=['GET', 'POST'])
@permission_required('Dean PGS approval (advisory committee)')
def dean_pgs_approval():
    if request.method == 'POST':
        action = request.form.get('action')
        user_id = session['user_id']

        if action == 'ADD_NOMINEE':
            adcid = request.form.get('adcid')
            advisor_id = request.form.get('advisor_id')
            if adcid and advisor_id:
                AdvisoryModel.save_nominee(adcid, advisor_id)
                flash('Nominee added successfully.', 'success')
            return redirect(url_for('academics.dean_pgs_approval', **request.args))

        elif action in ['APPROVE', 'REJECT']:
            adcid = request.form.get('adcid')
            status = 'A' if action == 'APPROVE' else 'R'
            remarks = request.form.get('remarks')
            if adcid:
                AdvisoryModel.approve_advisory(adcid, status, user_id, level='PGS', remarks=remarks)
                flash(f'Advisory record { "approved" if status == "A" else "rejected" } successfully.', 'success')
            return redirect(url_for('academics.dean_pgs_approval', **request.args))

    college_id = request.args.get('college_id')
    loc_id = session.get('selected_loc')
    colleges = DB.fetch_all("SELECT pk_collegeid as id, collegename as name FROM SMS_College_Mst WHERE fk_locid = ? ORDER BY collegename", [loc_id]) if loc_id else AcademicsModel.get_colleges_simple()

    if not college_id and colleges:
        college_id = str(colleges[0]['id'])

    session_id = request.args.get('session_id')
    if not session_id:
        curr_session = InfrastructureModel.get_current_session_id()
        session_id = str(curr_session) if curr_session else None

    degree_id = request.args.get('degree_id')
    branch_id = request.args.get('branch_id')
    page = request.args.get('page', 1, type=int)
    per_page = 10

    students = []
    total_students = 0
    if college_id:
        filters = {
            'college_id': college_id,
            'session_id': session_id,
            'degree_id': degree_id,
            'branch_id': branch_id
        }
        students, total_students = AdvisoryModel.get_students_for_advisory(filters, page=page, per_page=per_page)

    import math
    pagination = {
        'page': page,
        'per_page': per_page,
        'total': total_students,
        'total_pages': math.ceil(total_students / per_page) if total_students else 1,
        'has_prev': page > 1,
        'has_next': page < (math.ceil(total_students / per_page) if total_students else 1)
    }

    sid = request.args.get('sid')
    advisory_details = []
    current_adcid = None

    if sid:
        advisory_info = AdvisoryModel.get_student_advisory_committee(sid)
        if advisory_info:
            advisory_details = advisory_info.get('details', [])
            current_adcid = advisory_info.get('adcid')

    lookups = {
        'colleges': colleges,
        'sessions': InfrastructureModel.get_sessions(),
        'degrees': AcademicsModel.get_college_pg_degrees(college_id) if college_id else [],
        'branches': AcademicsModel.get_college_degree_specializations(college_id, degree_id) if (college_id and degree_id and str(degree_id) != '0') else [],
        'employees': DB.fetch_all("SELECT E.pk_empid as id, E.empname + ' || ' + ISNULL(E.empcode, '') + ' (' + ISNULL(D.description, 'No Dept') + ')' as name FROM SAL_Employee_Mst E LEFT JOIN Department_Mst D ON E.fk_deptid = D.pk_deptid WHERE E.employeeleftstatus = 'N' ORDER BY E.empname"),
        'advisory_types': [{'id': 5, 'name': 'Dean PGS Nominee'}]
    }

    active_filters = {
        'college_id': college_id,
        'session_id': session_id,
        'degree_id': degree_id,
        'branch_id': branch_id
    }

    return render_template('academics/dean_pgs_approval.html', 
                           lookups=lookups, 
                           filters=active_filters, 
                           students=clean_json_data(students), 
                           pagination=pagination,
                           advisory_details=advisory_details,
                           current_adcid=current_adcid,
                           sid=sid)
@academics_bp.route('/hod_approval', methods=['GET', 'POST'])
@permission_required('HOD Approval')
def hod_approval():
    if request.method == 'POST':
        action = request.form.get('action')
        user_id = session['user_id']
        adcid = request.form.get('adcid')
        
        if action in ['APPROVE', 'REJECT']:
            status = 'A' if action == 'APPROVE' else 'R'
            remarks = request.form.get('remarks')
            if adcid:
                AdvisoryModel.approve_advisory(adcid, status, user_id, level='HOD', remarks=remarks)
                flash(f'Advisory record { "approved" if status == "A" else "rejected" } by HOD successfully.', 'success')
            return redirect(url_for('academics.hod_approval', **request.args))

    filters = {
        'college_id': request.args.get('college_id'),
        'session_id': request.args.get('session_id'),
        'degree_id': request.args.get('degree_id'),
        'branch_id': request.args.get('branch_id')
    }
    
    sid = request.args.get('sid')
    students = []
    advisory_details = []
    current_adcid = None
    
    if all([filters['college_id'], filters['session_id'], filters['degree_id']]):
        students = AdvisoryModel.get_students_for_advisory(filters)
        
    if sid:
        advisory_info = AdvisoryModel.get_student_advisory_committee(sid)
        if advisory_info:
            advisory_details = advisory_info.get('details', [])
            current_adcid = advisory_info.get('adcid')

    lookups = AdvisoryModel.get_advisory_lookups(filters['college_id'], filters['degree_id'])
    
    return render_template('academics/hod_approval.html', 
                           lookups=lookups, 
                           filters=filters, 
                           students=students, 
                           advisory_details=advisory_details,
                           current_adcid=current_adcid,
                           sid=sid)

@academics_bp.route('/college_dean_approval', methods=['GET', 'POST'])
@permission_required('College Dean Approval')
def college_dean_approval():
    if request.method == 'POST':
        action = request.form.get('action')
        user_id = session['user_id']
        adcid = request.form.get('adcid')
        
        if action in ['APPROVE', 'REJECT']:
            status = 'A' if action == 'APPROVE' else 'R'
            remarks = request.form.get('remarks')
            if adcid:
                AdvisoryModel.approve_advisory(adcid, status, user_id, level='DEAN', remarks=remarks)
                flash(f'Advisory record { "approved" if status == "A" else "rejected" } by College Dean successfully.', 'success')
            return redirect(url_for('academics.college_dean_approval', **request.args))

    filters = {
        'college_id': request.args.get('college_id'),
        'session_id': request.args.get('session_id'),
        'degree_id': request.args.get('degree_id'),
        'branch_id': request.args.get('branch_id')
    }
    
    sid = request.args.get('sid')
    students = []
    advisory_details = []
    current_adcid = None
    
    if all([filters['college_id'], filters['session_id'], filters['degree_id']]):
        students = AdvisoryModel.get_students_for_advisory(filters)
        
    if sid:
        advisory_info = AdvisoryModel.get_student_advisory_committee(sid)
        if advisory_info:
            advisory_details = advisory_info.get('details', [])
            current_adcid = advisory_info.get('adcid')

    lookups = AdvisoryModel.get_advisory_lookups(filters['college_id'], filters['degree_id'])
    
    return render_template('academics/college_dean_approval.html', 
                           lookups=lookups, 
                           filters=filters, 
                           students=students, 
                           advisory_details=advisory_details,
                           current_adcid=current_adcid,
                           sid=sid)

@academics_bp.app_context_processor
def inject_academic_menu():
    if 'user_id' not in session:
        return dict(ACADEMIC_MENU_CONFIG=ACADEMIC_MENU_CONFIG, academic_tabs=[], academics_breadcrumb=[])

    curr_path = request.path.rstrip('/').lower()

    # Get user permissions for tab filtering
    from app.models import NavModel as _NavModel
    _user_id = session.get('user_id')
    _is_super = _NavModel._is_super_admin(_user_id)
    _all_rights = session.get('current_user_rights') or []
    _allowed_norm = {' '.join(str(r.get('PageName') or '').strip().lower().split())
                     for r in _all_rights if r.get('AllowView')}

    def _n(s):
        return ' '.join(str(s or '').strip().lower().split())

    academic_tabs = []
    academics_breadcrumb = []

    # Scan ALL groups; last matching group with visible tabs wins.
    # This ensures the "Admission" group (which comes after "Master and Config"
    # in the config) takes priority when both share the same page URLs.
    for main_cat, subs in ACADEMIC_MENU_CONFIG.items():
        for sub_cat, sub_subs in subs.items():
            for folder_name, pages in sub_subs.items():
                if not pages:
                    continue

                is_active_group = False
                tab_list = []
                for p_name in pages:
                    p_url = get_page_url(p_name).rstrip('/')
                    is_current = (curr_path == p_url.lower())

                    if _is_super or _n(p_name) in _allowed_norm:
                        tab_list.append({'name': p_name, 'url': p_url, 'active': is_current})

                    if is_current:
                        is_active_group = True
                        academics_breadcrumb = [main_cat, folder_name, p_name]

                # Only override when the new group has at least one visible tab
                if is_active_group and tab_list:
                    academic_tabs = tab_list

    return dict(ACADEMIC_MENU_CONFIG=ACADEMIC_MENU_CONFIG, academic_tabs=academic_tabs, academics_breadcrumb=academics_breadcrumb)

@academics_bp.route('/api/courses')
def get_all_courses_api():
    courses = CourseModel.get_all_courses()
    return jsonify(clean_json_data(courses))

@academics_bp.route('/api/student/<int:sid>/course_plan_data')
def get_student_course_plan_data_api(sid):
    current_plan = AdvisoryModel.get_student_course_plan(sid)
    available_courses = AdvisoryModel.get_available_courses(sid)
    credit_load = AdvisoryModel.get_credit_load(sid)
    required_load = AdvisoryModel.get_required_credit_load(sid)
    
    return jsonify(clean_json_data({
        'current_plan': current_plan,
        'available_courses': available_courses,
        'credit_load': credit_load,
        'required_load': required_load
    }))

@academics_bp.route('/course_plan', methods=['GET', 'POST'])
@permission_required('Prepare Course Plan')
def course_plan():
    if request.method == 'POST':
        sid = request.form.get('sid')
        # Expecting JSON-like structure or multiple arrays from form
        course_ids = request.form.getlist('course_id[]')
        types = request.form.getlist('type[]')
        
        courses = []
        for i in range(len(course_ids)):
            if course_ids[i] and str(course_ids[i]) != '0':
                courses.append({
                    'course_id': course_ids[i],
                    'type': types[i]
                })
        
        if sid and AdvisoryModel.save_course_plan(sid, courses, session['user_id']):
            flash('Course plan saved successfully.', 'success')
        else:
            flash('Error saving course plan.', 'danger')
        return redirect(url_for('academics.course_plan', **request.args))

    filters = {
        'college_id': request.args.get('college_id'),
        'session_id': request.args.get('session_id'),
        'degree_id': request.args.get('degree_id'),
        'branch_id': request.args.get('branch_id')
    }
    
    students = []
    if all([filters['college_id'], filters['session_id'], filters['degree_id']]):
        students = AdvisoryModel.get_students_for_advisory(filters)

    lookups = AdvisoryModel.get_advisory_lookups(filters['college_id'], filters['degree_id'])
    
    return render_template('academics/course_plan.html', 
                           lookups=lookups, 
                           filters=filters, 
                           students=students)

@academics_bp.route('/dean_pgs_course_plan_approval', methods=['GET', 'POST'])
@permission_required('Dean PGS approval (Course plan)')
def dean_pgs_course_plan_approval():
    if request.method == 'POST':
        action = request.form.get('action')
        sid = request.form.get('sid')
        user_id = session['user_id']
        
        if action in ['APPROVE', 'REJECT']:
            status = 'A' if action == 'APPROVE' else 'R'
            if sid:
                AdvisoryModel.approve_course_plan(sid, status, user_id, level='PGS')
                flash(f'Course plan { "approved" if status == "A" else "rejected" } successfully.', 'success')
            return redirect(url_for('academics.dean_pgs_course_plan_approval', **request.args))

    filters = {
        'college_id': request.args.get('college_id'),
        'session_id': request.args.get('session_id'),
        'degree_id': request.args.get('degree_id'),
        'branch_id': request.args.get('branch_id')
    }
    
    students = []
    if all([filters['college_id'], filters['session_id'], filters['degree_id']]):
        students = AdvisoryModel.get_pending_course_plan_approvals(filters, level='PGS')
    
    sid = request.args.get('sid')
    course_plan = []
    student_dtl = None
    if sid:
        course_plan = AdvisoryModel.get_student_course_plan(sid)
        student_dtl = AdvisoryModel.get_student_advisory_committee(sid) # Reusing this for basic student info

    lookups = AdvisoryModel.get_advisory_lookups(filters['college_id'], filters['degree_id'])
    return render_template('academics/dean_pgs_course_plan_approval.html', 
                           lookups=lookups, filters=filters, students=students, 
                           course_plan=course_plan, sid=sid, student=student_dtl)

@academics_bp.route('/pg_mandates_submission', methods=['GET', 'POST'])
@permission_required('PG Mandates Submission by HOD')
def pg_mandates_submission():
    filters = {
        'college_id': request.args.get('college_id'),
        'session_id': request.args.get('session_id'),
        'degree_id': request.args.get('degree_id'),
        'semester_id': request.args.get('semester_id'),
        'branch_id': request.args.get('branch_id')
    }

    if request.method == 'POST':
        action = request.form.get('action')
        user_id = session['user_id']
        
        if action == 'UPDATE_SINGLE':
            sid = request.form.get('sid')
            mandate_id = request.form.get('mandate_id')
            remarks = request.form.get('remarks')
            sub_date = request.form.get('sub_date')
            is_submitted = request.form.get('is_submitted') == '1'
            if sid and mandate_id:
                ResearchModel.update_mandate(sid, mandate_id, remarks, sub_date, is_submitted, user_id, filters)
                flash('Mandate updated successfully.', 'success')
            return redirect(url_for('academics.pg_mandates_submission', **request.args))

    students = []
    if all([filters['college_id'], filters['session_id'], filters['degree_id']]):
        students = ResearchModel.get_students_for_mandates(filters)

    # Use centralized AdvisoryModel lookups for correct discipline mapping (Int IDs)
    lookups = AdvisoryModel.get_advisory_lookups(filters['college_id'], filters['degree_id'])
    
    # Override colleges based on location (campus) from session
    loc_id = session.get('selected_loc')
    if loc_id:
        lookups['colleges'] = DB.fetch_all("SELECT pk_collegeid as id, collegename as name FROM SMS_College_Mst WHERE fk_locid = ? ORDER BY collegename", [loc_id])

    # Ensure degrees are filtered for PG/PhD specifically
    if filters['college_id'] and str(filters['college_id']) != '0':
        lookups['degrees'] = AcademicsModel.get_college_pg_degrees(filters['college_id'])

    # Standard class list (I to VIII)
    all_semesters = InfrastructureModel.get_all_semesters()
    lookups['semesters'] = [s for s in all_semesters if s.get('semesterorder', 0) <= 8]
    
    return render_template('academics/pg_mandates_submission.html', 
                           lookups=lookups, filters=filters, students=students)

@academics_bp.route('/pg_mandate_report')
@permission_required('PG Mandates Submission by HOD')
def pg_mandate_report():
    filters = {
        'college_id': request.args.get('college_id'),
        'session_id': request.args.get('session_id'),
        'degree_id': request.args.get('degree_id'),
        'branch_id': request.args.get('branch_id'),
        'format': request.args.get('format', 'pdf')
    }
    
    if not all([filters['college_id'], filters['session_id'], filters['degree_id']]):
        flash('Please apply filters first.', 'warning')
        return redirect(url_for('academics.pg_mandates_submission'))

    students = ResearchModel.get_students_for_mandates(filters)
    
    # Simple HTML-based report structure that mimics the .rpt layout
    html = render_template('reports/pg_mandates_report.html', students=students, filters=filters, 
                           now=datetime.now())

    if filters['format'] == 'excel':
        response = make_response(html)
        response.headers["Content-Disposition"] = "attachment; filename=PG_Mandates_Report.xls"
        response.headers["Content-Type"] = "application/vnd.ms-excel"
        return response
    elif filters['format'] == 'word':
        response = make_response(html)
        response.headers["Content-Disposition"] = "attachment; filename=PG_Mandates_Report.doc"
        response.headers["Content-Type"] = "application/msword"
        return response
    else:
        # Default PDF (browser print-friendly HTML)
        return html

@academics_bp.route('/add_with_approval_major_advisor', methods=['GET', 'POST'])
@permission_required('Addition/withdrawal Approval by Major Advisor')
def add_with_approval_major_advisor():
    from app.models.academics import AddWithModel
    user_id = session['user_id']
    
    if request.method == 'POST':
        action = request.form.get('action')
        pk_id = request.form.get('pk_id')
        remarks = request.form.get('remarks')
        
        if action in ['APPROVE', 'REJECT']:
            status = 'A' if action == 'APPROVE' else 'R'
            AddWithModel.approve_by_major_advisor(pk_id, status, remarks, user_id)
            flash(f'Request { "approved" if status == "A" else "rejected" } successfully.', 'success')
        
        return redirect(url_for('academics.add_with_approval_major_advisor', **request.args))

    filters = {
        'session_id': request.args.get('session_id', InfrastructureModel.get_current_session_id())
    }
    
    pending_requests = []
    processed_requests = []
    
    if filters['session_id']:
        pending_requests = AddWithModel.get_major_advisor_requests(filters, user_id, processed=False)
        processed_requests = AddWithModel.get_major_advisor_requests(filters, user_id, processed=True)

    lookups = {
        'sessions': InfrastructureModel.get_sessions()
    }
    
    return render_template('academics/add_with_approval_major_advisor.html', 
                           pending=pending_requests, 
                           processed=processed_requests,
                           lookups=lookups, filters=filters)

@academics_bp.route('/add_with_approval_teacher', methods=['GET', 'POST'])
@permission_required('Addition/withdrawal Approval by Teacher')
def add_with_approval_teacher():
    from app.models.academics import AddWithModel
    user_id = session['user_id']
    
    if request.method == 'POST':
        action = request.form.get('action')
        pk_id = request.form.get('pk_id')
        remarks = request.form.get('remarks')
        
        if action in ['APPROVE', 'REJECT']:
            status = 'A' if action == 'APPROVE' else 'R'
            AddWithModel.approve_by_teacher(pk_id, status, remarks, user_id)
            flash(f'Request { "approved" if status == "A" else "rejected" } successfully.', 'success')
        
        return redirect(url_for('academics.add_with_approval_teacher', **request.args))

    filters = {
        'session_id': request.args.get('session_id', InfrastructureModel.get_current_session_id())
    }
    
    pending_requests = []
    processed_requests = []
    
    if filters['session_id']:
        pending_requests = AddWithModel.get_teacher_requests(filters, user_id, processed=False)
        processed_requests = AddWithModel.get_teacher_requests(filters, user_id, processed=True)

    lookups = {
        'sessions': InfrastructureModel.get_sessions()
    }
    
    return render_template('academics/add_with_approval_teacher.html', 
                           pending=pending_requests, 
                           processed=processed_requests,
                           lookups=lookups, filters=filters)

@academics_bp.route('/student_thesis_detail', methods=['GET', 'POST'])
@permission_required('Student Thesis Detail')
def student_thesis_detail():
    filters = {
        'college_id': request.args.get('college_id'),
        'session_id': request.args.get('session_id'),
        'degree_id': request.args.get('degree_id'),
        'semester_id': request.args.get('semester_id'),
        'branch_id': request.args.get('branch_id')
    }

    if request.method == 'POST':
        action = request.form.get('action')
        user_id = session['user_id']
        
        if action == 'UPDATE_SINGLE':
            sid = request.form.get('sid')
            if sid:
                # Collect all fields for this sid
                data = {
                    'is_fee_remitted': request.form.get('is_fee_remitted') == '1',
                    'total_sem': request.form.get('total_sem'),
                    'thesis_sub_date': request.form.get('thesis_sub_date'),
                    'mc_date': request.form.get('mc_date'),
                    'viva_date': request.form.get('viva_date'),
                    'result_date': request.form.get('result_date'),
                    'thesis_not_no': request.form.get('thesis_not_no'),
                    'viva_not_no': request.form.get('viva_not_no'),
                    'result_not_no': request.form.get('result_not_no'),
                    'thesis_title': request.form.get('thesis_title'),
                    'adjudicator_remarks': request.form.get('adjudicator_remarks'),
                    'resubmission': request.form.get('resubmission'),
                    'remarks': request.form.get('remarks')
                }
                if ThesisModel.update_thesis_detail(sid, data, user_id):
                    flash('Thesis details updated successfully.', 'success')
                else:
                    flash('Error updating thesis details.', 'danger')
            return redirect(url_for('academics.student_thesis_detail', **request.args))

    students = []
    if all([filters['college_id'], filters['session_id'], filters['degree_id']]):
        students = ThesisModel.get_students_for_thesis(filters)

    lookups = AdvisoryModel.get_advisory_lookups(filters['college_id'], filters['degree_id'])
    
    # Campus-based college filter
    loc_id = session.get('selected_loc')
    if loc_id:
        lookups['colleges'] = DB.fetch_all("SELECT pk_collegeid as id, collegename as name FROM SMS_College_Mst WHERE fk_locid = ? ORDER BY collegename", [loc_id])

    # Filter for PG/PhD degrees
    if filters['college_id'] and str(filters['college_id']) != '0':
        lookups['degrees'] = AcademicsModel.get_college_pg_degrees(filters['college_id'])

    lookups['semesters'] = InfrastructureModel.get_all_semesters()
    
    return render_template('academics/student_thesis_detail.html', 
                           lookups=lookups, filters=filters, students=students)

@academics_bp.route('/student_thesis_report')
@permission_required('Student Thesis Detail')
def student_thesis_report():
    filters = {
        'college_id': request.args.get('college_id'),
        'session_id': request.args.get('session_id'),
        'degree_id': request.args.get('degree_id'),
        'branch_id': request.args.get('branch_id'),
        'format': request.args.get('format', 'pdf')
    }
    
    if not all([filters['college_id'], filters['session_id'], filters['degree_id']]):
        flash('Please apply filters first.', 'warning')
        return redirect(url_for('academics.student_thesis_detail'))

    students = ThesisModel.get_students_for_thesis(filters)
    
    # Get human-readable filter info for report header
    filters_info = {
        'college': DB.fetch_scalar("SELECT collegename FROM SMS_College_Mst WHERE pk_collegeid = ?", [filters['college_id']]),
        'session': DB.fetch_scalar("SELECT sessionname FROM SMS_AcademicSession_Mst WHERE pk_sessionid = ?", [filters['session_id']]),
        'degree': DB.fetch_scalar("SELECT degreename FROM SMS_Degree_Mst WHERE pk_degreeid = ?", [filters['degree_id']])
    }
    
    if filters['format'] == 'excel':
        html = render_template('reports/thesis_detail_report.html', students=students, filters=filters, now=datetime.now(), filters_info=filters_info)
        response = make_response(html)
        response.headers["Content-Disposition"] = "attachment; filename=Thesis_Detail_Report.xls"
        response.headers["Content-Type"] = "application/vnd.ms-excel"
        return response
    elif filters['format'] == 'word':
        html = render_template('reports/thesis_detail_report.html', students=students, filters=filters, now=datetime.now(), filters_info=filters_info)
        response = make_response(html)
        response.headers["Content-Disposition"] = "attachment; filename=Thesis_Detail_Report.doc"
        response.headers["Content-Type"] = "application/msword"
        return response
    else:
        from app.utils import generate_thesis_detail_pdf
        pdf_content = generate_thesis_detail_pdf(students, filters_info)
        return send_file(pdf_content, download_name="Thesis_Detail_Report.pdf", as_attachment=True, mimetype='application/pdf')

@academics_bp.route('/addition_withdrawal_approval_status', methods=['GET', 'POST'])
@permission_required('Addition/Withdrawal Approval Status')
def addition_withdrawal_approval_status():
    from app.models.academics import AddWithModel
    filters = {
        'college_id': request.args.get('college_id'),
        'session_id': request.args.get('session_id'),
        'degree_id': request.args.get('degree_id'),
        'semester_id': request.args.get('semester_id'),
        'branch_id': request.args.get('branch_id')
    }

    status_list = []
    if all([filters['session_id'], filters['degree_id']]):
        status_list = AddWithModel.get_add_with_status(filters)

    lookups = AdvisoryModel.get_advisory_lookups(filters['college_id'], filters['degree_id'])
    
    # Load colleges based on location (campus) from session if available
    loc_id = session.get('selected_loc')
    if loc_id:
        lookups['colleges'] = DB.fetch_all("SELECT pk_collegeid as id, collegename as name FROM SMS_College_Mst WHERE fk_locid = ? ORDER BY collegename", [loc_id])

    # Ensure degrees are filtered for PG/PhD specifically
    if filters['college_id'] and str(filters['college_id']) != '0':
        lookups['degrees'] = AcademicsModel.get_college_pg_degrees(filters['college_id'])

    lookups['semesters'] = InfrastructureModel.get_all_semesters()
    
    return render_template('academics/addition_withdrawal_approval_status.html', 
                           lookups=lookups, filters=filters, status_list=status_list)

@academics_bp.route('/add_with_status_report')
@permission_required('Addition/Withdrawal Approval Status')
def add_with_status_report():
    from app.models.academics import AddWithModel
    filters = {
        'session_id': request.args.get('session_id'),
        'degree_id': request.args.get('degree_id'),
        'semester_id': request.args.get('semester_id'),
        'branch_id': request.args.get('branch_id'),
        'format': request.args.get('format', 'pdf')
    }
    
    if not all([filters['session_id'], filters['degree_id']]):
        flash('Please apply filters first.', 'warning')
        return redirect(url_for('academics.addition_withdrawal_approval_status'))

    status_list = AddWithModel.get_add_with_status(filters)
    
    # Get human-readable filter info for report header
    filters_info = {
        'session': DB.fetch_scalar("SELECT sessionname FROM SMS_AcademicSession_Mst WHERE pk_sessionid = ?", [filters['session_id']]),
        'degree': DB.fetch_scalar("SELECT degreename FROM SMS_Degree_Mst WHERE pk_degreeid = ?", [filters['degree_id']])
    }
    
    if filters['format'] == 'excel':
        html = render_template('reports/add_with_status_report.html', status_list=status_list, filters=filters, now=datetime.now(), filters_info=filters_info)
        response = make_response(html)
        response.headers["Content-Disposition"] = "attachment; filename=Add_With_Status_Report.xls"
        response.headers["Content-Type"] = "application/vnd.ms-excel"
        return response
    elif filters['format'] == 'word':
        html = render_template('reports/add_with_status_report.html', status_list=status_list, filters=filters, now=datetime.now(), filters_info=filters_info)
        response = make_response(html)
        response.headers["Content-Disposition"] = "attachment; filename=Add_With_Status_Report.doc"
        response.headers["Content-Type"] = "application/msword"
        return response
    else:
        from app.utils import generate_add_with_status_pdf
        pdf_content = generate_add_with_status_pdf(status_list, filters_info)
        return send_file(pdf_content, download_name="Add_With_Status_Report.pdf", as_attachment=True, mimetype='application/pdf')

@academics_bp.route('/advisory_creation_approval_status', methods=['GET', 'POST'])
@permission_required('Advisory Creation And Approval Status')
def advisory_creation_approval_status():
    filters = {
        'college_id': request.args.get('college_id'),
        'session_id': request.args.get('session_id'),
        'degree_id': request.args.get('degree_id'),
        'semester_id': request.args.get('semester_id'),
        'branch_id': request.args.get('branch_id')
    }

    status_list = []
    if all([filters['college_id'], filters['session_id'], filters['degree_id']]):
        status_list = AdvisoryStatusModel.get_advisory_approval_status(filters)

    lookups = AdvisoryModel.get_advisory_lookups(filters['college_id'], filters['degree_id'])
    loc_id = session.get('selected_loc')
    if loc_id:
        lookups['colleges'] = DB.fetch_all("SELECT pk_collegeid as id, collegename as name FROM SMS_College_Mst WHERE fk_locid = ? ORDER BY collegename", [loc_id])
    if filters['college_id'] and str(filters['college_id']) != '0':
        lookups['degrees'] = AcademicsModel.get_college_pg_degrees(filters['college_id'])
    lookups['semesters'] = InfrastructureModel.get_all_semesters()
    
    return render_template('academics/advisory_creation_approval_status.html', 
                           lookups=lookups, filters=filters, status_list=status_list)


@academics_bp.route('/dean_pgs_approval_advisory', methods=['GET', 'POST'])
@permission_required('Dean PGS approval (advisory committee)')
def dean_pgs_approval_advisory():
    user_id = session.get('user_id')
    loc_id = session.get('selected_loc')
    
    # Force default session to '0' on page load if not specified
    lookups = AdvisoryModel.get_academic_context_lookups(user_id, loc_id, 'DEAN_PGS', request.args)
    if not request.args.get('session_id'):
        lookups['session_id'] = '0'

    filters = {
        'college_id': lookups.get('college_id', '0'),
        'session_id': lookups.get('session_id', '0'),
        'degree_id': lookups.get('degree_id', '0'),
        'branch_id': lookups.get('branch_id', '0'),
        'user_dept': lookups.get('user_dept')
    }
    
    if request.method == 'POST':
        action = request.form.get('action')
        selected_students = request.form.getlist('selected_students')
        if action and selected_students:
            AdvisoryModel.update_approval_status(selected_students, 'dean_pgs', action, session.get('emp_id'))
            flash(f"Successfully {action}ed the selected committees.", 'success')
        return redirect(url_for('academics.dean_pgs_approval_advisory', **filters))
    
    pending_students = AdvisoryModel.get_pending_approvals(filters, 'dean_pgs', session.get('emp_id'))
        
    return render_template('academics/dean_pgs_approval_advisory.html', lookups=lookups, filters=filters, pending_students=pending_students)

@academics_bp.route('/hod_approval_advisory', methods=['GET', 'POST'])
@permission_required('HOD Approval')
def hod_approval_advisory():
    user_id = session.get('user_id')
    loc_id = session.get('selected_loc')
    
    # Force default session to '0' on page load if not specified
    lookups = AdvisoryModel.get_academic_context_lookups(user_id, loc_id, 'HOD', request.args)
    if not request.args.get('session_id'):
        lookups['session_id'] = '0'
    
    filters = {
        'college_id': lookups.get('college_id', '0'),
        'session_id': lookups.get('session_id', '0'),
        'degree_id': lookups.get('degree_id', '0'),
        'branch_id': lookups.get('branch_id', '0'),
        'user_dept': lookups.get('user_dept')
    }
    
    edit_sid = request.args.get('edit_sid')
    edit_adcid = request.args.get('edit_adcid')
    
    if request.method == 'POST':
        action = request.form.get('action')
        adcid = request.form.get('adcid')
        remarks = request.form.get('remarks')
        
        if action and adcid and remarks:
            AdvisoryModel.update_approval_status(adcid, 'hod', action, session.get('emp_id'), remarks)
            flash(f"Successfully {action}ed the selected committee.", 'success')
        return redirect(url_for('academics.hod_approval_advisory', **filters))
    
    # Pagination
    page = request.args.get('page', 1, type=int)
    per_page = 10
    
    grid_students, total = AdvisoryModel.get_advisory_approvals_list(filters, 'hod', page=page, per_page=per_page)
    pagination = {'page': page, 'per_page': per_page, 'total': total, 'has_next': (page * per_page) < total}
    
    edit_student = None
    advisory_details = []
    
    if edit_sid:
        # Find student info
        from app.db import DB
        edit_student = DB.fetch_one("SELECT S.pk_sid, S.fullname, S.enrollmentno, S.AdmissionNo FROM SMS_Student_Mst S WHERE pk_sid = ?", [edit_sid])
        
        # Get advisory committee details
        advisory_details = DB.fetch_all('''
            SELECT 
                RM.statusname as role_name,
                E.pk_empid as fk_empid,
                E.empname as advisor_name,
                ISNULL(DES.designation, '') as designation,
                ISNULL(DEPT.description, '') as department
            FROM SMS_Advisory_Committee_Dtl D
            JOIN SMS_AdvisoryStatus_Mst RM ON D.fk_statusid = RM.pk_stid
            JOIN SAL_Employee_Mst E ON D.fk_empid = E.pk_empid
            LEFT JOIN SAL_Designation_Mst DES ON E.fk_desgid = DES.pk_desgid
            LEFT JOIN Department_Mst DEPT ON E.fk_deptid = DEPT.pk_deptid
            WHERE D.fk_adcid = ?
            ORDER BY RM.pk_stid
        ''', [edit_adcid])
        
    return render_template('academics/hod_approval_advisory.html', 
                           lookups=lookups, 
                           filters=filters, 
                           grid_students=clean_json_data(grid_students),
                           pagination=pagination,
                           edit_student=edit_student,
                           edit_adcid=edit_adcid,
                           advisory_details=advisory_details)

@academics_bp.route('/college_dean_approval_advisory', methods=['GET', 'POST'])
@permission_required('College Dean Approval')
def college_dean_approval_advisory():
    user_id = session.get('user_id')
    loc_id = session.get('selected_loc')
    
    lookups = AdvisoryModel.get_academic_context_lookups(user_id, loc_id, 'DEAN', request.args)
    if not request.args.get('session_id'):
        lookups['session_id'] = '0'
    
    filters = {
        'college_id': lookups.get('college_id', '0'),
        'session_id': lookups.get('session_id', '0'),
        'degree_id': lookups.get('degree_id', '0'),
        'branch_id': lookups.get('branch_id', '0'),
        'user_dept': lookups.get('user_dept')
    }
    
    if request.method == 'POST':
        action = request.form.get('action')
        selected_students = request.form.getlist('selected_students')
        if action and selected_students:
            AdvisoryModel.update_approval_status(selected_students, 'dean', action, session.get('emp_id'))
            flash(f"Successfully {action}ed the selected committees.", 'success')
        return redirect(url_for('academics.college_dean_approval_advisory', **filters))
    
    pending_students = AdvisoryModel.get_pending_approvals(filters, 'dean', session.get('emp_id'))
        
    return render_template('academics/college_dean_approval_advisory.html', lookups=lookups, filters=filters, pending_students=pending_students)

@academics_bp.route('/prepare_course_plan', methods=['GET', 'POST'])
@permission_required('Prepare Course Plan')
def prepare_course_plan():
    user_id = session.get('user_id')
    loc_id = session.get('selected_loc')
    
    filters = {
        'college_id': request.args.get('college_id'),
        'session_id': request.args.get('session_id'),
        'degree_id': request.args.get('degree_id'),
        'branch_id': request.args.get('branch_id'),
        'sid': request.args.get('sid')
    }
    
    if request.method == 'POST':
        sid = filters.get('sid')
        course_ids = request.form.getlist('course_ids[]')
        if sid and course_ids:
            AdvisoryModel.save_course_plan(sid, course_ids, session.get('emp_id'))
            flash("Course plan saved successfully.", 'success')
        return redirect(url_for('academics.prepare_course_plan', **filters))
    
    lookups = AdvisoryModel.get_academic_context_lookups(user_id, loc_id, 'TEACHER', request.args)
    students = []
    current_plan = []
    
    if filters['college_id'] and filters['degree_id'] and filters['branch_id']:
        students = AdvisoryModel.get_students_for_course_plan(filters)
        
    if filters['sid'] and filters['sid'] != '0':
        current_plan = AdvisoryModel.get_student_course_plan_advisory(filters['sid'])
        
    return render_template('academics/prepare_course_plan.html', lookups=lookups, filters=filters, students=students, current_plan=current_plan)

@academics_bp.route('/advisory_status_report')
@permission_required('Advisory Creation And Approval Status')
def advisory_status_report():
    filters = {
        'college_id': request.args.get('college_id'),
        'session_id': request.args.get('session_id'),
        'degree_id': request.args.get('degree_id'),
        'semester_id': request.args.get('semester_id'),
        'branch_id': request.args.get('branch_id'),
        'format': request.args.get('format', 'pdf')
    }
    
    if not all([filters['college_id'], filters['session_id'], filters['degree_id']]):
        flash('Please apply filters first.', 'warning')
        return redirect(url_for('academics.advisory_creation_approval_status'))

    status_list = AdvisoryStatusModel.get_advisory_approval_status(filters)
    
    filters_info = {
        'college': DB.fetch_scalar("SELECT collegename FROM SMS_College_Mst WHERE pk_collegeid = ?", [filters['college_id']]),
        'session': DB.fetch_scalar("SELECT sessionname FROM SMS_AcademicSession_Mst WHERE pk_sessionid = ?", [filters['session_id']]),
        'degree': DB.fetch_scalar("SELECT degreename FROM SMS_Degree_Mst WHERE pk_degreeid = ?", [filters['degree_id']])
    }

    if filters['format'] == 'excel':
        html = render_template('reports/advisory_status_report.html', status_list=status_list, filters=filters, now=datetime.now(), filters_info=filters_info)
        response = make_response(html)
        response.headers["Content-Disposition"] = "attachment; filename=Advisory_Status_Report.xls"
        response.headers["Content-Type"] = "application/vnd.ms-excel"
        return response
    elif filters['format'] == 'word':
        html = render_template('reports/advisory_status_report.html', status_list=status_list, filters=filters, now=datetime.now(), filters_info=filters_info)
        response = make_response(html)
        response.headers["Content-Disposition"] = "attachment; filename=Advisory_Status_Report.doc"
        response.headers["Content-Type"] = "application/msword"
        return response
    else:
        from app.utils import generate_advisory_status_pdf
        pdf_content = generate_advisory_status_pdf(status_list, filters_info)
        return send_file(pdf_content, download_name="Advisory_Status_Report.pdf", as_attachment=True, mimetype='application/pdf')

@academics_bp.route('/programme_of_work_pg', methods=['GET', 'POST'])
@permission_required('Programme of work(PG)')
def programme_of_work_pg():
    filters = {
        'college_id': request.args.get('college_id'),
        'session_id': request.args.get('session_id'),
        'degree_id': request.args.get('degree_id'),
        'semester_id': request.args.get('semester_id'),
        'branch_id': request.args.get('branch_id'),
        'year': request.args.get('year')
    }

    if request.method == 'POST':
        # Handle report generation
        return redirect(url_for('academics.programme_of_work_report', **request.form))

    # Lookups
    loc_id = session.get('selected_loc')
    if loc_id:
        colleges = DB.fetch_all("SELECT pk_collegeid as id, collegename as name FROM SMS_College_Mst WHERE fk_locid = ? ORDER BY collegename", [loc_id])
    else:
        colleges = AcademicsModel.get_colleges_simple()

    lookups = {
        'colleges': colleges,
        'sessions': InfrastructureModel.get_sessions(),
        'degrees': [],
        'branches': [],
        'semesters': InfrastructureModel.get_all_semesters()
    }

    if filters['college_id']:
        lookups['degrees'] = AcademicsModel.get_college_pg_degrees(filters['college_id'])

    if filters['college_id'] and filters['degree_id']:
         lookups['branches'] = DB.fetch_all("""
            SELECT DISTINCT B.Pk_BranchId as id, B.Branchname as name
            FROM SMS_BranchMst B
            INNER JOIN (
                SELECT branchid, fk_Coldgbrmapnewid as mapid FROM SMS_CollegeDegreeBranchMap_dtlnew
                UNION ALL
                SELECT fk_branchid as branchid, fk_Coldgbrmapid as mapid FROM SMS_CollegeDegreeBranchMap_dtl
            ) D ON B.Pk_BranchId = D.branchid
            INNER JOIN SMS_CollegeDegreeBranchMap_Mst MST ON D.mapid = MST.PK_Coldgbrid
            WHERE MST.fk_CollegeId = ? AND MST.fk_Degreeid = ?
            ORDER BY B.Branchname
        """, [filters['college_id'], filters['degree_id']])

    return render_template('academics/programme_of_work_pg.html', lookups=lookups, filters=filters)

@academics_bp.route('/programme_of_work_report')
@permission_required('Programme of work(PG)')
def programme_of_work_report():
    filters = {
        'college_id': request.args.get('college_id'),
        'session_id': request.args.get('session_id'),
        'degree_id': request.args.get('degree_id'),
        'branch_id': request.args.get('branch_id'),
        'semester_id': request.args.get('semester_id'),
        'format': request.args.get('format', 'pdf')
    }
    
    if not all([filters['college_id'], filters['session_id'], filters['degree_id']]):
        return "Please select College, Session and Degree to view report."

    # Fetch detailed students matching criteria using improved model method
    detailed_students = ResearchModel.get_students_for_work_programme(filters)
    
    if filters['format'] == 'excel':
        html = render_template('reports/programme_of_work_detailed.html', students=detailed_students, filters=filters, now=datetime.now())
        response = make_response(html)
        response.headers["Content-Disposition"] = "attachment; filename=Programme_of_Work.xls"
        response.headers["Content-Type"] = "application/vnd.ms-excel"
        return response
    elif filters['format'] == 'word':
        html = render_template('reports/programme_of_work_detailed.html', students=detailed_students, filters=filters, now=datetime.now())
        response = make_response(html)
        response.headers["Content-Disposition"] = "attachment; filename=Programme_of_Work.doc"
        response.headers["Content-Type"] = "application/msword"
        return response
    else:
        # PROFESSIONAL PDF GENERATION USING REPORTLAB
        from app.utils import generate_programme_of_work_pdf
        pdf_content = generate_programme_of_work_pdf(detailed_students)
        return send_file(pdf_content, download_name="Programme_of_Work.pdf", as_attachment=True, mimetype='application/pdf')

@academics_bp.route('/igrade_approval_teacher', methods=['GET', 'POST'])
@permission_required('I-Grade Approval by Teacher')
def igrade_approval_teacher():
    user_id = session['user_id']
    if request.method == 'POST':
        action = request.form.get('action') # 'A' or 'R' from radio/button
        pk_id = request.form.get('pk_id')
        remarks = request.form.get('remarks')
        if pk_id:
            IGradeModel.approve_by_teacher(pk_id, action, remarks, user_id)
            flash(f'Request { "approved" if action == "A" else "rejected" } successfully.', 'success')
        return redirect(url_for('academics.igrade_approval_teacher', **request.args))

    filters = {
        'session_id': request.args.get('session_id', InfrastructureModel.get_current_session_id())
    }
    
    pending = []
    processed = []
    if filters['session_id']:
        pending = IGradeModel.get_teacher_requests(filters, user_id, processed=False)
        processed = IGradeModel.get_teacher_requests(filters, user_id, processed=True)

    lookups = {'sessions': InfrastructureModel.get_sessions()}
    return render_template('academics/igrade_approval_teacher.html', 
                           lookups=lookups, filters=filters, pending=pending, processed=processed)

@academics_bp.route('/igrade_approval_dean_pgs', methods=['GET', 'POST'])
@permission_required('I-Grade Approval By Dean Pgs')
def igrade_approval_dean_pgs():
    user_id = session['user_id']
    if request.method == 'POST':
        action = request.form.get('action')
        pk_id = request.form.get('pk_id')
        remarks = request.form.get('remarks')
        if pk_id:
            IGradeModel.approve_by_dean(pk_id, action, remarks, user_id)
            flash(f'Request { "approved" if action == "A" else "rejected" } successfully.', 'success')
        return redirect(url_for('academics.igrade_approval_dean_pgs', **request.args))

    filters = {
        'session_id': request.args.get('session_id', InfrastructureModel.get_current_session_id())
    }
    
    pending = []
    processed = []
    if filters['session_id']:
        pending = IGradeModel.get_dean_requests(filters, processed=False)
        processed = IGradeModel.get_dean_requests(filters, processed=True)

    lookups = {'sessions': InfrastructureModel.get_sessions()}
    return render_template('academics/igrade_approval_dean_pgs.html', 
                           lookups=lookups, filters=filters, pending=pending, processed=processed)

@academics_bp.route('/igrade_approval_status', methods=['GET'])
@permission_required('I Grade Approval Status')
def igrade_approval_status():
    filters = {
        'college_id': request.args.get('college_id'),
        'session_id': request.args.get('session_id'),
        'degree_id': request.args.get('degree_id'),
        'semester_id': request.args.get('semester_id'),
        'branch_id': request.args.get('branch_id')
    }

    status_list = []
    if all([filters['college_id'], filters['session_id'], filters['degree_id']]):
        status_list = IGradeModel.get_igrade_status(filters)

    lookups = AdvisoryModel.get_advisory_lookups(filters['college_id'], filters['degree_id'])
    
    # Campus-based college filter
    loc_id = session.get('selected_loc')
    if loc_id:
        lookups['colleges'] = DB.fetch_all("SELECT pk_collegeid as id, collegename as name FROM SMS_College_Mst WHERE fk_locid = ? ORDER BY collegename", [loc_id])

    # Filter for PG/PhD degrees
    if filters['college_id'] and str(filters['college_id']) != '0':
        lookups['degrees'] = AcademicsModel.get_college_pg_degrees(filters['college_id'])

    lookups['semesters'] = InfrastructureModel.get_all_semesters()
    
    return render_template('academics/igrade_approval_status.html', 
                           lookups=lookups, filters=filters, status_list=status_list)

@academics_bp.route('/igrade_status_report')
@permission_required('I Grade Approval Status')
def igrade_status_report():
    filters = {
        'college_id': request.args.get('college_id'),
        'session_id': request.args.get('session_id'),
        'degree_id': request.args.get('degree_id'),
        'branch_id': request.args.get('branch_id'),
        'format': request.args.get('format', 'pdf')
    }
    
    if not all([filters['college_id'], filters['session_id'], filters['degree_id']]):
        flash('Please apply filters first.', 'warning')
        return redirect(url_for('academics.igrade_approval_status'))

    status_list = IGradeModel.get_igrade_status(filters)
    
    filters_info = {
        'college': DB.fetch_scalar("SELECT collegename FROM SMS_College_Mst WHERE pk_collegeid = ?", [filters['college_id']]),
        'session': DB.fetch_scalar("SELECT sessionname FROM SMS_AcademicSession_Mst WHERE pk_sessionid = ?", [filters['session_id']]),
        'degree': DB.fetch_scalar("SELECT degreename FROM SMS_Degree_Mst WHERE pk_degreeid = ?", [filters['degree_id']])
    }

    if filters['format'] == 'excel':
        html = render_template('reports/igrade_status_report.html', status_list=status_list, filters=filters, now=datetime.now(), filters_info=filters_info)
        response = make_response(html)
        response.headers["Content-Disposition"] = "attachment; filename=IGrade_Status_Report.xls"
        response.headers["Content-Type"] = "application/vnd.ms-excel"
        return response
    elif filters['format'] == 'word':
        html = render_template('reports/igrade_status_report.html', status_list=status_list, filters=filters, now=datetime.now(), filters_info=filters_info)
        response = make_response(html)
        response.headers["Content-Disposition"] = "attachment; filename=IGrade_Status_Report.doc"
        response.headers["Content-Type"] = "application/msword"
        return response
    else:
        from app.utils import generate_igrade_status_pdf
        pdf_content = generate_igrade_status_pdf(status_list, filters_info)
        return send_file(pdf_content, download_name="IGrade_Status_Report.pdf", as_attachment=True, mimetype='application/pdf')

@academics_bp.route('/course_allocation_pg', methods=['GET', 'POST'])
@permission_required('Course Allocation(For PG)')
def course_allocation_pg():
    if request.method == 'POST':
        action = request.form.get('action')
        sid = request.form.get('sid')
        exconfig_id = request.form.get('exconfig_id')
        user_id = session['user_id']
        
        if action == 'ALLOCATE':
            course_ids = request.form.getlist('course_id[]')
            semester_id = request.args.get('semester_id')
            if sid and exconfig_id and course_ids and semester_id:
                CourseAllocationModel.allocate_courses(sid, course_ids, exconfig_id, user_id, semester_id)
                flash('Courses allocated successfully.', 'success')
            return redirect(url_for('academics.course_allocation_pg', **request.args))
            
        elif action == 'DELETE':
            alloc_ids = request.form.getlist('alloc_id[]')
            if sid and alloc_ids:
                CourseAllocationModel.delete_allocated_courses(sid, alloc_ids)
                flash('Selected allocations deleted.', 'success')
            return redirect(url_for('academics.course_allocation_pg', **request.args))
            
        elif action == 'ADD_EXTERNAL':
            course_id = request.form.get('external_course_id')
            semester_id = request.args.get('semester_id')
            if sid and exconfig_id and course_id and semester_id:
                CourseAllocationModel.allocate_courses(sid, [course_id], exconfig_id, user_id, semester_id)
                flash('Course allocated successfully.', 'success')
            return redirect(url_for('academics.course_allocation_pg', **request.args))

    filters = {
        'college_id': request.args.get('college_id'),
        'session_id': request.args.get('session_id'),
        'degree_id': request.args.get('degree_id'),
        'semester_id': request.args.get('semester_id'),
        'branch_id': request.args.get('branch_id'),
        'exconfig_id': request.args.get('exconfig_id')
    }
    
    students = []
    exam_configs = []
    if filters['degree_id']:
        exam_configs = CourseAllocationModel.get_exam_configs(
            filters['degree_id'], 
            session_id=filters.get('session_id'), 
            semester_id=filters.get('semester_id')
        )
        
    if all([filters['college_id'], filters['session_id'], filters['degree_id'], filters['branch_id']]):
        students = CourseAllocationModel.get_students_for_allocation(filters)
        
    sid = request.args.get('sid')
    allocated_courses = []
    course_plan = []
    external_courses = []
    if sid:
        if filters['exconfig_id']:
            allocated_courses = CourseAllocationModel.get_allocated_courses(sid, filters['exconfig_id'])
        course_plan = CourseAllocationModel.get_student_course_plan_for_allocation(sid)
        external_courses = CourseAllocationModel.get_courses_not_in_plan(sid)

    lookups = AdvisoryModel.get_advisory_lookups(filters['college_id'], filters['degree_id'])
    # Load colleges based on location (campus) from session if available
    loc_id = session.get('selected_loc')
    if loc_id:
        lookups['colleges'] = DB.fetch_all("SELECT pk_collegeid as id, collegename as name FROM SMS_College_Mst WHERE fk_locid = ? ORDER BY collegename", [loc_id])
    else:
        lookups['colleges'] = AcademicsModel.get_colleges_simple()

    # Filter semesters to only show I to VIII as requested
    all_semesters = InfrastructureModel.get_all_semesters()
    lookups['semesters'] = [s for s in all_semesters if s.get('semesterorder', 0) <= 8]
    
    return render_template('academics/course_allocation_pg.html', 
                           lookups=lookups, filters=filters, students=students,
                           exam_configs=exam_configs, allocated_courses=allocated_courses,
                           course_plan=course_plan, sid=sid)

@academics_bp.route('/student/<int:sid>/alert')
@permission_required('Student BioData')
def student_alert(sid):
    # Fetch alerts/achievements/disciplinary actions for this student
    # For now, fetching basic info + any achievements
    info = StudentModel.get_student_info(sid)
    achievements = DB.fetch_all("SELECT * FROM SMS_StuAchievement_Dtl WHERE fk_sturegid = ?", [sid])
    disciplinary = DB.fetch_all("SELECT * FROM SMS_StuDisciplinary_Dtl WHERE fk_sid = ?", [sid])
    
    return render_template('academics/student_alert.html', info=info, achievements=achievements, disciplinary=disciplinary)

@academics_bp.route('/student/<int:sid>/advisors_popup')
@permission_required('Dean PGS approval (Course plan)')
def advisors_popup(sid):
    data = AdvisoryModel.get_student_advisory_committee(sid)
    return render_template('academics/advisors_popup.html', data=data)

@academics_bp.route('/course_offer_hod', methods=['GET', 'POST'])
@permission_required('Course Offer (By HOD)')
def course_offer_hod():
    user_id = session.get('user_id')
    emp_id = session.get('emp_id')
    dept_ctx = AcademicsModel.get_hod_department_context(emp_id) if emp_id else {'branch_ids': [], 'hr_departments': [], 'sms_dept_ids': []}
    hod_branch_ids = dept_ctx.get('branch_ids', [])
    sms_dept_ids = dept_ctx.get('sms_dept_ids', [])
    code_prefixes = dept_ctx.get('code_prefixes', [])

    if request.method == 'POST':
        filters = {
            'college_id': request.form.get('college_id', type=int),
            'session_id': request.form.get('session_id', type=int),
            'degree_id': request.form.get('degree_id', type=int),
            'semester_id': request.form.get('semester_id', type=int),
            'year_id': request.form.get('year_id', type=int),
            'exconfig_id': request.form.get('exconfig_id', type=int),
            'branch_id': request.form.get('branch_id', type=int) or 0
        }
        course_ids = request.form.getlist('course_ids[]')
        try:
            AcademicsModel.save_course_offer_by_hod(filters, course_ids, user_id, emp_id)
            flash('Course offer updated successfully!', 'success')
        except Exception:
            flash('Error updating course offer.', 'danger')
        return redirect(url_for('academics.course_offer_hod', **{k: v for k, v in filters.items() if v}))

    filters = {
        'college_id': request.args.get('college_id', type=int),
        'session_id': request.args.get('session_id', type=int),
        'degree_id': request.args.get('degree_id', type=int),
        'semester_id': request.args.get('semester_id', type=int),
        'year_id': request.args.get('year_id', type=int),
        'exconfig_id': request.args.get('exconfig_id', type=int),
        'branch_id': request.args.get('branch_id', type=int)
    }

    # Auto-select college based on selected location
    if not filters['college_id']:
        loc_id = session.get('selected_loc')
        if loc_id:
            colleges = DB.fetch_all("SELECT pk_collegeid as id FROM SMS_College_Mst WHERE fk_locid = ?", [loc_id])
            if len(colleges) == 1:
                filters['college_id'] = colleges[0]['id']

    # Auto-select current session by sessionorder
    if not filters['session_id']:
        sess = DB.fetch_one("SELECT TOP 1 pk_sessionid FROM SMS_AcademicSession_Mst ORDER BY sessionorder DESC, pk_sessionid DESC")
        if sess:
            filters['session_id'] = sess['pk_sessionid']

    lookups = {
        'colleges': AcademicsModel.get_colleges_simple(),
        'sessions': InfrastructureModel.get_sessions(),
        'degrees': [],
        'semesters': [],
        'years': AcademicsModel.get_degree_years(),
        'exam_configs': []
    }

    if filters['college_id']:
        lookups['degrees'] = AcademicsModel.get_college_degrees(filters['college_id'])
    if filters['degree_id']:
        lookups['semesters'] = AcademicsModel.get_degree_semesters_for_degree(filters['degree_id'])
        if filters['semester_id']:
            # Auto-calculate year if not provided but semester is
            if not filters['year_id']:
                year_row = DB.fetch_one("SELECT fk_degreeyearid FROM SMS_Semester_Mst WHERE pk_semesterid = ?", [filters['semester_id']])
                if year_row:
                    filters['year_id'] = year_row['fk_degreeyearid']
            
        lookups['exam_configs'] = CourseAllocationModel.get_exam_configs(
            filters['degree_id'], filters['session_id'], filters['semester_id']
        )

    courses = []
    selected_ids = set()
    if filters['degree_id'] and filters['semester_id']:
        # HOD offers courses for PG/PhD only. UG is offered by Dean.
        degree_info = DB.fetch_one("SELECT fk_degreetypeid FROM SMS_Degree_Mst WHERE pk_degreeid = ?", [filters['degree_id']])
        is_ug = False
        if degree_info:
            dtype = DB.fetch_one("SELECT isug FROM SMS_DegreeType_Mst WHERE pk_degreetypeid = ?", [degree_info['fk_degreetypeid']])
            is_ug = (dtype and dtype.get('isug') == 'B')

        if not is_ug:
            year_semesters = []
            if filters.get('year_id'):
                year_semesters = AcademicsModel.get_semesters_for_degree_year(filters['degree_id'], filters['year_id'])
            
            raw_courses = []
            if year_semesters:
                raw_courses = AcademicsModel.get_courses_for_degree_semesters(
                    filters['degree_id'], year_semesters, None, sms_dept_ids, code_prefixes, is_hod_view=True
                )
            else:
                raw_courses = AcademicsModel.get_courses_for_degree_semester(
                    filters['degree_id'], filters['semester_id'], None, sms_dept_ids, code_prefixes, is_hod_view=True
                )
            
            # Ensure Uniqueness to avoid repeats
            seen = set()
            for c in raw_courses:
                th = c.get('crhr_theory') or 0
                pr = c.get('crhr_practical') or 0
                code = (c.get('coursecode') or '').strip()
                name = (c.get('coursename') or '').strip()
                # Unique key to identify duplicates
                key = (code.upper(), name.upper(), th, pr)
                if key not in seen:
                    seen.add(key)
                    c['display_name'] = f"{code},({th}+{pr}), /{name}"
                    courses.append(c)

        master = AcademicsModel.get_course_offer_master(filters)
        if master:
            selected_ids = set(AcademicsModel.get_course_offer_selected_course_ids(master['Pk_courseallocid']))
        else:
            # Auto-check ONLY the courses that are officially mapped to this semester in the syllabus
            branch_filter = f" AND fk_branchid = {filters.get('branch_id')}" if filters.get('branch_id') else ""
            default_courses = DB.fetch_all(f"SELECT fk_courseid FROM SMS_Course_Mst_Dtl WHERE fk_degreeid = ? AND fk_semesterid = ?{branch_filter}", [filters.get('degree_id'), filters.get('semester_id')])
            selected_ids = set(c['fk_courseid'] for c in default_courses)

    return render_template('academics/course_offer_hod.html',
                           lookups=lookups,
                           filters=filters,
                           courses=courses,
                           selected_ids=selected_ids,
                           hod_departments=dept_ctx.get('hr_departments', []))

@academics_bp.route('/teacher_course_assignment', methods=['GET', 'POST'])
@permission_required('Teacher Course Assignment')
def teacher_course_assignment():
    user_id = session.get('user_id')
    emp_id = session.get('emp_id')
    dept_ctx = AcademicsModel.get_hod_department_context(emp_id) if emp_id else {'branch_ids': [], 'hr_departments': [], 'sms_dept_ids': []}
    hod_hr_dept_ids = [d['id'] for d in dept_ctx.get('hr_departments', [])]
    code_prefixes = dept_ctx.get('code_prefixes', [])
    include_all = request.args.get('include_all') == '1'

    if request.method == 'POST':
        action = request.form.get('action', 'SAVE')
        filters = {
            'college_id': request.form.get('college_id', type=int),
            'session_id': request.form.get('session_id', type=int),
            'degree_id': request.form.get('degree_id', type=int),
            'semester_id': request.form.get('semester_id', type=int),
            'year_id': request.form.get('year_id', type=int),
            'employee_id': request.form.get('employee_id'),
            'exconfig_id': request.form.get('exconfig_id', type=int),
            'coursetype': request.form.get('coursetype'),
            'batch_id': request.form.get('batch_id', type=int),
            'branch_id': request.form.get('branch_id', type=int) or 0
        }
        if filters.get('coursetype') == '0':
            filters['coursetype'] = None
        course_ids = request.form.getlist('course_ids[]')
        main_ids = request.form.getlist('main_course_ids[]')
        try:
            if action == 'DELETE':
                AcademicsModel.delete_teacher_course_assignment_courses(filters, course_ids)
                flash('Selected courses deleted.', 'success')
            else:
                AcademicsModel.save_teacher_course_assignment(filters, course_ids, main_ids, user_id)
                flash('Teacher course assignment saved successfully!', 'success')
        except Exception:
            flash('Error saving teacher course assignment.', 'danger')
        return redirect(url_for('academics.teacher_course_assignment', **{k: v for k, v in filters.items() if v}))

    filters = {
        'college_id': request.args.get('college_id', type=int),
        'session_id': request.args.get('session_id', type=int),
        'degree_id': request.args.get('degree_id', type=int),
        'semester_id': request.args.get('semester_id', type=int),
        'year_id': request.args.get('year_id', type=int),
        'employee_id': request.args.get('employee_id'),
        'exconfig_id': request.args.get('exconfig_id', type=int),
        'coursetype': request.args.get('coursetype'),
        'batch_id': request.args.get('batch_id', type=int),
        'branch_id': request.args.get('branch_id', type=int)
    }
    if filters.get('coursetype') == '0':
        filters['coursetype'] = None

    # Auto-select college based on employee location (if only one match)
    if not filters['college_id'] and emp_id:
        emp_loc = DB.fetch_one("SELECT fk_locid FROM SAL_Employee_Mst WHERE pk_empid = ?", [emp_id])
        if emp_loc and emp_loc.get('fk_locid'):
            colleges = DB.fetch_all("SELECT pk_collegeid as id FROM SMS_College_Mst WHERE fk_locid = ?", [emp_loc['fk_locid']])
            if len(colleges) == 1:
                filters['college_id'] = colleges[0]['id']

    # Auto-select current session by sessionorder
    if not filters['session_id']:
        sess = DB.fetch_one("SELECT TOP 1 pk_sessionid FROM SMS_AcademicSession_Mst ORDER BY sessionorder DESC, pk_sessionid DESC")
        if sess:
            filters['session_id'] = sess['pk_sessionid']

    lookups = {
        'colleges': AcademicsModel.get_colleges_simple(),
        'sessions': InfrastructureModel.get_sessions(),
        'degrees': [],
        'semesters': [],
        'years': AcademicsModel.get_degree_years(),
        'employees': [],
        'exam_configs': [],
        'batches': []
    }

    # Only load teachers if a semester is selected as per user request
    if filters['semester_id']:
        lookups['employees'] = AcademicsModel.get_teaching_employees(None if include_all else hod_hr_dept_ids)

    if filters['college_id']:
        lookups['degrees'] = AcademicsModel.get_college_degrees(filters['college_id'])
    if filters['degree_id']:
        lookups['semesters'] = AcademicsModel.get_degree_semesters_for_degree(filters['degree_id'])
        lookups['exam_configs'] = CourseAllocationModel.get_exam_configs(
            filters['degree_id'], filters['session_id'], filters['semester_id']
        )
    if filters['college_id'] and filters['degree_id'] and filters['semester_id'] and filters.get('coursetype'):
        lookups['batches'] = BatchModel.get_batches(
            filters['college_id'], filters['degree_id'], filters['semester_id'], filters['coursetype']
        )

    courses = []
    selected_ids = set()
    main_ids = set()
    if filters['degree_id'] and filters['semester_id'] and filters.get('coursetype'):
        sms_dept_ids = dept_ctx.get('sms_dept_ids', [])
        # Revert to fetching available departmental courses for HOD
        raw_courses = AcademicsModel.get_courses_offered_by_hod(
            filters, filters.get('coursetype'), sms_dept_ids, code_prefixes
        )
        
        # Uniqueness and formatting
        seen = set()
        for c in raw_courses:
            th = c.get('crhr_theory') or 0
            pr = c.get('crhr_practical') or 0
            code = (c.get('coursecode') or '').strip()
            name = (c.get('coursename') or '').strip()
            key = (code.upper(), name.upper(), th, pr)
            if key not in seen:
                seen.add(key)
                c['display_name'] = f"{code},({th}+{pr}), /{name}"
                courses.append(c)

        # Fetch LATEST assignment for this specific combination
        master = AcademicsModel.get_teacher_course_assignment_master(filters)
        if master:
            selected_ids, main_ids = AcademicsModel.get_teacher_course_assignment_details(master['pk_tcourseallocid'])

    return render_template('academics/teacher_course_assignment.html',
                           lookups=lookups,
                           filters=filters,
                           courses=courses,
                           selected_ids=selected_ids,
                           main_ids=main_ids,
                           include_all=include_all,
                           hod_departments=dept_ctx.get('hr_departments', []))

@academics_bp.route('/teacher_course_assignment_report')
@permission_required('Teacher Course Assignment')
def teacher_course_assignment_report():
    filters = {
        'college_id': request.args.get('college_id', type=int),
        'session_id': request.args.get('session_id', type=int),
        'degree_id': request.args.get('degree_id', type=int),
        'semester_id': request.args.get('semester_id', type=int),
        'exconfig_id': request.args.get('exconfig_id', type=int),
        'branch_id': request.args.get('branch_id', type=int)
    }

    if not all([filters['college_id'], filters['session_id'], filters['degree_id'], filters['semester_id']]):
        flash('Please select all mandatory filters to print report.', 'warning')
        return redirect(url_for('academics.teacher_course_assignment', **{k: v for k, v in filters.items() if v}))

    report_data = AcademicsModel.get_teacher_course_assignment_report_data(filters)

    if not report_data:
        flash('No data found for the selected filters.', 'info')
        return redirect(url_for('academics.teacher_course_assignment', **{k: v for k, v in filters.items() if v}))

    from app.utils import generate_teacher_course_assignment_pdf
    pdf_out = generate_teacher_course_assignment_pdf(report_data)

    response = make_response(pdf_out.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'inline; filename=Teacher_Course_Assignment_{datetime.now().strftime("%Y%m%d%H%M")}.pdf'
    return response

@academics_bp.route('/course_teacher_assignment_report_page', methods=['GET', 'POST'])
@permission_required('Course Offer And Teacher Course Assignment Report(For HOD)')
def course_teacher_assignment_report_page():
    user_id = session.get('user_id')
    emp_id = session.get('emp_id')

    if request.method == 'POST':
        filters = {
            'college_id': request.form.get('college_id', type=int),
            'session_id': request.form.get('session_id', type=int),
            'degree_id': request.form.get('degree_id', type=int),
            'semester_id': request.form.get('semester_id', type=int),
            'branch_id': request.form.get('branch_id', type=int),
            'coursetype': request.form.get('coursetype'),
            'action': request.form.get('action'),
            'rpt_format': request.form.get('rpt_format', 'pdf')
        }
    else:
        filters = {
            'college_id': request.args.get('college_id', type=int),
            'session_id': request.args.get('session_id', type=int),
            'degree_id': request.args.get('degree_id', type=int),
            'semester_id': request.args.get('semester_id', type=int),
            'branch_id': request.args.get('branch_id', type=int),
            'coursetype': request.args.get('coursetype')
        }

    # Auto-select college based on selected location if not set
    if not filters['college_id']:
        loc_id = session.get('selected_loc')
        if loc_id:
            colleges = DB.fetch_all("SELECT pk_collegeid as id FROM SMS_College_Mst WHERE fk_locid = ?", [loc_id])
            if len(colleges) == 1:
                filters['college_id'] = colleges[0]['id']

    # Auto-select current session by sessionorder if not set
    if not filters['session_id']:
        sess = DB.fetch_one("SELECT TOP 1 pk_sessionid FROM SMS_AcademicSession_Mst ORDER BY sessionorder DESC, pk_sessionid DESC")
        if sess:
            filters['session_id'] = sess['pk_sessionid']

    lookups = {
        'colleges': AcademicsModel.get_colleges_simple(),
        'sessions': InfrastructureModel.get_sessions(),
        'degrees': [],
        'semesters': [],
        'branches': []
    }

    if filters['college_id']:
        lookups['degrees'] = AcademicsModel.get_college_degrees(filters['college_id'])
    if filters['degree_id']:
        lookups['semesters'] = AcademicsModel.get_degree_semesters_for_degree(filters['degree_id'])
        lookups['branches'] = AcademicsModel.get_degree_branches(filters['degree_id'])

    report_data = []
    grouped_data = {}

    # Process View or Export
    if all([filters['college_id'], filters['session_id'], filters['degree_id'], filters['semester_id']]):
        if request.method == 'POST' and filters.get('action') == 'VIEW':
            report_data = AcademicsModel.get_teacher_course_assignment_report_data(filters)

            if filters.get('coursetype') and filters['coursetype'] != '0':
                report_data = [r for r in report_data if (r.get('coursetype') or '').strip() == filters['coursetype']]

            if filters['rpt_format'] == 'pdf':
                from app.utils import generate_teacher_course_assignment_pdf
                pdf_out = generate_teacher_course_assignment_pdf(report_data)
                response = make_response(pdf_out.getvalue())
                response.headers['Content-Type'] = 'application/pdf'
                response.headers['Content-Disposition'] = f'inline; filename=Teacher_Course_Assignment_{datetime.now().strftime("%Y%m%d%H%M")}.pdf'
                return response
            elif filters['rpt_format'] == 'excel':
                from app.utils import generate_teacher_course_assignment_excel
                excel_out = generate_teacher_course_assignment_excel(report_data)
                response = make_response(excel_out.getvalue())
                response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                response.headers['Content-Disposition'] = f'attachment; filename=Teacher_Course_Assignment_{datetime.now().strftime("%Y%m%d%H%M")}.xlsx'
                return response
            elif filters['rpt_format'] == 'word':
                from app.utils import generate_teacher_course_assignment_word
                word_out = generate_teacher_course_assignment_word(report_data)
                response = make_response(word_out.getvalue())
                response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.wordprocessingml.document'
                response.headers['Content-Disposition'] = f'attachment; filename=Teacher_Course_Assignment_{datetime.now().strftime("%Y%m%d%H%M")}.docx'
                return response

            # Re-group for grid display
            for row in report_data:
                t_key = (row['teacher_code'], row['teacher_name'])
                if t_key not in grouped_data:
                    grouped_data[t_key] = []
                grouped_data[t_key].append(row)

    return render_template('academics/reports/course_teacher_assignment_report_page.html',
                           lookups=lookups,
                           filters=filters,
                           grouped_data=grouped_data)

@academics_bp.route('/batch_master', methods=['GET', 'POST'])
@permission_required('Class - Batch Master')
def batch_master():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'DELETE':
            if AcademicsModel.delete_batch(request.form.get('id')):
                flash('Batch record deleted successfully!', 'success')
            else:
                flash('Error deleting record.', 'danger')
        else:
            if AcademicsModel.save_batch(request.form):
                flash('Batch record saved successfully!', 'success')
            else:
                flash('Error saving record.', 'danger')
        return redirect(url_for('academics.batch_master'))
    
    page = request.args.get('page', 1, type=int)
    per_page = 10
    items, total = AcademicsModel.get_batches(page=page, per_page=per_page)
    
    pagination = {
        'page': page,
        'per_page': per_page,
        'total': total,
        'total_pages': math.ceil(total / per_page) if total else 1,
        'has_prev': page > 1,
        'has_next': page < (math.ceil(total / per_page) if total else 1)
    }
    
    page_range = get_pagination_range(page, pagination['total_pages'])

    lookups = {
        'colleges': AcademicsModel.get_colleges_simple(),
        'sessions': InfrastructureModel.get_sessions(),
        'degrees': AcademicsModel.get_all_degrees(),
        'semesters': InfrastructureModel.get_all_semesters(),
        'branches': AcademicsModel.get_branches(),
        'batch_types': AcademicsModel.get_batch_type_lookup()
    }
    
    return render_template('academics/batch_master.html', items=items, lookups=lookups, pagination=pagination, page_range=page_range)

@academics_bp.route('/degree_crhr', methods=['GET', 'POST'])
@permission_required('Degree wise Credit Hours')
def degree_crhr():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'DELETE':
            if AcademicsModel.delete_degree_crhr(request.form.get('id')):
                flash('Credit hour record deleted successfully!', 'success')
            else:
                flash('Error deleting record.', 'danger')
        else:
            if AcademicsModel.save_degree_crhr(request.form):
                flash('Credit hour record saved successfully!', 'success')
            else:
                flash('Error saving record.', 'danger')
        return redirect(url_for('academics.degree_crhr'))
    
    page = request.args.get('page', 1, type=int)
    per_page = 10
    items, total = AcademicsModel.get_degree_crhr(page=page, per_page=per_page)
    
    pagination = {
        'page': page,
        'per_page': per_page,
        'total': total,
        'total_pages': math.ceil(total / per_page) if total else 1,
        'has_prev': page > 1,
        'has_next': page < (math.ceil(total / per_page) if total else 1)
    }
    
    page_range = get_pagination_range(page, pagination['total_pages'])

    lookups = {
        'degrees': AcademicsModel.get_all_degrees(),
        'semesters': InfrastructureModel.get_all_semesters()
    }

    return render_template('academics/degree_crhr.html', items=items, lookups=lookups, pagination=pagination, page_range=page_range)

@academics_bp.route('/degree_crhr_courseplan', methods=['GET', 'POST'])
@permission_required('Degree Wise Credit Hours(Course Plan)')
def degree_crhr_courseplan():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'DELETE':
            if AcademicsModel.delete_degree_crhr_courseplan(request.form.get('id')):
                flash('Course Plan record deleted successfully!', 'success')
            else:
                flash('Error deleting record.', 'danger')
        else:
            if AcademicsModel.save_degree_crhr_courseplan(request.form):
                flash('Course Plan record saved successfully!', 'success')
            else:
                flash('Error saving record.', 'danger')
        return redirect(url_for('academics.degree_crhr_courseplan'))

    page = request.args.get('page', 1, type=int)
    per_page = 10
    items, total = AcademicsModel.get_degree_crhr_courseplan_paginated(page=page, per_page=per_page)

    pagination = {
        'page': page,
        'per_page': per_page,
        'total': total,
        'total_pages': math.ceil(total / per_page) if total else 1,
        'has_prev': page > 1,
        'has_next': page < (math.ceil(total / per_page) if total else 1)
    }

    page_range = get_pagination_range(page, pagination['total_pages'])

    lookups = {
        'degrees': AcademicsModel.get_all_degrees(),
        'course_types': ClassificationModel.get_course_types()
    }

    return render_template('academics/degree_crhr_courseplan.html', items=items, lookups=lookups, pagination=pagination, page_range=page_range)

@academics_bp.route('/api/degree_crhr_courseplan/<int:plan_id>')
def get_degree_crhr_courseplan_details_api(plan_id):
    master, details = AcademicsModel.get_degree_crhr_courseplan_details(plan_id)
    return jsonify({'master': master, 'details': details})

@academics_bp.route('/api/mapping/<map_id>/details')
def get_mapping_details_api(map_id):
    master = DB.fetch_one("SELECT * FROM SMS_CollegeDegreeBranchMap_Mst WHERE PK_Coldgbrid = ?", [map_id])
    details = AcademicsModel.get_mapping_details(map_id)
    return jsonify({'master': master, 'details': details})

@academics_bp.route('/col_deg_spec_master', methods=['GET', 'POST'])
@permission_required('Col-Deg-Specialization Master')
def col_deg_spec_master():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'DELETE':
            if AcademicsModel.delete_mapping(request.form.get('id')):
                flash('Mapping deleted successfully!', 'success')
            else:
                flash('Error deleting mapping.', 'danger')
        else:
            if AcademicsModel.save_mapping(request.form):
                flash('College-Degree-Specialization mapping saved successfully!', 'success')
            else:
                flash('Error saving mapping.', 'danger')
        return redirect(url_for('academics.col_deg_spec_master'))
    
    page = request.args.get('page', 1, type=int)
    per_page = 10
    items, total = AcademicsModel.get_college_degree_mappings_paginated(page=page, per_page=per_page)
    
    pagination = {
        'page': page,
        'per_page': per_page,
        'total': total,
        'total_pages': math.ceil(total / per_page) if total else 1,
        'has_prev': page > 1,
        'has_next': page < (math.ceil(total / per_page) if total else 1)
    }
    
    page_range = get_pagination_range(page, pagination['total_pages'])

    lookups = {
        'colleges': AcademicsModel.get_colleges_simple(),
        'degrees': AcademicsModel.get_all_degrees(),
        'branches': AcademicsModel.get_branches(),
        'spec_types': [{'id': 'Major', 'name': 'Major'}, {'id': 'Minor', 'name': 'Minor'}, {'id': 'Supporting', 'name': 'Supporting'}, {'id': 'Optional', 'name': 'Optional'}]
    }
    
    return render_template('academics/col_deg_spec_master.html', items=items, lookups=lookups, pagination=pagination, page_range=page_range)

@academics_bp.route('/generic/<path:page_name>')
def generic_page_handler(page_name):
    # Try to find a specific template for this page
    template_name = page_name.lower().replace(' ', '_').replace('[', '').replace(']', '').replace('/', '_') + '.html'
    try:
        # Check if template exists in academics folder
        return render_template(f'academics/{template_name}', title=page_name)
    except:
        return render_template('academics/generic_page.html', title=page_name)

@academics_bp.route('/activity_course_master', methods=['GET', 'POST'])
@permission_required('Activity Course Master')
def activity_course_master():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'DELETE':
            if CourseActivityModel.delete_course_activity(request.form.get('id')):
                flash('Activity Course deleted successfully!', 'success')
            else:
                flash('Error deleting activity course.', 'danger')
        else:
            if CourseActivityModel.save_course_activity(request.form):
                flash('Activity Course saved successfully!', 'success')
            else:
                flash('Error saving activity course.', 'danger')
        return redirect(url_for('academics.activity_course_master'))

    page = request.args.get('page', 1, type=int)
    per_page = 10
    items, total = CourseActivityModel.get_course_activities(page=page, per_page=per_page)
    
    pagination = {
        'page': page,
        'per_page': per_page,
        'total': total,
        'total_pages': math.ceil(total / per_page) if total else 1,
        'has_prev': page > 1,
        'has_next': page < (math.ceil(total / per_page) if total else 1)
    }
    
    page_range = get_pagination_range(page, pagination['total_pages'])

    lookups = {
        'sessions': InfrastructureModel.get_sessions(),
        'semesters': InfrastructureModel.get_all_semesters(),
        'activities': ActivityCertificateModel.get_activities(),
        'categories': CourseActivityModel.get_activity_categories()
    }
    
    return render_template('academics/activity_course_master.html', items=items, lookups=lookups, pagination=pagination, page_range=page_range)

@academics_bp.route('/activity_master', methods=['GET', 'POST'])
@permission_required('Activity Master')
def activity_master():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'DELETE':
            if ActivityCertificateModel.delete_activity(request.form.get('id')):
                flash('Activity deleted successfully!', 'success')
            else:
                flash('Error deleting activity.', 'danger')
        else:
            if ActivityCertificateModel.save_activity(request.form):
                flash('Activity saved successfully!', 'success')
            else:
                flash('Error saving activity.', 'danger')
        return redirect(url_for('academics.activity_master'))

    page = request.args.get('page', 1, type=int)
    per_page = 10
    items, total = ActivityCertificateModel.get_activities_paginated(page=page, per_page=per_page)
    
    pagination = {
        'page': page,
        'per_page': per_page,
        'total': total,
        'total_pages': math.ceil(total / per_page) if total else 1,
        'has_prev': page > 1,
        'has_next': page < (math.ceil(total / per_page) if total else 1)
    }
    
    page_range = get_pagination_range(page, pagination['total_pages'])
    
    return render_template('academics/activity_master.html', items=items, pagination=pagination, page_range=page_range)

@academics_bp.route('/api/get_activity_course_details/<int:ca_id>')
def get_activity_course_details_api(ca_id):
    details = CourseActivityModel.get_course_activity_details(ca_id)
    from app.utils import clean_json_data
    return jsonify(clean_json_data(details))

@academics_bp.route('/package_master', methods=['GET', 'POST'])
@permission_required('Package Master')
def package_master():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'DELETE':
            if PackageMasterModel.delete_package(request.form.get('id')):
                flash('Package deleted successfully!', 'success')
            else:
                flash('Error deleting package.', 'danger')
        else:
            if PackageMasterModel.save_package(request.form):
                flash('Package saved successfully!', 'success')
            else:
                flash('Error saving package.', 'danger')
        return redirect(url_for('academics.package_master'))

    page = request.args.get('page', 1, type=int)
    per_page = 10
    items, total = PackageMasterModel.get_packages(page=page, per_page=per_page)
    
    pagination = {
        'page': page,
        'per_page': per_page,
        'total': total,
        'total_pages': math.ceil(total / per_page) if total else 1,
        'has_prev': page > 1,
        'has_next': page < (math.ceil(total / per_page) if total else 1)
    }
    
    page_range = get_pagination_range(page, pagination['total_pages'])

    lookups = {
        'degrees': AcademicsModel.get_all_degrees(),
        'semesters': InfrastructureModel.get_all_semesters(),
        'sessions': InfrastructureModel.get_sessions(),
        'courses': DB.fetch_all("SELECT pk_courseid as id, coursecode + ' - ' + coursename as name FROM SMS_Course_Mst ORDER BY coursecode")
    }
    
    return render_template('academics/package_master.html', items=items, lookups=lookups, pagination=pagination, page_range=page_range)

@academics_bp.route('/api/get_degree_semesters/<int:degree_id>')
def get_degree_semesters_api(degree_id):
    # Restrict to I to VIII as per live system standards
    from app.models.academics import InfrastructureModel
    semesters = InfrastructureModel.get_all_semesters()
    from app.utils import clean_json_data
    return jsonify(clean_json_data(semesters))
@academics_bp.route('/api/get_package_details/<int:pid>')
def get_package_details_api(pid):
    details = PackageMasterModel.get_package_details(pid)
    from app.utils import clean_json_data
    return jsonify(clean_json_data(details))

@academics_bp.route('/board_master', methods=['GET', 'POST'])
@permission_required('Board Master')
def board_master():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'DELETE':
            if BoardMasterModel.delete_board(request.form.get('id')):
                flash('Board deleted successfully!', 'success')
            else:
                flash('Error deleting board.', 'danger')
        else:
            if BoardMasterModel.save_board(request.form):
                flash('Board saved successfully!', 'success')
            else:
                flash('Error saving board.', 'danger')
        return redirect(url_for('academics.board_master'))

    page = request.args.get('page', 1, type=int)
    per_page = 10
    items, total = BoardMasterModel.get_boards(page=page, per_page=per_page)
    
    pagination = {
        'page': page,
        'per_page': per_page,
        'total': total,
        'total_pages': math.ceil(total / per_page) if total else 1,
        'has_prev': page > 1,
        'has_next': page < (math.ceil(total / per_page) if total else 1)
    }
    
    page_range = get_pagination_range(page, pagination['total_pages'])
    
    return render_template('academics/board_master.html', items=items, pagination=pagination, page_range=page_range)

@academics_bp.route('/certificates_master', methods=['GET', 'POST'])
@permission_required('Certificates Master')
def certificates_master():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'DELETE':
            if CertificateMasterModel.delete_certificate(request.form.get('id')):
                flash('Certificate deleted successfully!', 'success')
            else:
                flash('Error deleting certificate.', 'danger')
        else:
            if CertificateMasterModel.save_certificate(request.form):
                flash('Certificate saved successfully!', 'success')
            else:
                flash('Error saving certificate.', 'danger')
        return redirect(url_for('academics.certificates_master'))

    page = request.args.get('page', 1, type=int)
    per_page = 10
    items, total = CertificateMasterModel.get_certificates(page=page, per_page=per_page)
    
    pagination = {
        'page': page,
        'per_page': per_page,
        'total': total,
        'total_pages': math.ceil(total / per_page) if total else 1,
        'has_prev': page > 1,
        'has_next': page < (math.ceil(total / per_page) if total else 1)
    }
    
    page_range = get_pagination_range(page, pagination['total_pages'])
    
    return render_template('academics/certificates_master.html', items=items, pagination=pagination, page_range=page_range)

@academics_bp.route('/syllabus_creation', methods=['GET', 'POST'])
@permission_required('Syllabus Creation [Master]')
def syllabus_creation():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'SAVE':
            if SyllabusModel.save_syllabus(request.form):
                flash('Syllabus saved successfully!', 'success')
            else:
                flash('Error saving syllabus.', 'danger')
        return redirect(url_for('academics.syllabus_creation'))

    sessions = InfrastructureModel.get_sessions()
    degrees = AcademicsModel.get_all_degrees()

    return render_template('academics/syllabus_creation.html',
                           sessions=sessions,
                           degrees=degrees)

@academics_bp.route('/api/get_syllabus_courses', methods=['POST'])
def get_syllabus_courses_api():
    filters = request.json
    courses = SyllabusModel.get_syllabus_courses(filters)
    return jsonify(clean_json_data(courses))

@academics_bp.route('/api/get_syllabus', methods=['POST'])
def get_syllabus_api():
    data = request.json
    syllabus = SyllabusModel.get_syllabus(data.get('degree_id'), data.get('session_from'), data.get('session_to'), data.get('course_id'))
    return jsonify({'syllabus': syllabus})
@academics_bp.route('/course_detail_report', methods=['GET', 'POST'])
@permission_required('Course Detail Report')
def course_detail_report():
    if request.method == 'POST':
        filters = {
            'session_from': request.form.get('session_from'),
            'session_upto': request.form.get('session_upto'),
            'degree_id': request.form.get('degree_id'),
            'dept_id': request.form.get('dept_id'),
            'semester_id': request.form.get('semester_id'),
            'year_id': request.form.get('year_id')
        }
        rpt_format = request.form.get('rpt_format', '1') # 1: View, 2: Excel, 3: PDF
        
        data = CourseModel.get_course_report_data(filters)
        
        if rpt_format == '2': # Excel
            if not data:
                flash('No data found for selected filters', 'warning')
                return redirect(url_for('academics.course_detail_report'))
            df = pd.DataFrame(data)
            # Rename columns for clarity in Excel
            mapping = {
                'coursecode': 'Course Code', 'coursename': 'Course Name', 
                'crhr_theory': 'Theory', 'crhr_practical': 'Practical',
                'total_crhr': 'Total Credits', 'coursetype': 'Course Type',
                'dept_name': 'Department', 'degreename': 'Degree',
                'semester_roman': 'Semester', 'compulsory': 'Compulsory',
                'status': 'Status', 'session_from': 'Session From', 'session_upto': 'Session Upto'
            }
            df = df.rename(columns=mapping)
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                df.to_excel(writer, index=False, sheet_name='CourseDetails')
            output.seek(0)
            
            response = make_response(output.getvalue())
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            response.headers['Content-Disposition'] = 'attachment; filename=Course_Detail_Report.xlsx'
            return response

        elif rpt_format == '3': # PDF
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from datetime import datetime
            import os

            filter_names = {
                'session': DB.fetch_scalar("SELECT sessionname FROM SMS_AcademicSession_Mst WHERE pk_sessionid = ?", [filters['session_from']]) if filters.get('session_from') and filters['session_from'] != '0' else 'All',
                'degree': DB.fetch_scalar("SELECT degreename FROM SMS_Degree_Mst WHERE pk_degreeid = ?", [filters['degree_id']]) if filters.get('degree_id') and filters['degree_id'] != '0' else 'All',
                'year': DB.fetch_scalar("SELECT degreeyear_char FROM SMS_DegreeYear_Mst WHERE pk_degreeyearid = ?", [filters['year_id']]) if filters.get('year_id') and filters['year_id'] != '0' else 'All',
                'semester': DB.fetch_scalar("SELECT semester_roman FROM SMS_Semester_Mst WHERE pk_semesterid = ?", [filters['semester_id']]) if filters.get('semester_id') and filters['semester_id'] != '0' else 'All'
            }

            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=50)
            elements = []
            styles = getSampleStyleSheet()

            # Custom Styles
            title_style = ParagraphStyle('TitleStyle', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=14, alignment=1)
            subtitle_style = ParagraphStyle('SubtitleStyle', parent=styles['Normal'], fontName='Helvetica', fontSize=11, alignment=1, spaceAfter=20)
            
            # Header with Logo
            logo_path = os.path.join(current_app.root_path, 'static', 'images', 'logo.png')
            header_data = []
            if os.path.exists(logo_path):
                img = Image(logo_path, width=1*inch, height=1*inch)
                header_data.append([img, Paragraph("CHAUDHARY CHARAN SINGH HARYANA AGRICULTURAL UNIVERSITY, HISAR", title_style)])
            else:
                header_data.append(["", Paragraph("CHAUDHARY CHARAN SINGH HARYANA AGRICULTURAL UNIVERSITY, HISAR", title_style)])
            
            header_table = Table(header_data, colWidths=[1.2*inch, 6*inch])
            header_table.setStyle(TableStyle([
                ('ALIGN', (0,0), (0,0), 'LEFT'),
                ('ALIGN', (1,0), (1,0), 'CENTER'),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ]))
            elements.append(header_table)
            elements.append(Spacer(1, 10))

            # Subtitle
            subtitle_text = f"Courses For Session <b>{filter_names['session']}</b>, Degree <b>{filter_names['degree']}</b>, Year <b>{filter_names['year']}</b>, Semester <b>{filter_names['semester']}</b>"
            elements.append(Paragraph(subtitle_text, subtitle_style))

            # Data Table
            table_data = [['S.No.', 'Course Code', 'Course Name', 'Cr Hr\nTheory', 'Cr Hr\nPractical']]
            for i, row in enumerate(data, 1):
                table_data.append([
                    str(i),
                    row.get('coursecode', ''),
                    Paragraph(row.get('coursename', ''), styles['Normal']),
                    str(row.get('crhr_theory', '0')),
                    str(row.get('crhr_practical', '0'))
                ])

            t = Table(table_data, colWidths=[0.5*inch, 1.5*inch, 3.8*inch, 0.8*inch, 0.8*inch])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.white),
                ('TEXTCOLOR', (0,0), (-1,0), colors.black),
                ('ALIGN', (0,0), (-1,0), 'CENTER'),
                ('ALIGN', (0,1), (0,-1), 'CENTER'), # SNo
                ('ALIGN', (1,1), (1,-1), 'LEFT'),   # Code
                ('ALIGN', (2,1), (2,-1), 'LEFT'),   # Name
                ('ALIGN', (3,1), (-1,-1), 'CENTER'),# CrHr
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('FONTSIZE', (0,0), (-1,-1), 9),
                ('VALIGN', (0,0), (-1,-1), 'TOP'),
                ('GRID', (0,0), (-1,-1), 0.5, colors.black),
                ('BOTTOMPADDING', (0,0), (-1,0), 6),
                ('TOPPADDING', (0,0), (-1,0), 6),
            ]))
            elements.append(t)

            # Define the footer function
            def add_footer(canvas, doc):
                canvas.saveState()
                canvas.setFont('Helvetica', 9)
                page_num = canvas.getPageNumber()
                
                # Page number
                canvas.drawCentredString(A4[0]/2.0, 20, f"Page {page_num}")
                
                # Signatures
                canvas.setFont('Helvetica-Bold', 9)
                y_pos = 60
                date_str = datetime.now().strftime('%d/%m/%Y')
                
                # Left side
                canvas.drawString(40, y_pos + 15, f"Date : {date_str}")
                canvas.drawString(40, y_pos, "Place :")
                
                # Center
                canvas.drawCentredString(A4[0]/2.0, y_pos, "REGISTRAR")
                
                # Right side
                canvas.drawRightString(A4[0] - 40, y_pos + 15, "DIRECTOR OF ACADEMICS AND")
                canvas.drawRightString(A4[0] - 40, y_pos, "RESEARCH")
                
                canvas.restoreState()

            doc.build(elements, onFirstPage=add_footer, onLaterPages=add_footer)
            buffer.seek(0)
            
            response = make_response(buffer.getvalue())
            response.headers['Content-Type'] = 'application/pdf'
            response.headers['Content-Disposition'] = 'attachment; filename=Course_Detail_Report.pdf'
            return response

        # Default: Just show the data in the template
        sessions = InfrastructureModel.get_sessions()
        degrees = AcademicsModel.get_all_degrees()
        years = AcademicsModel.get_degree_years()
        semesters = InfrastructureModel.get_all_semesters()
        return render_template('academics/course_detail_report.html', 
                               data=data, filters=filters,
                               sessions=sessions, degrees=degrees, years=years, semesters=semesters)

    sessions = InfrastructureModel.get_sessions()
    degrees = AcademicsModel.get_all_degrees()
    years = AcademicsModel.get_degree_years()
    semesters = InfrastructureModel.get_all_semesters()
    return render_template('academics/course_detail_report.html', 
                           sessions=sessions, degrees=degrees, years=years, semesters=semesters)

@academics_bp.route('/api/get_degree_details/<int:degree_id>')
def get_degree_details_api(degree_id):
    # Returns minsem, maxsem, type (UG/PG), isdepartmentreq
    sql = """
    SELECT d.minsem, d.maxsem, dt.degreetype, d.isdepartmentreq, dt.isug
    FROM SMS_Degree_Mst d
    LEFT JOIN SMS_DegreeType_Mst dt ON d.fk_degreetypeid = dt.pk_degreetypeid
    WHERE d.pk_degreeid = ?
    """
    row = DB.fetch_one(sql, [degree_id])
    if row:
        return jsonify(row)
    return jsonify({})

@academics_bp.route('/api/degree/<int:degree_id>/branches')
def get_degree_branches_api(degree_id):
    branches = AcademicsModel.get_degree_branches(degree_id)
    from app.utils import clean_json_data
    return jsonify(clean_json_data(branches))

@academics_bp.route('/api/college/<int:college_id>/degree/<int:degree_id>/specializations')
def get_college_degree_specializations_api(college_id, degree_id):
    # Try the specific college-degree mapping first
    branches = AcademicsModel.get_college_degree_specializations(college_id, degree_id)
    if not branches:
        # Fallback to degree branches
        branches = AcademicsModel.get_degree_branches(degree_id)
    if not branches:
        # Final fallback: return all branches to ensure the dropdown works
        from app.db import DB
        branches = DB.fetch_all("SELECT Pk_BranchId as id, Branchname as name FROM SMS_BranchMst ORDER BY Branchname")
        
    from app.utils import clean_json_data
    return jsonify(clean_json_data(branches))

@academics_bp.route('/api/degree/<int:degree_id>/departments')
def get_degree_departments_api(degree_id):
    depts = AcademicsModel.get_degree_departments(degree_id)
    from app.utils import clean_json_data
    return jsonify(clean_json_data(depts))

@academics_bp.route('/api/get_semesters_range/<int:min_sem>/<int:max_sem>')
def get_semesters_range_api(min_sem, max_sem):
    sql = "SELECT pk_semesterid, semester_roman, semester_char FROM SMS_Semester_Mst WHERE semesterorder BETWEEN ? AND ? ORDER BY semesterorder"
    rows = DB.fetch_all(sql, [min_sem, max_sem])
    from app.utils import clean_json_data
    return jsonify(clean_json_data(rows))

@academics_bp.route('/api/get_semester_year/<int:semester_id>')
def get_semester_year_api(semester_id):
    sql = """
    SELECT dy.degreeyear_char, dy.pk_degreeyearid
    FROM SMS_Semester_Mst s
    INNER JOIN SMS_DegreeYear_Mst dy ON s.fk_degreeyearid = dy.pk_degreeyearid
    WHERE s.pk_semesterid = ?
    """
    row = DB.fetch_one(sql, [semester_id])
    from app.utils import clean_json_data
    return jsonify(clean_json_data(row if row else {}))

@academics_bp.route('/api/get_courses_filtered')
def get_courses_filtered_api():
    filters = {
        'degree_id': request.args.get('degree_id'),
        'semester_id': request.args.get('semester_id'),
        'branch_id': request.args.get('branch_id'),
        'session_id': request.args.get('session_id')
    }
    courses = CourseModel.get_courses_filtered(filters)
    from app.utils import clean_json_data
    return jsonify(clean_json_data(courses))

@academics_bp.route('/api/get_exam_configs')
def get_exam_configs_api():
    degree_id = request.args.get('degree_id', type=int)
    session_id = request.args.get('session_id', type=int)
    semester_id = request.args.get('semester_id', type=int)
    if not degree_id:
        return jsonify([])
    configs = CourseAllocationModel.get_exam_configs(degree_id, session_id, semester_id)
    return jsonify(clean_json_data(configs))

@academics_bp.route('/api/get_batches')
def get_batches_with_type_api():
    college_id = request.args.get('college_id', type=int)
    degree_id = request.args.get('degree_id', type=int)
    semester_id = request.args.get('semester_id', type=int)
    type_tp = request.args.get('type_tp') or 'T'
    if not college_id or not degree_id or not semester_id:
        return jsonify([])
    batches = BatchModel.get_batches(college_id, degree_id, semester_id, type_tp)
    return jsonify(clean_json_data(batches))

@academics_bp.route('/api/get_all_departments')
def get_all_departments_api():
    depts = DB.fetch_all("SELECT Pk_BranchId as id, Branchname as name FROM SMS_BranchMst ORDER BY Branchname")
    return jsonify(depts)

@academics_bp.route('/minor_advisor', methods=['GET', 'POST'])
@permission_required('Member of Minor and supporting')
def minor_advisor():
    if request.method == 'POST':
        action = request.form.get('action')
        sid = request.form.get('sid')
        advisor_id = request.form.get('advisor_id')
        role_id = request.form.get('role_id')
        user_id = session['user_id']

        if action == 'SUBMIT' and sid:
            DB.execute("UPDATE SMS_Advisory_Committee_Mst SET submitdate = GETDATE() WHERE fk_stid = ?", [sid])
            flash('Committee submitted successfully!', 'success')
            return redirect(url_for('academics.minor_advisor', **request.args))

        if sid and advisor_id and role_id:
            # Upsert into unified committee structure
            mst = DB.fetch_one("SELECT pk_adcid FROM SMS_Advisory_Committee_Mst WHERE fk_stid = ?", [sid])
            if not mst:
                stu = DB.fetch_one("SELECT fk_collegeid, fk_adm_session, fk_degreeid, fk_branchid FROM SMS_Student_Mst WHERE pk_sid=?", [sid])
                if stu:
                    res = DB.execute("""
                        INSERT INTO SMS_Advisory_Committee_Mst (fk_colgid, fk_sessionid, fk_degreeid, fk_branchid, fk_stid, createdby, creationdate, approvalstatus)
                        VALUES (?, ?, ?, ?, ?, ?, GETDATE(), 'P')
                    """, [stu['fk_collegeid'], stu['fk_adm_session'], stu['fk_degreeid'], stu['fk_branchid'], sid, user_id])
                    mst = DB.fetch_one("SELECT pk_adcid FROM SMS_Advisory_Committee_Mst WHERE fk_stid = ?", [sid])

            if mst:
                adcid = mst['pk_adcid']
                DB.execute("DELETE FROM SMS_Advisory_Committee_Dtl WHERE fk_adcid=? AND fk_statusid=?", [adcid, role_id])
                DB.execute("INSERT INTO SMS_Advisory_Committee_Dtl (fk_adcid, fk_statusid, fk_empid) VALUES (?, ?, ?)",
                           [adcid, role_id, advisor_id])
                if action == 'UPDATE':
                    flash('Committee member updated successfully!', 'success')
                else:
                    flash('Committee member assigned successfully!', 'success')
            else:
                flash('Error creating advisory master.', 'danger')
        return redirect(url_for('academics.minor_advisor', **request.args))
    college_id = request.args.get('college_id')
    loc_id = session.get('selected_loc')
    colleges = DB.fetch_all("SELECT pk_collegeid as id, collegename as name FROM SMS_College_Mst WHERE fk_locid = ? ORDER BY collegename", [loc_id]) if loc_id else AcademicsModel.get_colleges_simple()
    
    if not college_id and colleges:
        college_id = str(colleges[0]['id'])

    session_id = request.args.get('session_id')
    if not session_id:
        curr_session = InfrastructureModel.get_current_session_id()
        session_id = str(curr_session) if curr_session else None

    degree_id = request.args.get('degree_id')
    branch_id = request.args.get('branch_id')
    page = request.args.get('page', 1, type=int)
    per_page = 10
    
    students = []
    total_students = 0
    if college_id:
        filters = {
            'college_id': college_id,
            'session_id': session_id,
            'degree_id': degree_id,
            'branch_id': branch_id
        }
        students, total_students = AdvisoryModel.get_students_for_advisory(filters, page=page, per_page=per_page)

    import math
    pagination = {
        'page': page,
        'per_page': per_page,
        'total': total_students,
        'total_pages': math.ceil(total_students / per_page) if total_students else 1,
        'has_prev': page > 1,
        'has_next': page < (math.ceil(total_students / per_page) if total_students else 1)
    }

    lookups = {
        'colleges': colleges,
        'sessions': InfrastructureModel.get_sessions(),
        'degrees': AcademicsModel.get_college_pg_degrees(college_id) if college_id else [],
        'branches': AcademicsModel.get_college_degree_specializations(college_id, degree_id) if (college_id and degree_id and str(degree_id) != '0') else [],
        'employees': DB.fetch_all("SELECT E.pk_empid as id, E.empname + ' || ' + ISNULL(E.empcode, '') + ' (' + ISNULL(D.description, 'No Dept') + ')' as name FROM SAL_Employee_Mst E LEFT JOIN Department_Mst D ON E.fk_deptid = D.pk_deptid WHERE E.employeeleftstatus = 'N' ORDER BY E.empname")
    }

    active_filters = {
        'college_id': college_id,
        'session_id': session_id,
        'degree_id': degree_id,
        'branch_id': branch_id
    }
    
    return render_template('academics/minor_advisor.html',
                           lookups=lookups,
                           students=clean_json_data(students),
                           grid_students=clean_json_data(students),
                           pagination=pagination,
                           filters=active_filters)
@academics_bp.route('/minor_advisor_report/<int:sid>')
@permission_required('Member of Minor and supporting')
def minor_advisor_report(sid):
    from flask import make_response
    from app.utils_pdf import generate_advisory_committee_report
    
    # 1. Fetch Student Details
    stu_query = """
        SELECT S.fullname, S.AdmissionNo, S.enrollmentno, 
               CLG.collegename, DEPT.description as department_name,
               B_MAJ.Branchname as major_name, B_MIN.Branchname as minor_name, B_SUP.Branchname as supporting_name,
               ACM.hod_date, ACM.Collegedean_date, ACM.deanpgs_date,
               E_HOD.empname as hod_name, E_DEAN.empname as dean_name, E_PGS.empname as deanpgs_name
        FROM SMS_Student_Mst S
        LEFT JOIN SMS_College_Mst CLG ON S.fk_collegeid = CLG.pk_collegeid
        LEFT JOIN SMS_stuDiscipline_dtl SD ON S.pk_sid = SD.fk_sturegid
        LEFT JOIN SMS_BranchMst B_MAJ ON SD.fk_desciplineidMajor = B_MAJ.Pk_BranchId
        LEFT JOIN Department_Mst DEPT ON B_MAJ.fk_deptidDdo = DEPT.pk_deptid
        LEFT JOIN SMS_BranchMst B_MIN ON SD.fk_desciplineidMinor = B_MIN.Pk_BranchId
        LEFT JOIN SMS_BranchMst B_SUP ON SD.fk_desciplineidSupporting = B_SUP.Pk_BranchId
        LEFT JOIN SMS_Advisory_Committee_Mst ACM ON S.pk_sid = ACM.fk_stid
        LEFT JOIN SAL_Employee_Mst E_HOD ON ACM.hod_id = E_HOD.pk_empid
        LEFT JOIN SAL_Employee_Mst E_DEAN ON ACM.collegedean_id = E_DEAN.pk_empid
        LEFT JOIN SAL_Employee_Mst E_PGS ON ACM.deanpgs_id = E_PGS.pk_empid
        WHERE S.pk_sid = ?
    """
    student_info = DB.fetch_one(stu_query, [sid])
    if not student_info:
        flash("Student not found for report.", "danger")
        return redirect(url_for('academics.minor_advisor'))

    # 2. Fetch Committee Details
    com_query = """
        SELECT ACD.*, E.empname + ' || ' + ISNULL(E.empcode, '') as advisor_name,
               DEPT.description as department, DESG.designation,
               (SELECT TOP 1 B.Branchname FROM SMS_BranchMst B WHERE B.fk_deptidDdo = E.fk_deptid) as specialization,
               CASE ACD.fk_statusid
                    WHEN 1 THEN 'Major Advisor' WHEN 2 THEN 'Co-Advisor'
                    WHEN 3 THEN 'Member From Minor Subject' WHEN 4 THEN 'Member From Supporting Subject'
                    WHEN 5 THEN 'Dean PGS Nominee' WHEN 6 THEN 'Member From Major Subject'
                    ELSE 'Member'
               END as role_name
        FROM SMS_Advisory_Committee_Dtl ACD
        INNER JOIN SMS_Advisory_Committee_Mst ACM ON ACD.fk_adcid = ACM.pk_adcid
        INNER JOIN SAL_Employee_Mst E ON ACD.fk_empid = E.pk_empid
        LEFT JOIN Department_Mst DEPT ON E.fk_deptid = DEPT.pk_deptid
        LEFT JOIN SAL_Designation_Mst DESG ON E.fk_desgid = DESG.pk_desgid
        WHERE ACM.fk_stid = ?
    """
    committee_data = DB.fetch_all(com_query, [sid])
    
    buffer = generate_advisory_committee_report(student_info, committee_data)
    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=Committee_Report_{sid}.pdf'
    return response

@academics_bp.route('/api/student/<int:sid>/advisory_committee')
def get_student_advisory_committee_api(sid):
    info = AdvisoryModel.get_student_advisory_committee(sid)
    if not info:
        return jsonify({'error': 'Not found'}), 404
    return jsonify(clean_json_data(info))

@academics_bp.route('/api/student/<int:sid>/specializations')
def get_student_specializations_api(sid):
    # Fetch Disciplines (Major, Minor, Supporting)
    disc_sql = """
        SELECT fk_desciplineidMajor as major_id, fk_desciplineidMinor as minor_id, 
               fk_desciplineidSupporting as supporting_id
        FROM SMS_stuDiscipline_dtl
        WHERE fk_sturegid = ?
    """
    disc = DB.fetch_one(disc_sql, [sid])
    
    # Fetch Major Advisor (statusid = 1)
    adv_sql = """
        SELECT ACD.fk_empid as advisor_id
        FROM SMS_Advisory_Committee_Dtl ACD
        INNER JOIN SMS_Advisory_Committee_Mst ACM ON ACD.fk_adcid = ACM.pk_adcid
        WHERE ACM.fk_stid = ? AND ACD.fk_statusid = 1
    """
    adv = DB.fetch_one(adv_sql, [sid])
    
    res = {
        'major_id': disc['major_id'] if disc else None,
        'minor_id': disc['minor_id'] if disc else None,
        'supporting_id': disc['supporting_id'] if disc else None,
        'advisor_id': adv['advisor_id'] if adv else None
    }
    return jsonify(res)

@academics_bp.route('/api/student/<int:sid>/committee')
def get_student_committee_api(sid):
    filter_type = request.args.get('type', 'all')
    sql = """
        SELECT ACD.*, S.fullname as student_name, 
               E.empname + ' || ' + E.empcode as advisor_name,
               DESG.designation as designation, DEPT.description as department,
               B.Branchname as specialization,
               CASE ACD.fk_statusid
                    WHEN 1 THEN 'Major Advisor' WHEN 2 THEN 'Co-Advisor'
                    WHEN 3 THEN 'Member From Minor Subject' WHEN 4 THEN 'Member From Supporting Subject'
                    WHEN 5 THEN 'Dean PGS Nominee' WHEN 6 THEN 'Member From Major Subject'
                    ELSE 'Member'
               END as role_name
        FROM SMS_Advisory_Committee_Dtl ACD
        INNER JOIN SMS_Advisory_Committee_Mst ACM ON ACD.fk_adcid = ACM.pk_adcid
        INNER JOIN SMS_Student_Mst S ON ACM.fk_stid = S.pk_sid
        INNER JOIN SAL_Employee_Mst E ON ACD.fk_empid = E.pk_empid
        LEFT JOIN SAL_Designation_Mst DESG ON E.fk_desgid = DESG.pk_desgid
        LEFT JOIN Department_Mst DEPT ON E.fk_deptid = DEPT.pk_deptid
        LEFT JOIN SMS_BranchMst B ON S.fk_branchid = B.Pk_BranchId
        WHERE ACM.fk_stid = ?
    """
    if filter_type == 'major':
        sql += " AND ACD.fk_statusid = 1"
    elif filter_type == 'minor':
        sql += " AND ACD.fk_statusid != 5"
        
    sql += " ORDER BY ACD.fk_statusid"
    rows = DB.fetch_all(sql, [sid])
    return jsonify(clean_json_data(rows))

@academics_bp.route('/api/advisor/<string:advisor_id>/students')
def get_advisor_students_api(advisor_id):
    sql = """
        SELECT S.fullname, S.AdmissionNo, S.enrollmentno, DEG.degreename, B.Branchname
        FROM SMS_Advisory_Committee_Dtl ACD
        JOIN SMS_Advisory_Committee_Mst ACM ON ACD.fk_adcid = ACM.pk_adcid
        JOIN SMS_Student_Mst S ON ACM.fk_stid = S.pk_sid
        LEFT JOIN SMS_Degree_Mst DEG ON S.fk_degreeid = DEG.pk_degreeid
        LEFT JOIN SMS_BranchMst B ON S.fk_branchid = B.Pk_BranchId
        WHERE ACD.fk_empid = ? AND ACD.fk_statusid = 1
        ORDER BY S.fullname
    """
    rows = DB.fetch_all(sql, [advisor_id])
    return jsonify(clean_json_data(rows))

@academics_bp.route('/specialization_assignment', methods=['GET', 'POST'])
@permission_required('Specialization Assignment')
def specialization_assignment():
    if request.method == 'POST':
        if AdvisoryModel.save_student_discipline(request.form, session['user_id']):
            flash('Specialization assigned successfully!', 'success')
        else:
            flash('Error assigning specialization. Please check all fields.', 'danger')
        return redirect(url_for('academics.specialization_assignment', **request.args))

    college_id = request.args.get('college_id')
    loc_id = session.get('selected_loc')
    colleges = DB.fetch_all("SELECT pk_collegeid as id, collegename as name FROM SMS_College_Mst WHERE fk_locid = ? ORDER BY collegename", [loc_id]) if loc_id else AcademicsModel.get_colleges_simple()

    if not college_id and colleges:
        college_id = str(colleges[0]['id'])

    session_id = request.args.get('session_id')
    if not session_id:
        curr_session = InfrastructureModel.get_current_session_id()
        session_id = str(curr_session) if curr_session else None

    degree_id = request.args.get('degree_id')
    branch_id = request.args.get('filter_branch_id')
    page = request.args.get('page', 1, type=int)
    per_page = 10

    students = []
    total_students = 0
    if college_id:
        filters = {
            'college_id': college_id,
            'session_id': session_id,
            'degree_id': degree_id,
            'branch_id': branch_id
        }
        students, total_students = AdvisoryModel.get_students_for_advisory(filters, page=page, per_page=per_page)

    import math
    pagination = {
        'page': page,
        'per_page': per_page,
        'total': total_students,
        'total_pages': math.ceil(total_students / per_page) if total_students else 1,
        'has_prev': page > 1,
        'has_next': page < (math.ceil(total_students / per_page) if total_students else 1)
    }

    lookups = {
        'colleges': colleges,
        'sessions': InfrastructureModel.get_sessions(),
        'degrees': AcademicsModel.get_college_pg_degrees(college_id) if college_id else [],
        'branches': AcademicsModel.get_college_degree_specializations(college_id, degree_id) if (college_id and degree_id and str(degree_id) != '0') else [],
        'all_branches': AcademicsModel.get_branches()
    }

    # Context info for grid labels
    degree_name = ''
    session_name = ''
    if degree_id:
        deg = next((d for d in lookups['degrees'] if str(d['id']) == str(degree_id)), None)
        degree_name = deg['name'] if deg else ''
    if session_id:
        sess = next((s for s in lookups['sessions'] if str(s['id']) == str(session_id)), None)
        session_name = sess['name'] if sess else ''

    active_filters = {
        'college_id': college_id,
        'session_id': session_id,
        'degree_id': degree_id,
        'filter_branch_id': branch_id
    }

    return render_template('academics/specialization_assignment.html',
                           lookups=lookups,
                           students=clean_json_data(students),
                           pagination=pagination,
                           filters=active_filters,
                           degree_name=degree_name,
                           session_name=session_name)
@academics_bp.route('/admission_no_configuration', methods=['GET', 'POST'])
@permission_required('Admission No Configuration')
def admission_no_configuration():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'DELETE':
            if AdmissionModel.delete_admission_config(request.form.get('id')):
                flash('Configuration deleted successfully!', 'success')
            else:
                flash('Error deleting configuration.', 'danger')
        else:
            if AdmissionModel.save_admission_config(request.form):
                flash('Configuration saved successfully!', 'success')
            else:
                flash('Error saving configuration.', 'danger')
        return redirect(url_for('academics.admission_no_configuration'))

    page = request.args.get('page', 1, type=int)
    per_page = 10
    loc_id = session.get('selected_loc')
    items, total = AdmissionModel.get_all_configs(page=page, per_page=per_page, loc_id=loc_id)
    
    pagination = {
        'page': page,
        'per_page': per_page,
        'total': total,
        'total_pages': math.ceil(total / per_page) if total else 1,
        'has_prev': page > 1,
        'has_next': page < (math.ceil(total / per_page) if total else 1)
    }
    
    page_range = get_pagination_range(page, pagination['total_pages'])

    if loc_id:
        colleges = DB.fetch_all("SELECT pk_collegeid as id, collegename as name FROM SMS_College_Mst WHERE fk_locid = ? ORDER BY collegename", [loc_id])
    else:
        colleges = AcademicsModel.get_colleges_simple()

    college_id = request.args.get('college_id')
    filters = {
        'college_id': college_id
    }
    lookups = {
        'colleges': colleges,
        'degrees': AcademicsModel.get_college_all_degrees(college_id) if college_id else AcademicsModel.get_all_degrees(),
        'sessions': InfrastructureModel.get_sessions()
    }
    
    return render_template('academics/admission_no_configuration.html', items=items, lookups=lookups, pagination=pagination, page_range=page_range, filters=filters)

@academics_bp.route('/api/admission/get_config')
def api_get_admission_config():
    college_id = request.args.get('college_id')
    degree_id = request.args.get('degree_id')
    session_id = request.args.get('session_id')
    if not all([college_id, degree_id, session_id]):
        return jsonify({'error': 'Missing parameters'}), 400
    
    config = AdmissionModel.get_admission_config(college_id, degree_id, session_id)
    if config:
        # HAU ERP reversed logic: suffix col is Prefix, Prefix col is Suffix
        return jsonify({
            'prefix': config['suffix'] or '',
            'suffix': config['Prefix'] or '',
            'separator': config['Separator'] or ''
        })
    return jsonify({'error': 'No config found'}), 404

@academics_bp.route('/admission_no_generation', methods=['GET', 'POST'])
@permission_required('Admission No Generation')
def admission_no_generation():
    if request.method == 'POST':
        action = request.form.get('action')
        student_ids = request.form.getlist('student_ids')
        admission_nos = request.form.getlist('new_nos[]')
        
        if action == 'SAVE':
            success_count = 0
            for i in range(len(student_ids)):
                if admission_nos[i]:
                    if AdmissionModel.save_admission_no(student_ids[i], admission_nos[i]):
                        success_count += 1
            flash(f'Successfully updated {success_count} admission numbers.', 'success')
        
        return redirect(url_for('academics.admission_no_generation', **request.args))

    filters = {
        'college_id': request.args.get('college_id'),
        'session_id': request.args.get('session_id'),
        'degree_id': request.args.get('degree_id'),
        'branch_id': request.args.get('branch_id')
    }
    
    students = []
    config = None
    next_serial = 1
    
    if all([filters['college_id'], filters['session_id'], filters['degree_id']]):
        students = AdmissionModel.get_students_for_generation(filters)
        config_raw = AdmissionModel.get_admission_config(filters['college_id'], filters['degree_id'], filters['session_id'])
        if config_raw:
            config = {
                'prefix': config_raw['suffix'] or '',
                'suffix': config_raw['Prefix'] or '',
                'separator': config_raw['Separator'] or ''
            }
            next_serial = AdmissionModel.get_next_serial(
                filters['college_id'], filters['degree_id'], filters['session_id'],
                config['prefix'], config['suffix'], config['separator']
            )
        else:
            flash('No admission number configuration found for this selection.', 'warning')

    # Lookups
    loc_id = session.get('selected_loc')
    if loc_id:
        colleges = DB.fetch_all("SELECT pk_collegeid as id, collegename as name FROM SMS_College_Mst WHERE fk_locid = ? ORDER BY collegename", [loc_id])
    else:
        colleges = AcademicsModel.get_colleges_simple()

    lookups = {
        'colleges': colleges,
        'sessions': InfrastructureModel.get_sessions(),
        'degrees': AcademicsModel.get_college_all_degrees(filters['college_id']) if filters['college_id'] else [],
        'branches': AcademicsModel.get_college_degree_specializations(filters['college_id'], filters['degree_id']) if (filters['college_id'] and filters['degree_id']) else []
    }
    
    return render_template('academics/admission_no_generation.html', 
                           lookups=lookups, filters=filters, students=students, 
                           config=config, next_serial=next_serial)

    pagination = {
        'page': page,
        'per_page': per_page,
        'total': total,
        'total_pages': math.ceil(total / per_page) if total else 1,
        'has_prev': page > 1,
        'has_next': page < (math.ceil(total / per_page) if total else 1)
    }
    
    page_range = get_pagination_range(page, pagination['total_pages'])

    colleges = AcademicsModel.get_colleges_simple()
    sessions = InfrastructureModel.get_sessions()
    degrees = AcademicsModel.get_all_degrees()
    
    # Create a copy of filters and remove 'page' if it exists
    page_filters = dict(request.args)
    if 'page' in page_filters:
        del page_filters['page']
    
    return render_template('academics/admission_no_generation.html', 
                           colleges=colleges, sessions=sessions, degrees=degrees,
                           students=students, config=config, filters=page_filters,
                           pagination=pagination, page_range=page_range)

@academics_bp.route('/college_degree_seat_detail', methods=['GET', 'POST'])
@permission_required(' College Degree Wise Seat Detail')
def college_degree_seat_detail():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'DELETE':
            if SeatDetailModel.delete_seat_detail(request.form.get('id')):
                flash('Seat detail deleted successfully!', 'success')
            else:
                flash('Error deleting seat detail.', 'danger')
        else:
            if SeatDetailModel.save_seat_detail(request.form):
                flash('Seat detail saved successfully!', 'success')
            else:
                flash('Error saving seat detail.', 'danger')
        return redirect(url_for('academics.college_degree_seat_detail'))

    page = request.args.get('page', 1, type=int)
    per_page = 10
    items, total = SeatDetailModel.get_seat_details(page=page, per_page=per_page)
    
    pagination = {
        'page': page,
        'per_page': per_page,
        'total': total,
        'total_pages': math.ceil(total / per_page) if total else 1,
        'has_prev': page > 1,
        'has_next': page < (math.ceil(total / per_page) if total else 1)
    }
    
    page_range = get_pagination_range(page, pagination['total_pages'])

    lookups = {
        'colleges': AcademicsModel.get_colleges_simple(),
        'degrees': AcademicsModel.get_all_degrees(),
        'branches': AcademicsModel.get_branches(),
        'sessions': InfrastructureModel.get_sessions()
    }
    
    return render_template('academics/college_degree_seat_detail.html', items=items, lookups=lookups, pagination=pagination, page_range=page_range)

@academics_bp.route('/api/get_pgs_limit/<int:limit_id>')
def get_pgs_limit_api(limit_id):
    record = PgsCourseLimitModel.get_limit_by_id(limit_id)
    if record:
        return jsonify(clean_json_data(record))
    return jsonify({'error': 'Record not found'}), 404

@academics_bp.route('/pgs_course_limit_master', methods=['GET', 'POST'])
@permission_required('PGS Course Limit [Master]')
def pgs_course_limit_master():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'DELETE':
            if PgsCourseLimitModel.delete_limit(request.form.get('id')):
                flash('Limit deleted successfully!', 'success')
            else:
                flash('Error deleting limit.', 'danger')
        else:
            data = request.form.to_dict()
            data['user_id'] = session['user_id']
            if PgsCourseLimitModel.save_limit(data):
                flash('Course limit saved successfully!', 'success')
            else:
                flash('Error saving course limit. Check all fields.', 'danger')
        return redirect(url_for('academics.pgs_course_limit_master'))

    page = request.args.get('page', 1, type=int)
    per_page = 10
    items, total = PgsCourseLimitModel.get_limits(page=page, per_page=per_page)
    
    pagination = {
        'page': page,
        'per_page': per_page,
        'total': total,
        'total_pages': math.ceil(total / per_page) if total else 1,
        'has_prev': page > 1,
        'has_next': page < (math.ceil(total / per_page) if total else 1)
    }
    
    page_range = get_pagination_range(page, pagination['total_pages'])

    # Fetch lookup data
    colleges = AcademicsModel.get_colleges_simple()
    # Filter courses to show only relevant ones if needed, or all.
    # Restricted to only FIVE PGS courses as requested
    courses = DB.fetch_all("""
        SELECT pk_courseid as id, '[' + coursecode + ']~' + coursename as name
        FROM SMS_Course_Mst
        WHERE pk_courseid IN (1142, 1147, 1144, 1145, 1143)
        ORDER BY coursecode
    """)

    sessions = InfrastructureModel.get_sessions()
    return render_template('academics/pgs_course_limit_master.html', 
                           items=items, 
                           colleges=colleges, 
                           courses=courses, 
                           sessions=sessions,
                           pagination=pagination,
                           page_range=page_range)

@academics_bp.route('/batch_assignment', methods=['GET', 'POST'])
@permission_required('Batch Assignment')
def batch_assignment():
    if request.method == 'POST':
        student_ids = request.form.getlist('chk_student')
        # batch_id comes from the "Assign to Batch" select; assign_batch_id is the hidden fallback
        batch_id = request.form.get('batch_id') or request.form.get('assign_batch_id')
        type_tp = request.form.get('type_tp')
        if student_ids and batch_id and batch_id not in ('0', '') and type_tp:
            BatchModel.assign_batch(student_ids, batch_id, type_tp)
            flash('Batch assigned successfully.', 'success')
        elif student_ids and not batch_id:
            flash('Please select a batch to assign.', 'warning')
        return redirect(url_for('academics.batch_assignment', **request.args))

    filters = {
        'college_id': request.args.get('college_id'),
        'session_id': request.args.get('session_id'),
        'degree_id': request.args.get('degree_id'),
        'semester_id': request.args.get('semester_id'),
        'type_tp': request.args.get('type_tp', 'T'),
        'branch_id': request.args.get('branch_id'),
        'batch_id': request.args.get('batch_id')
    }

    students = []
    batches = []
    if all([filters['college_id'], filters['session_id'], filters['degree_id']]):
        students = BatchModel.get_students_for_batch(filters)
        if filters['semester_id']:
            batches = BatchModel.get_batches(filters['college_id'], filters['degree_id'], filters['semester_id'], filters['type_tp'])

    # Lookups
    loc_id = session.get('selected_loc')
    if loc_id:
        colleges = DB.fetch_all("SELECT pk_collegeid as id, collegename as name FROM SMS_College_Mst WHERE fk_locid = ? ORDER BY collegename", [loc_id])
    else:
        colleges = AcademicsModel.get_colleges_simple()

    lookups = {
        'colleges': colleges,
        'sessions': InfrastructureModel.get_sessions(),
        'degrees': AcademicsModel.get_college_all_degrees(filters['college_id']) if filters['college_id'] else [],
        'semesters': InfrastructureModel.get_all_semesters(),
        'branches': AcademicsModel.get_college_degree_specializations(filters['college_id'], filters['degree_id']) if (filters['college_id'] and filters['degree_id']) else [],
        'batches': batches
    }

    return render_template('academics/batch_assignment.html', lookups=lookups, filters=filters, students=students, now=datetime.now())

@academics_bp.route('/batch_assignment_report')
@permission_required('Batch Assignment')
def batch_assignment_report():
    filters = {
        'college_id': request.args.get('college_id'),
        'session_id': request.args.get('session_id'),
        'degree_id': request.args.get('degree_id'),
        'semester_id': request.args.get('semester_id'),
        'type_tp': request.args.get('type_tp', 'T'),
        'batch_id': request.args.get('batch_id'),
        'format': request.args.get('format', 'pdf')
    }
    
    if not all([filters['college_id'], filters['session_id'], filters['degree_id']]):
        flash('Please apply filters first.', 'warning')
        return redirect(url_for('academics.batch_assignment'))

    report_data = BatchModel.get_batch_report(filters)
    
    filters_info = {
        'college': DB.fetch_scalar("SELECT collegename FROM SMS_College_Mst WHERE pk_collegeid = ?", [filters['college_id']]),
        'session': DB.fetch_scalar("SELECT sessionname FROM SMS_AcademicSession_Mst WHERE pk_sessionid = ?", [filters['session_id']]),
        'degree': DB.fetch_scalar("SELECT degreename FROM SMS_Degree_Mst WHERE pk_degreeid = ?", [filters['degree_id']]),
        'type': 'Theory' if filters['type_tp'] == 'T' else 'Practical'
    }

    if filters['format'] == 'excel':
        html = render_template('reports/batch_assignment_report.html', report_data=report_data, filters=filters, now=datetime.now(), filters_info=filters_info)
        response = make_response(html)
        response.headers["Content-Disposition"] = "attachment; filename=Batch_Assignment_Report.xls"
        response.headers["Content-Type"] = "application/vnd.ms-excel"
        return response
    elif filters['format'] == 'word':
        html = render_template('reports/batch_assignment_report.html', report_data=report_data, filters=filters, now=datetime.now(), filters_info=filters_info)
        response = make_response(html)
        response.headers["Content-Disposition"] = "attachment; filename=Batch_Assignment_Report.doc"
        response.headers["Content-Type"] = "application/msword"
        return response
    else:
        from app.utils import generate_batch_assignment_pdf
        pdf_content = generate_batch_assignment_pdf(report_data, filters_info)
        return send_file(pdf_content, download_name="Batch_Assignment_Report.pdf", as_attachment=True, mimetype='application/pdf')

@academics_bp.route('/api/admission/get_batches')
def api_get_batches():
    college_id = request.args.get('college_id')
    degree_id = request.args.get('degree_id')
    semester_id = request.args.get('semester_id')
    type_tp = request.args.get('type_tp')
    
    if not all([college_id, degree_id, semester_id, type_tp]):
        return jsonify([])
        
    batches = BatchModel.get_batches(college_id, degree_id, semester_id, type_tp)
    return jsonify(batches)

@academics_bp.route('/student_course_approval_advisor', methods=['GET', 'POST'])
@permission_required('Course Allocation Approval For UG/PG[By Advisor]')
def student_course_approval_advisor():
    user_id = session.get('user_id')
    emp_id = session.get('emp_id')
    
    if request.method == 'POST':
        if AdvisorApprovalModel.save_approvals(request.form, emp_id, user_id):
            flash('Approvals updated successfully!', 'success')
        else:
            flash('Error updating approvals.', 'danger')
        return redirect(url_for('academics.student_course_approval_advisor', **request.args))

    filters = {
        'college_id': request.args.get('college_id', type=int),
        'session_id': request.args.get('session_id', type=int),
        'degree_id': request.args.get('degree_id', type=int),
        'semester_id': request.args.get('semester_id', type=int),
        'branch_id': request.args.get('branch_id', type=int),
        'exconfig_id': request.args.get('exconfig_id', type=int),
        'view': request.args.get('view')
    }

    students = []
    if filters['view'] == '1' and all([filters['college_id'], filters['session_id'], filters['degree_id'], filters['exconfig_id']]):
        students = AdvisorApprovalModel.get_students_for_approval(filters, emp_id)

    # Lookups
    loc_id = session.get('selected_loc')
    if loc_id:
        colleges = DB.fetch_all("SELECT pk_collegeid as id, collegename as name FROM SMS_College_Mst WHERE fk_locid = ? ORDER BY collegename", [loc_id])
    else:
        colleges = AcademicsModel.get_colleges_simple()

    lookups = {
        'colleges': colleges,
        'sessions': InfrastructureModel.get_sessions(),
        'degrees': AcademicsModel.get_college_all_degrees(filters['college_id']) if filters['college_id'] else [],
        'semesters': InfrastructureModel.get_all_semesters(),
        'branches': AcademicsModel.get_college_degree_specializations(filters['college_id'], filters['degree_id']) if (filters['college_id'] and filters['degree_id']) else [],
        'exam_configs': CourseAllocationModel.get_exam_configs(filters['degree_id'], filters['session_id'], filters['semester_id']) if filters['degree_id'] else [],
        'years': AcademicsModel.get_degree_years()
    }

    return render_template('academics/student_course_approval_advisor.html', 
                           lookups=lookups, filters=filters, students=clean_json_data(students))

@academics_bp.route('/course_allocation_approval', methods=['GET', 'POST'])
@permission_required('Course Allocation Approval For UG/PG[By Teacher]')
def course_allocation_approval():
    user_id = session.get('user_id')
    emp_id = session.get('emp_id')

    if request.method == 'POST':
        if TeacherApprovalModel.save_approvals(request.form, emp_id, user_id):
            flash('Approvals updated successfully!', 'success')
        else:
            flash('Error updating approvals.', 'danger')
        return redirect(url_for('academics.course_allocation_approval', **request.args))

    filters = {
        'college_id': request.args.get('college_id', type=int),
        'session_id': request.args.get('session_id', type=int),
        'degree_id': request.args.get('degree_id', type=int),
        'semester_id': request.args.get('semester_id', type=int),
        'branch_id': request.args.get('branch_id', type=int),
        'exconfig_id': request.args.get('exconfig_id', type=int),
        'view': request.args.get('view'),
    }

    students = []
    if filters['view'] == '1' and all([filters['college_id'], filters['session_id'], filters['degree_id'], filters['exconfig_id']]):
        students = TeacherApprovalModel.get_students_for_approval(filters, emp_id)

    # Lookups
    loc_id = session.get('selected_loc')
    if loc_id:
        colleges = DB.fetch_all(
            "SELECT pk_collegeid as id, collegename as name FROM SMS_College_Mst WHERE fk_locid = ? ORDER BY collegename",
            [loc_id]
        )
    else:
        colleges = AcademicsModel.get_colleges_simple()

    lookups = {
        'colleges': colleges,
        'sessions': InfrastructureModel.get_sessions(),
        'degrees': AcademicsModel.get_college_all_degrees(filters['college_id']) if filters['college_id'] else [],
        'semesters': InfrastructureModel.get_all_semesters(),
        'branches': AcademicsModel.get_college_degree_specializations(filters['college_id'], filters['degree_id']) if (filters['college_id'] and filters['degree_id']) else [],
        'exam_configs': CourseAllocationModel.get_exam_configs(filters['degree_id'], filters['session_id'], filters['semester_id']) if filters['degree_id'] else [],
        'years': AcademicsModel.get_degree_years(),
    }

    return render_template(
        'academics/student_course_approval_teacher.html',
        lookups=lookups,
        filters=filters,
        students=clean_json_data(students),
    )

@academics_bp.route('/student_course_approval_dsw', methods=['GET', 'POST'])
@permission_required('Course Allocation Approval For UG/PG[By DSW]')
def student_course_approval_dsw():
    user_id = session.get('user_id')
    emp_id = session.get('emp_id')

    if request.method == 'POST':
        if DswApprovalModel.save_approvals(request.form, emp_id, user_id):
            flash('Approvals updated successfully!', 'success')
        else:
            flash('Error updating approvals.', 'danger')
        return redirect(url_for('academics.student_course_approval_dsw', **request.args))

    filters = {
        'college_id': request.args.get('college_id', type=int),
        'session_id': request.args.get('session_id', type=int),
        'degree_id': request.args.get('degree_id', type=int),
        'semester_id': request.args.get('semester_id', type=int),
        'branch_id': request.args.get('branch_id', type=int),
        'exconfig_id': request.args.get('exconfig_id', type=int),
        'view': request.args.get('view'),
        'pending': request.args.get('pending'),
    }

    students = []
    if filters['view'] == '1' and all([filters['college_id'], filters['session_id'], filters['degree_id'], filters['semester_id'], filters['exconfig_id']]):
        students = DswApprovalModel.get_students_for_approval(filters, emp_id)

    loc_id = session.get('selected_loc')
    if loc_id:
        colleges = DB.fetch_all(
            "SELECT pk_collegeid as id, collegename as name FROM SMS_College_Mst WHERE fk_locid = ? ORDER BY collegename",
            [loc_id]
        )
    else:
        colleges = AcademicsModel.get_colleges_simple()

    lookups = {
        'colleges': colleges,
        'sessions': InfrastructureModel.get_sessions(),
        'degrees': AcademicsModel.get_college_all_degrees(filters['college_id']) if filters['college_id'] else [],
        'semesters': InfrastructureModel.get_all_semesters(),
        'branches': AcademicsModel.get_college_degree_specializations(filters['college_id'], filters['degree_id']) if (filters['college_id'] and filters['degree_id']) else [],
        'exam_configs': CourseAllocationModel.get_exam_configs(filters['degree_id'], filters['session_id'], filters['semester_id']) if filters['degree_id'] else [],
        'years': AcademicsModel.get_degree_years(),
    }

    return render_template(
        'academics/student_course_approval_dsw.html',
        lookups=lookups,
        filters=filters,
        students=clean_json_data(students),
    )

@academics_bp.route('/api/student_course_approval_dsw_pending')
@permission_required('Course Allocation Approval For UG/PG[By DSW]')
def student_course_approval_dsw_pending():
    emp_id = session.get('emp_id')
    filters = {
        'college_id': request.args.get('college_id', type=int),
        'session_id': request.args.get('session_id', type=int),
        'degree_id': request.args.get('degree_id', type=int),
        'semester_id': request.args.get('semester_id', type=int),
        'branch_id': request.args.get('branch_id', type=int),
        'exconfig_id': request.args.get('exconfig_id', type=int),
    }
    
    if not all([filters['college_id'], filters['session_id'], filters['degree_id'], filters['semester_id'], filters['exconfig_id']]):
        return jsonify([])
        
    students = DswApprovalModel.get_pending_students(filters, emp_id)
    return jsonify(clean_json_data(students))

@academics_bp.route('/student_course_approval_library', methods=['GET', 'POST'])
@permission_required('Course Allocation Approval For UG/PG[By Library]')
def student_course_approval_library():
    user_id = session.get('user_id')
    emp_id = session.get('emp_id')

    if request.method == 'POST':
        if LibraryApprovalModel.save_approvals(request.form, emp_id, user_id):
            flash('Approvals updated successfully!', 'success')
        else:
            flash('Error updating approvals.', 'danger')
        return redirect(url_for('academics.student_course_approval_library', **request.args))

    filters = {
        'college_id': request.args.get('college_id', type=int),
        'session_id': request.args.get('session_id', type=int),
        'degree_id': request.args.get('degree_id', type=int),
        'semester_id': request.args.get('semester_id', type=int),
        'branch_id': request.args.get('branch_id', type=int),
        'exconfig_id': request.args.get('exconfig_id', type=int),
        'view': request.args.get('view'),
    }

    students = []
    if filters['view'] == '1' and all([filters['college_id'], filters['session_id'], filters['degree_id'], filters['semester_id'], filters['exconfig_id']]):
        students = LibraryApprovalModel.get_students_for_approval(filters, emp_id)

    loc_id = session.get('selected_loc')
    if loc_id:
        colleges = DB.fetch_all(
            "SELECT pk_collegeid as id, collegename as name FROM SMS_College_Mst WHERE fk_locid = ? ORDER BY collegename",
            [loc_id]
        )
    else:
        colleges = AcademicsModel.get_colleges_simple()

    lookups = {
        'colleges': colleges,
        'sessions': InfrastructureModel.get_sessions(),
        'degrees': AcademicsModel.get_college_all_degrees(filters['college_id']) if filters['college_id'] else [],
        'semesters': InfrastructureModel.get_all_semesters(),
        'branches': AcademicsModel.get_college_degree_specializations(filters['college_id'], filters['degree_id']) if (filters['college_id'] and filters['degree_id']) else [],
        'exam_configs': CourseAllocationModel.get_exam_configs(filters['degree_id'], filters['session_id'], filters['semester_id']) if filters['degree_id'] else [],
        'years': AcademicsModel.get_degree_years(),
    }

    return render_template(
        'academics/student_course_approval_library.html',
        lookups=lookups,
        filters=filters,
        students=clean_json_data(students),
    )

@academics_bp.route('/api/student_course_approval_library_pending')
@permission_required('Course Allocation Approval For UG/PG[By Library]')
def student_course_approval_library_pending():
    emp_id = session.get('emp_id')
    filters = {
        'college_id': request.args.get('college_id', type=int),
        'session_id': request.args.get('session_id', type=int),
        'degree_id': request.args.get('degree_id', type=int),
        'semester_id': request.args.get('semester_id', type=int),
        'branch_id': request.args.get('branch_id', type=int),
        'exconfig_id': request.args.get('exconfig_id', type=int),
    }
    
    if not all([filters['college_id'], filters['session_id'], filters['degree_id'], filters['semester_id'], filters['exconfig_id']]):
        return jsonify([])
        
    students = LibraryApprovalModel.get_pending_students(filters, emp_id)
    return jsonify(clean_json_data(students))

@academics_bp.route('/student_course_approval_fee', methods=['GET', 'POST'])
@permission_required('Course Allocation Fee Approval For UG/PG')
def student_course_approval_fee():
    user_id = session.get('user_id')
    emp_id = session.get('emp_id')

    if request.method == 'POST':
        if FeeApprovalModel.save_approvals(request.form, emp_id, user_id):
            flash('Approvals updated successfully!', 'success')
        else:
            flash('Error updating approvals.', 'danger')
        return redirect(url_for('academics.student_course_approval_fee', **request.args))

    filters = {
        'college_id': request.args.get('college_id', type=int),
        'session_id': request.args.get('session_id', type=int),
        'degree_id': request.args.get('degree_id', type=int),
        'semester_id': request.args.get('semester_id', type=int),
        'branch_id': request.args.get('branch_id', type=int),
        'exconfig_id': request.args.get('exconfig_id', type=int),
        'view': request.args.get('view'),
    }

    students = []
    if filters['view'] == '1' and all([filters['college_id'], filters['session_id'], filters['degree_id'], filters['semester_id'], filters['exconfig_id']]):
        students = FeeApprovalModel.get_students_for_approval(filters, emp_id)

    loc_id = session.get('selected_loc')
    if loc_id:
        colleges = DB.fetch_all(
            "SELECT pk_collegeid as id, collegename as name FROM SMS_College_Mst WHERE fk_locid = ? ORDER BY collegename",
            [loc_id]
        )
    else:
        colleges = AcademicsModel.get_colleges_simple()

    lookups = {
        'colleges': colleges,
        'sessions': InfrastructureModel.get_sessions(),
        'degrees': AcademicsModel.get_college_all_degrees(filters['college_id']) if filters['college_id'] else [],
        'semesters': InfrastructureModel.get_all_semesters(),
        'branches': AcademicsModel.get_college_degree_specializations(filters['college_id'], filters['degree_id']) if (filters['college_id'] and filters['degree_id']) else [],
        'exam_configs': CourseAllocationModel.get_exam_configs(filters['degree_id'], filters['session_id'], filters['semester_id']) if filters['degree_id'] else [],
        'years': AcademicsModel.get_degree_years(),
    }

    return render_template(
        'academics/student_course_approval_fee.html',
        lookups=lookups,
        filters=filters,
        students=clean_json_data(students),
    )

@academics_bp.route('/api/student_course_approval_fee_pending')
@permission_required('Course Allocation Fee Approval For UG/PG')
def student_course_approval_fee_pending():
    emp_id = session.get('emp_id')
    filters = {
        'college_id': request.args.get('college_id', type=int),
        'session_id': request.args.get('session_id', type=int),
        'degree_id': request.args.get('degree_id', type=int),
        'semester_id': request.args.get('semester_id', type=int),
        'branch_id': request.args.get('branch_id', type=int),
        'exconfig_id': request.args.get('exconfig_id', type=int),
    }
    
    if not all([filters['college_id'], filters['session_id'], filters['degree_id'], filters['semester_id'], filters['exconfig_id']]):
        return jsonify([])
        
    students = FeeApprovalModel.get_pending_students(filters, emp_id)
    return jsonify(clean_json_data(students))

@academics_bp.route('/student_course_approval_dean', methods=['GET', 'POST'])
@permission_required('Course Allocation Approval For UG/PG [By Dean]')
def student_course_approval_dean():
    from app.models import DeanApprovalModel
    user_id = session.get('user_id')
    emp_id = session.get('emp_id')

    if request.method == 'POST':
        if DeanApprovalModel.save_approvals(request.form, emp_id, user_id):
            flash('Approvals updated successfully!', 'success')
        else:
            flash('Error updating approvals.', 'danger')
        return redirect(url_for('academics.student_course_approval_dean', **request.args))

    filters = {
        'college_id': request.args.get('college_id', type=int),
        'session_id': request.args.get('session_id', type=int),
        'degree_id': request.args.get('degree_id', type=int),
        'semester_id': request.args.get('semester_id', type=int),
        'branch_id': request.args.get('branch_id', type=int),
        'exconfig_id': request.args.get('exconfig_id', type=int),
        'view': request.args.get('view'),
        'pending': request.args.get('pending'),
    }

    students = []
    pending_students = []
    if filters['view'] == '1' and all([filters['college_id'], filters['session_id'], filters['degree_id'], filters['semester_id'], filters['exconfig_id']]):
        students = DeanApprovalModel.get_students_for_approval(filters, emp_id)
        if filters.get('pending') == '1':
            pending_students = DeanApprovalModel.get_pending_students(filters, emp_id)

    loc_id = session.get('selected_loc')
    if loc_id:
        colleges = DB.fetch_all(
            "SELECT pk_collegeid as id, collegename as name FROM SMS_College_Mst WHERE fk_locid = ? ORDER BY collegename",
            [loc_id]
        )
    else:
        colleges = AcademicsModel.get_colleges_simple()

    lookups = {
        'colleges': colleges,
        'sessions': InfrastructureModel.get_sessions(),
        'degrees': AcademicsModel.get_college_all_degrees(filters['college_id']) if filters['college_id'] else [],
        'semesters': InfrastructureModel.get_all_semesters(),
        'branches': AcademicsModel.get_college_degree_specializations(filters['college_id'], filters['degree_id']) if (filters['college_id'] and filters['degree_id']) else [],
        'exam_configs': CourseAllocationModel.get_exam_configs(filters['degree_id'], filters['session_id'], filters['semester_id']) if filters['degree_id'] else [],
        'years': AcademicsModel.get_degree_years(),
    }

    return render_template(
        'academics/student_course_approval_dean.html',
        lookups=lookups,
        filters=filters,
        students=clean_json_data(students),
        pending_students=clean_json_data(pending_students),
    )

@academics_bp.route('/student_course_approval_deanpgs', methods=['GET', 'POST'])
@permission_required('Course Allocation Approval For PG [By DeanPGS]')
def student_course_approval_deanpgs():
    from app.models import DeanPgsApprovalModel
    user_id = session.get('user_id')
    emp_id = session.get('emp_id')

    if request.method == 'POST':
        if DeanPgsApprovalModel.save_approvals(request.form, emp_id, user_id):
            flash('Approvals updated successfully!', 'success')
        else:
            flash('Error updating approvals.', 'danger')
        return redirect(url_for('academics.student_course_approval_deanpgs', **request.args))

    filters = {
        'college_id': request.args.get('college_id', type=int),
        'session_id': request.args.get('session_id', type=int),
        'degree_id': request.args.get('degree_id', type=int),
        'semester_id': request.args.get('semester_id', type=int),
        'branch_id': request.args.get('branch_id', type=int),
        'exconfig_id': request.args.get('exconfig_id', type=int),
        'view': request.args.get('view'),
        'pending': request.args.get('pending'),
    }

    students = []
    pending_students = []
    if filters['view'] == '1' and all([filters['college_id'], filters['session_id'], filters['degree_id'], filters['semester_id'], filters['exconfig_id']]):
        students = DeanPgsApprovalModel.get_students_for_approval(filters, emp_id)
        if filters.get('pending') == '1':
            pending_students = DeanPgsApprovalModel.get_pending_students(filters, emp_id)

    loc_id = session.get('selected_loc')
    if loc_id:
        colleges = DB.fetch_all(
            "SELECT pk_collegeid as id, collegename as name FROM SMS_College_Mst WHERE fk_locid = ? ORDER BY collegename",
            [loc_id]
        )
    else:
        colleges = AcademicsModel.get_colleges_simple()

    lookups = {
        'colleges': colleges,
        'sessions': InfrastructureModel.get_sessions(),
        'degrees': AcademicsModel.get_college_pg_degrees(filters['college_id']) if filters['college_id'] else [],
        'semesters': InfrastructureModel.get_all_semesters(),
        'branches': AcademicsModel.get_college_degree_specializations(filters['college_id'], filters['degree_id']) if (filters['college_id'] and filters['degree_id']) else [],
        'exam_configs': CourseAllocationModel.get_exam_configs(filters['degree_id'], filters['session_id'], filters['semester_id']) if filters['degree_id'] else [],
        'years': AcademicsModel.get_degree_years(),
    }

    return render_template(
        'academics/student_course_approval_deanpgs.html',
        lookups=lookups,
        filters=filters,
        students=clean_json_data(students),
        pending_students=clean_json_data(pending_students),
    )


@academics_bp.route('/degree_complete_detail', methods=['GET', 'POST'])
def degree_complete_detail():
    from datetime import datetime
    filters = {
        'college_id': request.args.get('college_id', ''),
        'session_id': request.args.get('session_id', ''),
        'degree_id': request.args.get('degree_id', ''),
        'semester_id': request.args.get('semester_id', ''),
        'completion_date': request.args.get('completion_date', datetime.now().strftime('%Y-%m-%d')),
        'from_date': request.args.get('from_date', datetime.now().strftime('%Y-%m-%d')),
        'to_date': request.args.get('to_date', datetime.now().strftime('%Y-%m-%d')),
        'view': request.args.get('view')
    }

    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'SAVE':
            student_ids = request.form.getlist('chk_student')
            comp_date = request.form.get('completion_date')
            if student_ids and comp_date:
                count = AdmissionModel.save_degree_complete(student_ids, comp_date, filters['college_id'], filters['degree_id'], filters['session_id'], filters['semester_id'], session.get('user_id', 'Admin'))
                flash(f'Degree complete updated for {count} students.', 'success')
            else:
                flash('Please select students and provide a date.', 'warning')
            return redirect(url_for('academics.degree_complete_detail', **request.args))

    students = []
    completed_students = []
    
    if filters['college_id'] and filters['session_id'] and filters['degree_id'] and filters['semester_id']:
        if filters['view'] == '1':
            students = AdmissionModel.get_students_for_degree_complete(filters, is_completed=False)
            completed_students = AdmissionModel.get_students_for_degree_complete(filters, is_completed=True)
            
    loc_id = session.get('selected_loc')
    if loc_id:
        colleges = DB.fetch_all("SELECT pk_collegeid as id, collegename as name FROM SMS_College_Mst WHERE fk_locid = ? ORDER BY collegename", [loc_id])
    else:
        colleges = AcademicsModel.get_colleges_simple()

    lookups = {
        'colleges': colleges,
        'sessions': InfrastructureModel.get_sessions(),
        'degrees': AcademicsModel.get_college_all_degrees(filters['college_id']) if filters['college_id'] else [],
        'semesters': DB.fetch_all("SELECT pk_semesterid, semester_roman FROM SMS_Semester_Mst ORDER BY pk_semesterid")
    }

    return render_template('academics/degree_complete_detail.html', lookups=lookups, filters=filters, students=students, completed_students=completed_students, now=datetime.now())

@academics_bp.route('/degree_complete_report')
def degree_complete_report():
    import io
    from datetime import datetime
    filters = {
        'college_id': request.args.get('college_id', ''),
        'degree_id': request.args.get('degree_id', ''),
        'from_date': request.args.get('from_date', datetime.now().strftime('%d/%m/%Y')),
        'to_date': request.args.get('to_date', datetime.now().strftime('%d/%m/%Y'))
    }
    
    if not all([filters['college_id'], filters['degree_id'], filters['from_date'], filters['to_date']]):
        flash('Please select all filters for the report.', 'warning')
        return redirect(url_for('academics.degree_complete_detail'))
        
    report_data = AdmissionModel.get_degree_complete_report(filters)
    
    if not report_data:
        flash('No Record Found!', 'warning')
        return redirect(url_for('academics.degree_complete_detail', **request.args))
        
    HINDI_MONTHS = {
        1: 'जनवरी', 2: 'फरवरी', 3: 'मार्च', 4: 'अप्रैल',
        5: 'मई',    6: 'जून',   7: 'जुलाई', 8: 'अगस्त',
        9: 'सितम्बर', 10: 'अक्तूबर', 11: 'नवम्बर', 12: 'दिसम्बर'
    }

    def get_division(cgpa):
        if cgpa is None or cgpa == '':
            return '', ''
        try:
            v = float(cgpa)
        except (ValueError, TypeError):
            return '', ''
        if v >= 7.5:
            return 'First Division with Distinction', 'प्रथम श्रेणी श्रेष्ठता के साथ'
        elif v >= 6.0:
            return 'First Division', 'प्रथम श्रेणी '
        elif v >= 5.0:
            return 'Second Division', 'द्वितीय श्रेणी'
        else:
            return 'Third Division', 'तृतीय श्रेणी'

    expected_cols = ['SrNo', 'DegreeNo', 'AdmissionNo', 'StudentName', 'StudentNameHindi', 'SonDaughter', 'HisHer', 'सुपुत्री सुपुत्र', 'FatherName', 'FatherNameHindi', 'MotherName', 'MotherNameHindi', 'Completion', 'Completion1', 'पूर्णता की तिथि', 'Discipline', 'DegreeNameHindi', 'MarksObtain', 'MaxMarks', 'Division', 'श्रेणी', 'AadharNumber', 'PhotoFileName', 'Percentages', 'MajorField', 'Branchname_hindi']

    import pandas as pd
    out_df = pd.DataFrame(columns=expected_cols)
    for i, row in enumerate(report_data):
        dt = row.get('dgcompleteddate', '')
        if hasattr(dt, 'strftime'):
            comp_date    = dt.strftime('%d/%m/%Y')
            completion1  = dt.strftime('%B, %Y')
            hindi_date   = f"{HINDI_MONTHS.get(dt.month, '')}, {dt.year}"
        else:
            comp_date   = str(dt) if dt else ''
            completion1 = comp_date
            hindi_date  = comp_date

        cgpa = row.get('final_cgpa', '')
        try:
            cgpa_val = round(float(cgpa), 2) if cgpa != '' else ''
        except (ValueError, TypeError):
            cgpa_val = ''
        pct = round(float(cgpa_val) * 10, 2) if cgpa_val != '' else ''
        div_en, div_hi = get_division(cgpa_val)

        out_df.loc[i] = {
            'SrNo':               i + 1,
            'DegreeNo':           '',
            'AdmissionNo':        row.get('AdmissionNo', ''),
            'StudentName':        row.get('fullname', ''),
            'StudentNameHindi':   row.get('fullname_h', '') or english_to_hindi(row.get('fullname', '')),
            'SonDaughter':        'Daughter' if row.get('gender') == 'F' else 'Son',
            'HisHer':             'her'       if row.get('gender') == 'F' else 'his',
            'सुपुत्री सुपुत्र':  'सुपुत्री'  if row.get('gender') == 'F' else 'सुपुत्र',
            'FatherName':         row.get('fname', ''),
            'FatherNameHindi':    row.get('Fnamehindi', '') or english_to_hindi(row.get('fname', '')),
            'MotherName':         row.get('mname', ''),
            'MotherNameHindi':    row.get('Mnamehindi', '') or english_to_hindi(row.get('mname', '')),
            'Completion':         comp_date,
            'Completion1':        completion1,
            'पूर्णता की तिथि':   hindi_date,
            'Discipline':         row.get('degreename', ''),
            'DegreeNameHindi':    row.get('degreename_hindi', '') or english_to_hindi(row.get('degreename', '')),
            'MarksObtain':        cgpa_val,
            'MaxMarks':           10 if cgpa_val != '' else '',
            'Division':           div_en,
            'श्रेणी':             div_hi,
            'AadharNumber':       row.get('AdharNo', ''),
            'PhotoFileName':      row.get('StuImage', ''),
            'Percentages':        pct,
            'MajorField':         row.get('Branchname', ''),
            'Branchname_hindi':   row.get('Branchname_hindi', '') or english_to_hindi(row.get('Branchname', '')),
        }
        
    # --- Professional Excel Report with openpyxl ---
    import openpyxl
    from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Degree Complete'

    # ── Color palette ──────────────────────────────────────────────
    CLR_HEADER_BG   = '1F3864'   # deep navy
    CLR_HEADER_FG   = 'FFFFFF'   # white
    CLR_TITLE_BG    = '2E75B6'   # medium blue
    CLR_TITLE_FG    = 'FFFFFF'
    CLR_SUBHDR_BG   = 'D6E4F0'   # pale blue
    CLR_SUBHDR_FG   = '1F3864'
    CLR_ROW_ODD     = 'FFFFFF'
    CLR_ROW_EVEN    = 'EBF5FB'
    CLR_FOOTER_BG   = 'F2F2F2'
    CLR_ACCENT      = 'F39C12'   # golden accent for distinction rows

    # ── Border helper ──────────────────────────────────────────────
    thin  = Side(style='thin',   color='AAAAAA')
    thick = Side(style='medium', color='1F3864')
    def mk_border(left=thin, right=thin, top=thin, bottom=thin):
        return Border(left=left, right=right, top=top, bottom=bottom)

    outer_border = Border(left=thick, right=thick, top=thick, bottom=thick)

    # ── Column definitions ─────────────────────────────────────────
    # (header label, field key in out_df, width, alignment)
    columns = [
        ('Sr.No',             'SrNo',              6,  'center'),
        ('Degree No',         'DegreeNo',           12, 'center'),
        ('Admission No',      'AdmissionNo',        16, 'center'),
        ('Student Name',      'StudentName',        22, 'left'),
        ('Name (Hindi)',       'StudentNameHindi',   22, 'left'),
        ('Son/Daughter',      'SonDaughter',        12, 'center'),
        ('His/Her',           'HisHer',              8, 'center'),
        ('सुपुत्री/सुपुत्र', 'सुपुत्री सुपुत्र',  14, 'center'),
        ('Father Name',       'FatherName',         22, 'left'),
        ('Father (Hindi)',    'FatherNameHindi',    22, 'left'),
        ('Mother Name',       'MotherName',         22, 'left'),
        ('Mother (Hindi)',    'MotherNameHindi',    22, 'left'),
        ('Completion Date',   'Completion',         16, 'center'),
        ('Completion1',       'Completion1',        16, 'center'),
        ('पूर्णता की तिथि',   'पूर्णता की तिथि',   18, 'center'),
        ('Discipline',        'Discipline',         22, 'left'),
        ('Degree (Hindi)',    'DegreeNameHindi',    22, 'left'),
        ('Marks Obtained',    'MarksObtain',        14, 'center'),
        ('Max Marks',         'MaxMarks',           10, 'center'),
        ('Division',          'Division',           28, 'left'),
        ('श्रेणी',            'श्रेणी',             24, 'left'),
        ('Aadhar Number',     'AadharNumber',       18, 'center'),
        ('Photo File',        'PhotoFileName',      26, 'left'),
        ('Percentage',        'Percentages',        12, 'center'),
        ('Major Field',       'MajorField',         20, 'left'),
        ('Branch (Hindi)',    'Branchname_hindi',   22, 'left'),
    ]
    num_cols = len(columns)

    # ── ROW 1 : Main Title (merged) ────────────────────────────────
    ws.row_dimensions[1].height = 36
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    c = ws.cell(1, 1,
        value='HARYANA AGRICULTURAL UNIVERSITY, HISAR')
    c.font      = Font(name='Calibri', bold=True, size=18, color=CLR_HEADER_FG)
    c.fill      = PatternFill('solid', fgColor=CLR_HEADER_BG)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border    = mk_border(left=thick, right=thick, top=thick, bottom=thin)

    # ── ROW 2 : Sub-Title ─────────────────────────────────────────
    ws.row_dimensions[2].height = 26
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
    c = ws.cell(2, 1, value='Student Degree Complete Report')
    c.font      = Font(name='Calibri', bold=True, size=14, color=CLR_TITLE_FG)
    c.fill      = PatternFill('solid', fgColor=CLR_TITLE_BG)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border    = mk_border(left=thick, right=thick, top=thin, bottom=thin)

    # ── ROW 3 : Date range info ────────────────────────────────────
    ws.row_dimensions[3].height = 18
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=num_cols)
    date_info = f"Period: {filters['from_date']}  to  {filters['to_date']}     |     Generated on: {datetime.now().strftime('%d %B %Y  %I:%M %p')}"
    c = ws.cell(3, 1, value=date_info)
    c.font      = Font(name='Calibri', italic=True, size=10, color='555555')
    c.fill      = PatternFill('solid', fgColor='EAF2FF')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border    = mk_border(left=thick, right=thick, top=thin, bottom=thick)

    # ── ROW 4 : Column Headers ────────────────────────────────────
    ws.row_dimensions[4].height = 30
    for col_idx, (label, _, width, _align) in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        c = ws.cell(4, col_idx, value=label)
        c.font      = Font(name='Calibri', bold=True, size=10, color=CLR_SUBHDR_FG)
        c.fill      = PatternFill('solid', fgColor=CLR_SUBHDR_BG)
        c.alignment = Alignment(horizontal='center', vertical='center',
                                wrap_text=True)
        lft = thick if col_idx == 1        else thin
        rgt = thick if col_idx == num_cols else thin
        c.border = Border(left=lft, right=rgt, top=thick, bottom=thick)

    # ── DATA ROWS ─────────────────────────────────────────────────
    for row_idx, (_, row_data) in enumerate(out_df.iterrows(), start=5):
        ws.row_dimensions[row_idx].height = 17
        is_even      = (row_idx % 2 == 0)
        is_distinct  = 'distinction' in str(row_data.get('Division', '')).lower()
        row_bg = 'FFF8E7' if is_distinct else (CLR_ROW_EVEN if is_even else CLR_ROW_ODD)

        for col_idx, (_, field, _, align) in enumerate(columns, start=1):
            val = row_data.get(field, '')
            c   = ws.cell(row_idx, col_idx, value=val)
            c.font      = Font(name='Calibri', size=9,
                               bold=(col_idx == 1),
                               color='1A1A1A')
            c.fill      = PatternFill('solid', fgColor=row_bg)
            c.alignment = Alignment(horizontal=align, vertical='center',
                                    wrap_text=False)
            lft = thick if col_idx == 1        else thin
            rgt = thick if col_idx == num_cols else thin
            c.border = Border(left=lft, right=rgt, top=thin, bottom=thin)

    # ── FOOTER ROW ────────────────────────────────────────────────
    last_row = 5 + len(out_df)
    ws.row_dimensions[last_row].height = 18
    ws.merge_cells(start_row=last_row, start_column=1,
                   end_row=last_row, end_column=num_cols)
    c = ws.cell(last_row, 1,
        value=f'Total Records: {len(out_df)}     |     HAU ERP System  —  Confidential')
    c.font      = Font(name='Calibri', bold=True, size=9, color='555555')
    c.fill      = PatternFill('solid', fgColor=CLR_FOOTER_BG)
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.border    = mk_border(left=thick, right=thick, top=thick, bottom=thick)

    # ── Freeze panes below header row ────────────────────────────
    ws.freeze_panes = 'A5'

    # ── Auto-filter on header row ─────────────────────────────────
    ws.auto_filter.ref = (
        f'A4:{get_column_letter(num_cols)}{4 + len(out_df)}'
    )

    # ── Print settings ────────────────────────────────────────────
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth  = 1
    ws.print_title_rows       = '1:4'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name="DegreeComplete Report (5).xlsx", as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@academics_bp.route('/api/get_college_degrees')
def get_college_degrees_query_api():
    college_id = request.args.get('college_id')
    type_tp = request.args.get('type_tp')
    if not college_id:
        return jsonify([])
    if type_tp == 'U':
        degrees = AcademicsModel.get_college_ug_degrees(college_id)
    else:
        degrees = AcademicsModel.get_college_degrees(college_id)
    return jsonify(clean_json_data(degrees))

@academics_bp.route('/api/get_degree_branches')
def get_degree_branches_query_api():
    college_id = request.args.get('college_id')
    degree_id = request.args.get('degree_id')
    if not degree_id:
        return jsonify([])
    branches = AcademicsModel.get_degree_branches(degree_id)
    return jsonify(clean_json_data(branches))

@academics_bp.route('/api/get_teachers')
def get_teachers_api():
    college_id = request.args.get('college_id')
    if not college_id:
        return jsonify([])
    teachers = AdvisorAllocationModel.get_teachers_for_dropdown(college_id)
    return jsonify(clean_json_data(teachers))

@academics_bp.route('/api/employee_mapped_degrees')
def get_employee_mapped_degrees():
    """Return degree IDs already mapped to a given user."""
    user_id = request.args.get('user_id')
    if not user_id:
        return jsonify([])
    rows = DB.fetch_all("""
        SELECT FK_DegreeID AS degree_id
        FROM SMS_EmployeeDegreeMap
        WHERE FK_USERID = ?
    """, [user_id])
    return jsonify([r['degree_id'] for r in rows])
